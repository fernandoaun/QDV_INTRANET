from __future__ import annotations

import sqlite3
from datetime import date, datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from qdv_salmuera.data.db import DB
from qdv_salmuera.config.settings import SECURITY_DELETE_CODE
from qdv_salmuera.utils.validators import validate_float, validate_int, is_int_ok, is_float_ok, fmt_num
from qdv_salmuera.utils.dates import iso_to_ddmmyyyy, parse_date_ddmmyyyy, date_to_iso
from qdv_salmuera.ui.widgets import ScrollableFrame
from qdv_salmuera.ui.dialogs import HistorialDialog, CodigoSeguridadDialog
from qdv_salmuera.ui.produccion_ui_helpers import add_operador_via_dialog, bind_excel_field_navigation
from qdv_salmuera.ui.theme import QDV_COLORS
from qdv_salmuera.ui.module_timer import PersistentModuleTimer, TimerConfig
from qdv_salmuera.ui.module_labels import module_label
from qdv_salmuera.utils.module_defaults import build_daily_lot, get_current_username


class ReactorWindow(tk.Toplevel):
    def __init__(self, master, db: DB):
        super().__init__(master)
        self.db = db

        self.title(f"Química del Valle - Producción - {module_label('reactor')}")
        self.geometry("1200x800")
        self.minsize(860, 600)

        now = datetime.now()
        self.fixed_date_ddmmyyyy = now.strftime("%d/%m/%Y")
        self.fixed_date_iso = now.strftime("%Y-%m-%d")

        self.var_fecha = tk.StringVar(value=self.fixed_date_ddmmyyyy)
        self.var_hora = tk.StringVar(value=now.strftime("%H:%M"))

        self.operadores = self.db.fetch_operadores()
        self.var_operador = tk.StringVar()
        self.cbo_operador: Optional[ttk.Combobox] = None

        self.var_lote = tk.StringVar()
        self.var_ph = tk.StringVar()
        self.var_temperatura = tk.StringVar()
        self.var_densidad = tk.StringVar()
        self.var_concentracion_tabla = tk.StringVar()
        self.var_exceso_naoh = tk.StringVar()
        self.var_exceso_na2co3 = tk.StringVar()

        self.txt_obs: Optional[tk.Text] = None

        self.vcmd_float = (self.register(validate_float), "%P")
        self.vcmd_int = (self.register(validate_int), "%P")

        self.btn_guardar: Optional[ttk.Button] = None
        self.btn_historial: Optional[ttk.Button] = None
        self.btn_borrar: Optional[ttk.Button] = None
        self.lbl_status: Optional[ttk.Label] = None

        self.tree: Optional[ttk.Treeview] = None
        self.tree_cols: List[str] = []
        self._selected_row_id: Optional[int] = None

        # Timer Reactor (2h)
        self.var_timer = tk.StringVar(value="02:00:00")
        self.lbl_timer = None
        self._timer = None

        self._build_ui()
        self._set_default_operador_lote()
        self._wire_validation()
        self._refresh_save_state()
        self._load_day_table()

        self._init_timer()

    def _clear_row(self) -> None:
        """Limpia solo los campos editables (fecha/hora quedan automáticas)."""
        self.var_ph.set("")
        self.var_temperatura.set("")
        self.var_densidad.set("")
        self.var_concentracion_tabla.set("")
        self.var_exceso_naoh.set("")
        self.var_exceso_na2co3.set("")
        if self.txt_obs is not None:
            self.txt_obs.delete("1.0", "end")
        self._set_default_operador_lote()

    def _set_default_operador_lote(self) -> None:
        username = get_current_username(self).strip()
        if username:
            if username not in self.operadores:
                try:
                    self.db.add_operador(username)
                except sqlite3.IntegrityError:
                    pass
                self.operadores = self.db.fetch_operadores()
                if self.cbo_operador is not None:
                    self.cbo_operador["values"] = self.operadores
            self.var_operador.set(username)
        next_correlative = self.db.get_daily_sample_count("reactor", self.fixed_date_iso) + 1
        self.var_lote.set(build_daily_lot(next_correlative, self.fixed_date_iso))

    def _init_timer(self) -> None:
        def _fetch_last_iso() -> Optional[str]:
            last = self.db.fetch_last_reactor()
            return (last or {}).get("created_at_iso")

        def _on_overdue(overdue: bool) -> None:
            if self.lbl_timer is not None:
                self.lbl_timer.configure(foreground=("red" if overdue else QDV_COLORS["fg"]))

        self._timer = PersistentModuleTimer(
            owner=self,
            config=TimerConfig(interval_seconds=2 * 60 * 60, default_label="02:00:00"),
            fetch_last_created_at_iso=_fetch_last_iso,
            out_var=self.var_timer,
            on_overdue_change=_on_overdue,
        )
        self._timer.start()

    # =========================
    # UI
    # =========================
    def _build_ui(self) -> None:
        outer = ttk.Frame(self, padding=16)
        outer.pack(fill="both", expand=True)

        # Advertencias no bloqueantes (estilo centralizado en ui/theme.py)
        self._warn_entry_style = "QDV.Warn.TEntry"

        header = ttk.Frame(outer)
        header.pack(fill="x", pady=(0, 8))
        ttk.Label(header, text=module_label("reactor"), font=("Segoe UI", 18, "bold")).pack(side="left")
        ttk.Label(header, text=f"Día: {self.fixed_date_ddmmyyyy}", font=("Segoe UI", 11)).pack(side="right")

        # Cronómetro
        timer_box = ttk.LabelFrame(outer, text="Cronómetro de análisis (cada 2 h)", padding=8)
        timer_box.pack(fill="x", pady=(0, 10))
        trow = ttk.Frame(timer_box)
        trow.pack(fill="x")
        ttk.Label(trow, text="Tiempo restante:").pack(side="left")
        self.lbl_timer = ttk.Label(trow, textvariable=self.var_timer, font=("Segoe UI", 14, "bold"))
        self.lbl_timer.pack(side="left", padx=(8, 0))

        sc = ScrollableFrame(outer)
        sc.pack(fill="both", expand=True)
        body = sc.inner

        # Form (una fila tipo Excel)
        reg = ttk.LabelFrame(body, text=f"Carga de {module_label('reactor')}", padding=12)
        reg.pack(fill="x", pady=(0, 12))

        # Columnas (UNA sola fuente de verdad: widths en caracteres)
        cols: List[Tuple[str, str, int]] = [
            ("fecha", "Fecha", 10),
            ("hora", "Hora", 8),
            ("operador", "Operador", 15),
            ("op_add", "", 3),
            ("lote", "Lote", 12),
            ("ph", "pH", 6),
            ("temperatura", "Temp", 8),
            ("densidad", "Densidad", 8),
            ("conc", "Concentración (min. 300 gr/lts)", 24),
            ("naoh", "Ex NaOH (max. 0.16 gr/lts)", 22),
            ("na2co3", "Ex Na2 CO3 (max. 0.45 gr/lts)", 24),
            ("obs", "Observaciones", 30),
        ]

        # Planilla con scroll horizontal (encabezados + fila de carga)
        sheet_wrap = ttk.Frame(reg)
        sheet_wrap.pack(fill="x")

        sheet_canvas = tk.Canvas(
            sheet_wrap,
            background=QDV_COLORS.get("card_elevated", QDV_COLORS["card"]),
            highlightthickness=0,
            bd=0,
        )
        sheet_hsb = ttk.Scrollbar(sheet_wrap, orient="horizontal", command=sheet_canvas.xview)
        sheet_canvas.configure(xscrollcommand=sheet_hsb.set)

        sheet_canvas.pack(side="top", fill="x", expand=False)
        sheet_hsb.pack(side="bottom", fill="x")

        sheet = ttk.Frame(sheet_canvas)
        _sheet_win = sheet_canvas.create_window((0, 0), window=sheet, anchor="nw")

        def _sync_sheet_scroll(_event=None) -> None:
            sheet_canvas.configure(scrollregion=sheet_canvas.bbox("all"))

        def _sync_sheet_width(event) -> None:
            # Mantener altura y permitir ancho > ventana (scroll horizontal)
            sheet_canvas.itemconfigure(_sheet_win, height=event.height)

        sheet.bind("<Configure>", _sync_sheet_scroll)
        sheet_canvas.bind("<Configure>", _sync_sheet_width)

        # Encabezados (mismo width que inputs)
        for i, (_k, title, wch) in enumerate(cols):
            ttk.Label(
                sheet,
                text=title,
                width=wch,
                anchor="center",
                font=("Segoe UI", 9, "bold"),
                foreground="#334155",
            ).grid(row=0, column=i, padx=(2, 2), pady=(0, 6), sticky="w")

        # Inputs (fila)
        ttk.Entry(sheet, textvariable=self.var_fecha, state="readonly", width=cols[0][2]).grid(row=1, column=0, padx=(2, 2), sticky="w")
        ttk.Entry(sheet, textvariable=self.var_hora, state="readonly", width=cols[1][2]).grid(row=1, column=1, padx=(2, 2), sticky="w")

        self.cbo_operador = ttk.Combobox(
            sheet,
            textvariable=self.var_operador,
            values=self.operadores,
            state="readonly",
            width=cols[2][2],
        )
        self.cbo_operador.grid(row=1, column=2, padx=(2, 2), sticky="w")
        ttk.Button(sheet, text="+", width=cols[3][2], command=self._add_operador).grid(row=1, column=3, padx=(2, 2), sticky="w")

        self.ent_lote = ttk.Entry(sheet, textvariable=self.var_lote, width=cols[4][2], state="readonly")
        self.ent_lote.grid(row=1, column=4, padx=(2, 2), sticky="w")

        self.ent_ph = ttk.Entry(sheet, textvariable=self.var_ph, validate="key", validatecommand=self.vcmd_float, width=cols[5][2])
        self.ent_ph.grid(row=1, column=5, padx=(2, 2), sticky="w")

        self.ent_temperatura = ttk.Entry(
            sheet,
            textvariable=self.var_temperatura,
            validate="key",
            validatecommand=self.vcmd_float,
            width=cols[6][2],
        )
        self.ent_temperatura.grid(row=1, column=6, padx=(2, 2), sticky="w")

        self.ent_densidad = ttk.Entry(
            sheet,
            textvariable=self.var_densidad,
            validate="key",
            validatecommand=self.vcmd_float,
            width=cols[7][2],
        )
        self.ent_densidad.grid(row=1, column=7, padx=(2, 2), sticky="w")

        self.ent_conc = ttk.Entry(
            sheet,
            textvariable=self.var_concentracion_tabla,
            validate="key",
            validatecommand=self.vcmd_float,
            width=cols[8][2],
        )
        self.ent_conc.grid(row=1, column=8, padx=(2, 2), sticky="w")

        self.ent_naoh = ttk.Entry(
            sheet,
            textvariable=self.var_exceso_naoh,
            validate="key",
            validatecommand=self.vcmd_float,
            width=cols[9][2],
        )
        self.ent_naoh.grid(row=1, column=9, padx=(2, 2), sticky="w")

        self.ent_na2co3 = ttk.Entry(
            sheet,
            textvariable=self.var_exceso_na2co3,
            validate="key",
            validatecommand=self.vcmd_float,
            width=cols[10][2],
        )
        self.ent_na2co3.grid(row=1, column=10, padx=(2, 2), sticky="w")

        self.txt_obs = tk.Text(sheet, height=2, width=cols[11][2], wrap="word")
        self.txt_obs.configure(
            bg=QDV_COLORS.get("input_inset", QDV_COLORS["input_bg"]),
            fg=QDV_COLORS["fg"],
            insertbackground=QDV_COLORS["accent"],
            relief="flat",
            highlightthickness=0,
            padx=2,
            pady=1,
        )
        self.txt_obs.grid(row=1, column=11, padx=(2, 2), sticky="w")

        # Navegación por teclado tipo Excel
        focus_order = [
            self.cbo_operador,
            self.ent_lote,
            self.ent_ph,
            self.ent_temperatura,
            self.ent_densidad,
            self.ent_conc,
            self.ent_naoh,
            self.ent_na2co3,
            self.txt_obs,
        ]
        bind_excel_field_navigation(focus_order, self.on_save)

        # Acciones (guardar/historial/cerrar)
        act = ttk.Frame(body)
        act.pack(fill="x", pady=(0, 10))
        self.lbl_status = ttk.Label(act, text="", font=("Segoe UI", 10))
        self.lbl_status.pack(side="left", fill="x", expand=True)

        self.btn_guardar = ttk.Button(act, text="Guardar", command=self.on_save)
        self.btn_guardar.pack(side="right", padx=(8, 0))

        self.btn_historial = ttk.Button(act, text="Historial", command=self.on_historial)
        self.btn_historial.pack(side="right", padx=(8, 0))

        ttk.Button(act, text="Cerrar", command=self.destroy).pack(side="right")

        # Foco inicial
        self.after(50, lambda: self.cbo_operador.focus_set())

        # Tabla inferior
        bottom = ttk.LabelFrame(body, text="Registros del día - Eliminar seleccionado", padding=12)
        bottom.pack(fill="both", expand=True)

        toolbar = ttk.Frame(bottom)
        toolbar.pack(fill="x", pady=(0, 8))
        ttk.Label(toolbar, text="Seleccione una fila para borrar.", font=("Segoe UI", 9), foreground=QDV_COLORS["muted"]).pack(side="left")
        self.btn_borrar = ttk.Button(toolbar, text="Borrar seleccionado", command=self.on_delete_selected)
        self.btn_borrar.pack(side="right")

        tv_wrap = ttk.Frame(bottom)
        tv_wrap.pack(fill="both", expand=True)

        self.tree_cols = [
            "id",
            "fecha",
            "hora",
            "operador",
            "lote",
            "ph",
            "temperatura",
            "densidad",
            "conc_tabla",
            "naoh",
            "na2co3",
            "observaciones",
            "creado",
        ]

        self.tree = ttk.Treeview(tv_wrap, columns=self.tree_cols, show="headings", style="Treeview")
        vsb = ttk.Scrollbar(tv_wrap, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(bottom, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        # Headings
        widths = {
            "id": 60,
            "fecha": 90,
            "hora": 70,
            "operador": 130,
            "lote": 120,
            "ph": 70,
            "temperatura": 100,
            "densidad": 90,
            "conc_tabla": 130,
            "naoh": 120,
            "na2co3": 130,
            "observaciones": 260,
            "creado": 160,
        }

        for col_key in self.tree_cols:
            title = col_key
            anchor = "center"
            if col_key in ("operador", "lote", "observaciones", "creado"):
                anchor = "w"
            if col_key == "ph":
                title = "pH"
            if col_key == "conc_tabla":
                title = "Conc. tabla"
            if col_key == "naoh":
                title = "Exceso NaOH"
            if col_key == "na2co3":
                title = "Exceso Na2CO3"
            if col_key == "temperatura":
                title = "Temperatura"
            if col_key == "densidad":
                title = "Densidad"

            self.tree.heading(col_key, text=title)
            self.tree.column(col_key, width=widths.get(col_key, 120), anchor=anchor, stretch=False)

        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)

    # =========================
    # Validación / estados
    # =========================
    def _wire_validation(self) -> None:
        vars_to_trace = [
            self.var_operador,
            self.var_lote,
            self.var_ph,
            self.var_temperatura,
            self.var_densidad,
            self.var_concentracion_tabla,
            self.var_exceso_naoh,
            self.var_exceso_na2co3,
        ]
        for v in vars_to_trace:
            v.trace_add("write", lambda *_: self._refresh_save_state())

        if self.txt_obs is not None:
            self.txt_obs.bind("<KeyRelease>", lambda _e: self._refresh_save_state())

    def _add_operador(self) -> None:
        name = add_operador_via_dialog(self, self.db)
        if name is None:
            return
        self.operadores = self.db.fetch_operadores()
        if self.cbo_operador is not None:
            self.cbo_operador["values"] = self.operadores
        self.var_operador.set(name)

    def _validate_all(self) -> Tuple[bool, str]:
        op = (self.var_operador.get() or "").strip()
        if not op:
            return False, "Operador es obligatorio."
        if op not in self.operadores:
            return False, "Operador inválido. Seleccione uno del desplegable o agregue uno nuevo con +."

        lote = (self.var_lote.get() or "").strip()
        if not lote:
            return False, "Lote es obligatorio."

        def _req_float(var: tk.StringVar, label: str) -> Tuple[bool, float | None, str]:
            s = (var.get() or "").strip()
            if not is_float_ok(s):
                return False, None, f"{label} es obligatorio y debe ser numérico."
            try:
                return True, float(s.replace(",", ".")), "OK"
            except Exception:
                return False, None, f"{label} es obligatorio y debe ser numérico (usar punto)."

        ok, ph, msg = _req_float(self.var_ph, "pH")
        if not ok:
            return False, msg
        ok, temp, msg = _req_float(self.var_temperatura, "Temperatura")
        if not ok:
            return False, msg
        ok, dens, msg = _req_float(self.var_densidad, "Densidad")
        if not ok:
            return False, msg
        ok, conc, msg = _req_float(self.var_concentracion_tabla, "Concentración de tabla")
        if not ok:
            return False, msg
        ok, naoh, msg = _req_float(self.var_exceso_naoh, "Exceso NaOH")
        if not ok:
            return False, msg
        ok, na2co3, msg = _req_float(self.var_exceso_na2co3, "Exceso Na2CO3")
        if not ok:
            return False, msg

        return True, "OK"

    def _apply_recommended_warnings(self) -> List[str]:
        """
        Advertencias NO bloqueantes basadas en rangos recomendados.
        Devuelve lista de mensajes (si hay).
        """
        msgs: List[str] = []

        def _warn_entry(entry: Optional[ttk.Entry], warn: bool) -> None:
            if entry is None:
                return
            try:
                entry.configure(style=(self._warn_entry_style if warn else "TEntry"))
            except Exception:
                pass

        # Ex NaOH (referencia informativa)
        naoh_s = (self.var_exceso_naoh.get() or "").strip()
        naoh_warn = False
        if is_float_ok(naoh_s):
            try:
                naoh = float(naoh_s.replace(",", "."))
                naoh_warn = naoh > 0.16
            except Exception:
                naoh_warn = False
        _warn_entry(getattr(self, "ent_naoh", None), naoh_warn)
        if naoh_warn:
            msgs.append("Ex NaOH por encima del rango recomendado (max. 0.16).")

        # Concentración (referencia operativa: mínimo recomendado)
        conc_s = (self.var_concentracion_tabla.get() or "").strip()
        conc_warn = False
        if is_float_ok(conc_s):
            try:
                conc = float(conc_s.replace(",", "."))
                conc_warn = conc < 300
            except Exception:
                conc_warn = False
        _warn_entry(getattr(self, "ent_conc", None), conc_warn)
        if conc_warn:
            msgs.append("Concentración por debajo del rango recomendado (min. 300 gr/lts).")

        # Ex Na2CO3 (referencia informativa)
        na2_s = (self.var_exceso_na2co3.get() or "").strip()
        na2_warn = False
        if is_float_ok(na2_s):
            try:
                na2 = float(na2_s.replace(",", "."))
                na2_warn = na2 > 0.45
            except Exception:
                na2_warn = False
        _warn_entry(getattr(self, "ent_na2co3", None), na2_warn)
        if na2_warn:
            msgs.append("Ex Na2 CO3 por encima del rango recomendado (max. 0.45).")

        return msgs

    def _refresh_save_state(self) -> None:
        if self.btn_guardar is None or self.lbl_status is None:
            return

        ok, msg = self._validate_all()
        if ok:
            warn_msgs = self._apply_recommended_warnings()
            if warn_msgs:
                self.lbl_status.config(
                    text=" | ".join(warn_msgs),
                    foreground=QDV_COLORS.get("warning", "#F59E0B"),
                    font=("Segoe UI", 10, "bold"),
                )
            else:
                self.lbl_status.config(text="")
        else:
            self.lbl_status.config(text=msg, foreground="red", font=("Segoe UI", 12, "bold"))

        self.btn_guardar.configure(state=("normal" if ok else "disabled"))

    # =========================
    # CRUD / Tabla
    # =========================
    def _load_day_table(self) -> None:
        if self.tree is None:
            return

        registros = self.db.fetch_reactor_by_date(self.fixed_date_iso)

        for item in self.tree.get_children():
            self.tree.delete(item)

        for r in registros:
            row = [
                r["id"],
                iso_to_ddmmyyyy(r["fecha_iso"]),
                r["hora_hm"],
                r["operador"],
                r["lote"],
                fmt_num(r["ph"]),
                fmt_num(r["temperatura"]),
                fmt_num(r["densidad"]),
                fmt_num(r["concentracion_tabla"]),
                fmt_num(r["exceso_naoh"]),
                fmt_num(r["exceso_na2co3"]),
                (r.get("observaciones", "") or ""),
                r["created_at_iso"],
            ]
            self.tree.insert("", "end", values=row)

        self._selected_row_id = None

    def _on_tree_select(self, _event) -> None:
        if self.tree is None:
            return
        sel = self.tree.selection()
        if not sel:
            self._selected_row_id = None
            return
        vals = self.tree.item(sel[0], "values")
        if not vals:
            self._selected_row_id = None
            return
        try:
            self._selected_row_id = int(vals[0])
        except Exception:
            self._selected_row_id = None

    def on_delete_selected(self) -> None:
        if self._selected_row_id is None:
            messagebox.showwarning("Atención", "Seleccione un registro en la tabla para borrarlo.")
            return

        dlg = CodigoSeguridadDialog(self, title="Borrar registro - Código de seguridad")
        self.wait_window(dlg)
        if dlg.result is None:
            return
        if dlg.result != SECURITY_DELETE_CODE:
            messagebox.showerror("Código incorrecto", "Código de seguridad inválido. No se borró el registro.")
            return

        if not messagebox.askyesno("Confirmar", f"¿Confirma borrar el registro ID {self._selected_row_id}?"):
            return

        try:
            self.db.delete_reactor_by_id(self._selected_row_id)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo borrar el registro.\n\nDetalle: {e}")
            return

        self._load_day_table()
        messagebox.showinfo("OK", f"Registro ID {self._selected_row_id} eliminado.")

    # =========================
    # Guardar
    # =========================
    def on_save(self) -> None:
        try:
            ok, msg = self._validate_all()
            if not ok:
                messagebox.showerror("No se puede guardar", msg)
                return

            # Mantener coherencia con el resto del sistema: hora/created_at según momento real de guardado
            now = datetime.now()
            self.var_hora.set(now.strftime("%H:%M"))

            data = {
                "fecha_iso": self.fixed_date_iso,
                "hora_hm": now.strftime("%H:%M"),
                "operador": self.var_operador.get().strip(),
                "lote": self.var_lote.get().strip(),
                "ph": float(self.var_ph.get().strip().replace(",", ".")),
                "temperatura": float(self.var_temperatura.get().strip().replace(",", ".")),
                "densidad": float(self.var_densidad.get().strip().replace(",", ".")),
                "concentracion_tabla": float(self.var_concentracion_tabla.get().strip().replace(",", ".")),
                "exceso_naoh": float(self.var_exceso_naoh.get().strip().replace(",", ".")),
                "exceso_na2co3": float(self.var_exceso_na2co3.get().strip().replace(",", ".")),
                "observaciones": self.txt_obs.get("1.0", "end").strip() if self.txt_obs is not None else "",
                "created_at_iso": now.isoformat(timespec="seconds"),
            }

            self.db.insert_reactor(data)
            self._load_day_table()
            self._clear_row()
            self._refresh_save_state()
            self.after(50, lambda: self.cbo_operador.focus_set())
            if self._timer is not None:
                self._timer.reset_from_created_at_iso(data["created_at_iso"])
            # Sin popup de éxito: carga rápida continua.
        except Exception as e:
            messagebox.showerror("Error al guardar", f"{e}")

    def destroy(self):
        try:
            if self._timer is not None:
                self._timer.stop()
        except Exception:
            pass
        super().destroy()

    # =========================
    # Historial / Export
    # =========================
    def on_historial(self) -> None:
        if not OPENPYXL_OK:
            messagebox.showerror("Falta dependencia", "Para exportar a Excel instalá:\n\npip install openpyxl")
            return

        hoy_ddmmyyyy = datetime.now().strftime("%d/%m/%Y")
        dlg = HistorialDialog(self, default_desde=hoy_ddmmyyyy, default_hasta=hoy_ddmmyyyy)
        self.wait_window(dlg)
        if not dlg.result:
            return

        desde_s, hasta_s = dlg.result
        d1 = parse_date_ddmmyyyy(desde_s)
        d2 = parse_date_ddmmyyyy(hasta_s)
        if not d1 or not d2:
            messagebox.showerror("Error", "Rango de fechas inválido.")
            return

        desde_iso = date_to_iso(d1)
        hasta_iso = date_to_iso(d2)

        try:
            registros = self.db.fetch_reactor_range(desde_iso, hasta_iso)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el historial.\n\nDetalle: {e}")
            return

        if not registros:
            messagebox.showinfo("Sin datos", "No hay registros en el rango seleccionado.")
            return

        default_name = f"historial_reactor_{desde_iso}_a_{hasta_iso}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Guardar reporte Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not save_path:
            return

        try:
            self._export_excel(registros, save_path)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el Excel.\n\nDetalle: {e}")
            return

        messagebox.showinfo("OK", f"Reporte generado:\n{save_path}")

    def _export_excel(self, registros: List[Dict[str, Any]], path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Reactor"

        headers = [
            "ID",
            "Fecha",
            "Hora",
            "Operador",
            "Lote",
            "pH",
            "Temperatura",
            "Densidad",
            "Conc. tabla",
            "Exceso NaOH",
            "Exceso Na2CO3",
            "Observaciones",
            "Creado (ISO)",
        ]
        ws.append(headers)

        for r in registros:
            ws.append([
                r["id"],
                iso_to_ddmmyyyy(r["fecha_iso"]),
                r["hora_hm"],
                r["operador"],
                r["lote"],
                r["ph"],
                r["temperatura"],
                r["densidad"],
                r["concentracion_tabla"],
                r["exceso_naoh"],
                r["exceso_na2co3"],
                r.get("observaciones", ""),
                r["created_at_iso"],
            ])

        # Ajustar ancho de columnas (similar al módulo salmuera)
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter]:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)

        wb.save(path)

