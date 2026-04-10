from __future__ import annotations

import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from qdv_salmuera.data.db import DB

from qdv_salmuera.config.settings import (
    DEFAULT_OPERATORS,
    SECURITY_DELETE_CODE,
    VOLTAGE_MIN,
    VOLTAGE_MAX,
    ANALYSIS_INTERVAL_SECONDS,
)

from qdv_salmuera.utils.validators import (
    validate_float,
    validate_int,
    is_int_ok,
    is_float_ok,
    fmt_num,
)

from qdv_salmuera.utils.dates import (
    iso_to_ddmmyyyy,
    parse_date_ddmmyyyy,
    date_to_iso,
    format_hhmmss,
)

from qdv_salmuera.ui.widgets import ScrollableFrame
from qdv_salmuera.ui.dialogs import HistorialDialog, CodigoSeguridadDialog
from qdv_salmuera.ui.produccion_ui_helpers import add_operador_via_dialog
from qdv_salmuera.ui.salmuera_edit import EditRegistroDialog
from qdv_salmuera.ui.theme import QDV_COLORS
from qdv_salmuera.ui.module_labels import module_label
from qdv_salmuera.utils.module_defaults import build_daily_lot, get_current_username


class CircuitoSalmueraWindow(tk.Toplevel):
    def __init__(self, master, db: DB):
        super().__init__(master)

        self.db = db

        self.title(f"Química del Valle - Producción - {module_label('salmuera')}")
        self.geometry("1280x860")
        self.minsize(980, 680)

        now = datetime.now()
        self.fixed_date_ddmmyyyy = now.strftime("%d/%m/%Y")
        self.fixed_date_iso = now.strftime("%Y-%m-%d")

        self.var_fecha = tk.StringVar(value=self.fixed_date_ddmmyyyy)
        self.var_hora = tk.StringVar(value=now.strftime("%H:%M"))

        self.var_electrolizador = tk.StringVar()
        self.var_celdas = tk.StringVar()
        self.var_turno = tk.StringVar()

        self.var_amperaje = tk.StringVar()
        self.var_caudal_agua = tk.StringVar()
        self.var_caudal_salmuera = tk.StringVar()

        self.var_hipo_conc = tk.StringVar()
        self.var_hipo_exceso_soda = tk.StringVar()

        self.var_sal_temp = tk.StringVar()
        self.var_sal_conc = tk.StringVar()
        self.var_sal_ph = tk.StringVar()

        self.var_soda_conc = tk.StringVar()
        self.var_declor_ph = tk.StringVar()

        self.operadores = self.db.fetch_operadores()
        self.var_operador = tk.StringVar()
        self.cbo_operador = None
        self.var_lote = tk.StringVar()

        self.txt_obs = None

        self.voltage_vars: List[tk.StringVar] = []
        self.voltage_container = None
        self.voltage_block = None
        self.var_voltaje_total = tk.StringVar(value="")

        self.vcmd_float = (self.register(validate_float), "%P")
        self.vcmd_int = (self.register(validate_int), "%P")

        self.tree = None
        self.tree_cols = []
        self._selected_row_id: Optional[int] = None

        self.btn_guardar = None
        self.btn_historial = None
        self.btn_borrar = None
        self.lbl_status = None

        # Cronómetro análisis (un vencimiento por electrolizador; fuente DB)
        self._next_due_by_electrolizador: Dict[int, datetime] = {}
        self._electro_timer_vars: Dict[int, tk.StringVar] = {}
        self._electro_timer_labels: Dict[int, ttk.Label] = {}
        self._timer_electro_inner: Optional[ttk.Frame] = None
        self.var_motivo_atraso = tk.StringVar()
        self.ent_motivo_atraso = None
        self._timer_job = None
        self._prev_form_timer_seconds_left: Optional[int] = None

        # Advertencias rojas grandes
        self.warn_caudad = None
        self.warn_voltajes = None

        self._build_ui()
        self._set_default_operador_lote()
        self._rebuild_electro_timer_panel()
        self._start_timer()
        self._wire_validation()
        self._refresh_save_state()
        self._load_day_table()

    def _form_electrolizador_int_optional(self) -> Optional[int]:
        s = self.var_electrolizador.get().strip()
        if not s or not is_int_ok(s):
            return None
        try:
            v = int(s)
        except Exception:
            return None
        return v if v > 0 else None

    def _normalize_naive_dt(self, dt: datetime) -> datetime:
        if getattr(dt, "tzinfo", None) is not None:
            return dt.astimezone().replace(tzinfo=None)
        return dt

    def _raw_seconds_until_due(self, electrolizador: int) -> int:
        due = self._next_due_by_electrolizador.get(int(electrolizador))
        if due is None:
            return ANALYSIS_INTERVAL_SECONDS
        due = self._normalize_naive_dt(due)
        return int((due - datetime.now()).total_seconds())

    def _sync_motivo_atraso_enabled(self, fe: Optional[int], cur: Optional[int]) -> None:
        ent = getattr(self, "ent_motivo_atraso", None)
        if ent is None:
            return
        overdue = fe is not None and cur is not None and cur <= 0
        if overdue:
            ent.configure(state="normal")
        else:
            self.var_motivo_atraso.set("")
            ent.configure(state="disabled")

    def _rebuild_electro_timer_panel(self) -> None:
        self._next_due_by_electrolizador.clear()
        self._electro_timer_vars.clear()
        self._electro_timer_labels.clear()
        inner = getattr(self, "_timer_electro_inner", None)
        if inner is None:
            return
        for w in inner.winfo_children():
            w.destroy()
        try:
            ids = self.db.fetch_distinct_salmuera_electrolizador_ids()
        except Exception:
            ids = []
        if not ids:
            ttk.Label(inner, text="No hay electrolizadores con registros históricos.").pack(anchor="w")
            return
        now = datetime.now()
        for eid in ids:
            due: Optional[datetime] = None
            try:
                iso = self.db.fetch_last_salmuera_created_at_iso_for_electrolizador(eid)
                if iso:
                    created = datetime.fromisoformat(str(iso))
                    created = self._normalize_naive_dt(created)
                    due = created + timedelta(seconds=ANALYSIS_INTERVAL_SECONDS)
            except Exception:
                due = now + timedelta(seconds=ANALYSIS_INTERVAL_SECONDS)
            if due is not None:
                self._next_due_by_electrolizador[int(eid)] = self._normalize_naive_dt(due)
            row = ttk.Frame(inner)
            row.pack(fill="x", pady=1)
            ttk.Label(row, text=f"Electrolizador {eid}", width=16).pack(side="left")
            var = tk.StringVar(value="--:--:--")
            lbl = ttk.Label(row, textvariable=var, font=("Segoe UI", 12, "bold"))
            lbl.pack(side="left", padx=(8, 0))
            self._electro_timer_vars[int(eid)] = var
            self._electro_timer_labels[int(eid)] = lbl

    def _start_timer(self):
        # Cancela el timer anterior si existe
        if getattr(self, "_timer_job", None) is not None:
            try:
                self.after_cancel(self._timer_job)
            except Exception:
                pass
            self._timer_job = None

        self._prev_form_timer_seconds_left = None

        # Arranca el loop
        self._tick_timer()


    def _stop_timer(self):
        job = getattr(self, "_timer_job", None)
        if job is not None:
            try:
                self.after_cancel(job)
            except Exception:
                pass
        self._timer_job = None

    def _tick_timer(self):
        try:
            if not self.winfo_exists():
                return
        except Exception:
            return

        try:
            now = datetime.now()
            for eid, var in self._electro_timer_vars.items():
                lbl = self._electro_timer_labels.get(eid)
                if lbl is None:
                    continue
                due = self._next_due_by_electrolizador.get(eid)
                if due is None:
                    var.set("--:--:--")
                    lbl.config(foreground=QDV_COLORS.get("muted", QDV_COLORS["fg"]))
                    continue
                nd = self._normalize_naive_dt(due)
                remaining = int((nd - now).total_seconds())
                var.set(format_hhmmss(max(0, remaining)))
                lbl.config(foreground=(QDV_COLORS["danger"] if remaining <= 0 else QDV_COLORS["fg"]))

            fe = self._form_electrolizador_int_optional()
            if fe is None:
                self._sync_motivo_atraso_enabled(None, None)
                self._prev_form_timer_seconds_left = None
            else:
                cur = self._raw_seconds_until_due(fe)
                self._sync_motivo_atraso_enabled(fe, cur)
                prev = self._prev_form_timer_seconds_left
                crossed = prev is not None and (prev > 0) != (cur > 0)
                need_refresh = crossed or (prev is None and cur <= 0)
                self._prev_form_timer_seconds_left = cur
                if need_refresh:
                    try:
                        self._refresh_save_state()
                    except Exception:
                        pass

        except Exception as e:
            try:
                for var in self._electro_timer_vars.values():
                    var.set(f"ERR: {type(e).__name__}")
            except Exception:
                pass

        finally:
            self._timer_job = self.after(1000, self._tick_timer)


    def destroy(self):
        # importante: frenar after para que no quede corriendo
        try:
            self._stop_timer()
        except Exception:
            pass
        super().destroy()


    def _add_operador(self):
        name = add_operador_via_dialog(self, self.db)
        if name is None:
            return
        self.operadores = self.db.fetch_operadores()
        self.cbo_operador["values"] = self.operadores
        self.var_operador.set(name)
        self._refresh_save_state()

    def _set_default_operador_lote(self):
        username = get_current_username(self).strip()
        if username:
            if username not in self.operadores:
                try:
                    self.db.add_operador(username)
                except sqlite3.IntegrityError:
                    pass
                self.operadores = self.db.fetch_operadores()
                if self.cbo_operador is not None:
                    self.cbo_operador["values"] = self.operadores
            self.var_operador.set(username)
        next_correlative = self.db.get_daily_sample_count("salmuera", self.fixed_date_iso) + 1
        self.var_lote.set(build_daily_lot(next_correlative, self.fixed_date_iso))

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # Padding lateral razonable, menos margen vertical para ganar altura a la tabla
        outer = ttk.Frame(self, padding=(12, 6))
        outer.grid(row=0, column=0, sticky="nsew")

        # Solo la fila de la tabla crece: ~todo el slack vertical va al Treeview
        _r = 0
        header = ttk.Frame(outer)
        header.grid(row=_r, column=0, sticky="ew", pady=(0, 4))
        _r += 1
        ttk.Label(header, text=module_label("salmuera"), font=("Segoe UI", 16, "bold")).pack(side="left")
        ttk.Label(header, text=f"Día: {self.fixed_date_ddmmyyyy}", font=("Segoe UI", 11)).pack(side="right")

        timer_box = ttk.LabelFrame(
            outer,
            text=f"Cronómetro próximo análisis (cada {ANALYSIS_INTERVAL_SECONDS // 3600} h) — por electrolizador",
            padding=4,
        )
        timer_box.grid(row=_r, column=0, sticky="ew", pady=(0, 4))
        _r += 1
        self._timer_electro_inner = ttk.Frame(timer_box)
        self._timer_electro_inner.pack(fill="x")
        trow = ttk.Frame(timer_box)
        trow.pack(fill="x", pady=(6, 0))
        ttk.Label(trow, text="Motivo atraso (si el electrolizador cargado está vencido):").pack(side="left", padx=(0, 6))
        self.ent_motivo_atraso = ttk.Entry(trow, textvariable=self.var_motivo_atraso, width=50)
        self.ent_motivo_atraso.pack(side="left", fill="x", expand=True)
        self.ent_motivo_atraso.configure(state="disabled")

        # Planilla horizontal: una sola línea de carga (orden según imagen)
        planilla_box = ttk.LabelFrame(outer, text=f"PLANILLA - {module_label('salmuera')}", padding=4)
        planilla_box.grid(row=_r, column=0, sticky="ew", pady=(0, 4))
        _r += 1

        # Contenedor con scroll horizontal para la fila de datos
        planilla_canvas = tk.Canvas(planilla_box, highlightthickness=0)
        hsb_planilla = ttk.Scrollbar(planilla_box, orient="horizontal", command=planilla_canvas.xview)
        planilla_canvas.configure(xscrollcommand=hsb_planilla.set)
        hsb_planilla.pack(side="bottom", fill="x")
        planilla_canvas.pack(fill="x")

        self.planilla_inner = ttk.Frame(planilla_canvas)
        self.planilla_inner_id = planilla_canvas.create_window((0, 0), window=self.planilla_inner, anchor="nw")
        self.planilla_inner.bind("<Configure>", lambda e: planilla_canvas.configure(scrollregion=planilla_canvas.bbox("all")))
        planilla_canvas.bind("<Configure>", lambda e: planilla_canvas.itemconfig(self.planilla_inner_id, width=max(e.width, self.planilla_inner.winfo_reqwidth())))

        pf = self.planilla_inner
        pad = 2
        w_s = 8
        w_m = 10
        col = 0

        _py = 1
        # Fila 0: títulos de agrupación (centrados sobre cada grupo; vacío donde no hay grupo)
        for _ in range(5):
            ttk.Label(pf, text="").grid(row=0, column=col, padx=pad, pady=_py); col += 1
        lbl_volt = ttk.Label(pf, text="Voltaje Celdas")
        lbl_volt.grid(row=0, column=col, columnspan=21, padx=pad, pady=_py, sticky="ew"); col += 21
        ttk.Label(pf, text="").grid(row=0, column=col, padx=pad, pady=_py); col += 1
        lbl_caud = ttk.Label(pf, text="Caudales (lts/h)")
        lbl_caud.grid(row=0, column=col, columnspan=2, padx=pad, pady=_py, sticky="ew"); col += 2
        lbl_hipo = ttk.Label(pf, text="Hipoclorito")
        lbl_hipo.grid(row=0, column=col, columnspan=2, padx=pad, pady=_py, sticky="ew"); col += 2
        lbl_sal = ttk.Label(pf, text="Salmuera Salida Celdas")
        lbl_sal.grid(row=0, column=col, columnspan=3, padx=pad, pady=_py, sticky="ew"); col += 3
        lbl_soda = ttk.Label(pf, text="Soda Salida Celdas")
        lbl_soda.grid(row=0, column=col, columnspan=1, padx=pad, pady=_py, sticky="ew"); col += 1
        lbl_decl = ttk.Label(pf, text="Declorinacion")
        lbl_decl.grid(row=0, column=col, columnspan=1, padx=pad, pady=_py, sticky="ew"); col += 1
        ttk.Label(pf, text="").grid(row=0, column=col, padx=pad, pady=_py); col += 1

        # Fila 1: encabezados (labels 1..N y Total se rellenan en _rebuild_voltage_fields)
        col = 0
        ttk.Label(pf, text="Fecha").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Electrolizador").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Turno").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Hora").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Celdas").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        self.voltage_block = ttk.Frame(pf)
        self.voltage_block.grid(row=1, column=col, rowspan=2, columnspan=20, sticky="nw", padx=pad, pady=_py)
        self.voltage_container = self.voltage_block
        col += 20
        ttk.Label(pf, text="Total").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Amp").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Agua").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Salmuer").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Concentracion Cloro").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Exceso Soda").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Temp").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Concentración").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Ph").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Concentración").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Ph").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Operador").grid(row=1, column=col, padx=pad, pady=_py); col += 1
        ttk.Label(pf, text="Lote").grid(row=1, column=col, padx=pad, pady=_py); col += 1

        # Fila 2: una sola línea de datos (Total y resto alineados con sus labels)
        col = 0
        ttk.Entry(pf, textvariable=self.var_fecha, width=w_m, state="readonly").grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_electrolizador, width=w_s, validate="key", validatecommand=self.vcmd_int).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Combobox(pf, textvariable=self.var_turno, values=["M", "T", "N"], width=6, state="readonly").grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_hora, width=6, state="readonly").grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_celdas, width=6, validate="key", validatecommand=self.vcmd_int).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        # columnas 5-24: contenido del voltage_block (labels + entries se crean en _rebuild_voltage_fields)
        col = 25
        self.ent_voltaje_total = ttk.Entry(pf, textvariable=self.var_voltaje_total, width=w_s, state="readonly")
        self.ent_voltaje_total.grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_amperaje, width=w_s, validate="key", validatecommand=self.vcmd_float).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_caudal_agua, width=w_s, validate="key", validatecommand=self.vcmd_float).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_caudal_salmuera, width=w_s, validate="key", validatecommand=self.vcmd_float).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_hipo_conc, width=w_s, validate="key", validatecommand=self.vcmd_float).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_hipo_exceso_soda, width=w_s, validate="key", validatecommand=self.vcmd_float).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_sal_temp, width=w_s, validate="key", validatecommand=self.vcmd_float).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_sal_conc, width=w_s, validate="key", validatecommand=self.vcmd_float).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_sal_ph, width=w_s, validate="key", validatecommand=self.vcmd_float).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_soda_conc, width=w_s, validate="key", validatecommand=self.vcmd_float).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        ttk.Entry(pf, textvariable=self.var_declor_ph, width=w_s, validate="key", validatecommand=self.vcmd_float).grid(row=2, column=col, padx=pad, pady=_py); col += 1
        op_f = ttk.Frame(pf)
        op_f.grid(row=2, column=col, sticky="w", padx=pad, pady=_py)
        self.cbo_operador = ttk.Combobox(op_f, textvariable=self.var_operador, values=self.operadores, width=18, state="readonly")
        self.cbo_operador.pack(side="left")
        ttk.Button(op_f, text="+", width=2, command=self._add_operador).pack(side="left", padx=(4, 0))
        col += 1
        ttk.Entry(pf, textvariable=self.var_lote, width=w_m, state="readonly").grid(row=2, column=col, padx=pad, pady=_py)

        # Advertencias (debajo de la planilla, sin bloquear)
        self.warn_voltajes = ttk.Label(
            outer,
            text=f"Voltajes por celda entre {VOLTAGE_MIN} y {VOLTAGE_MAX}. Caudal salmuera > caudal agua.",
            font=("Segoe UI", 8),
            foreground=QDV_COLORS["muted"],
        )
        self.warn_voltajes.grid(row=_r, column=0, sticky="w", pady=(0, 2))
        _r += 1
        self.warn_caudad = None

        # Observaciones (compacto; misma funcionalidad)
        obs_frame = ttk.LabelFrame(outer, text="Observaciones", padding=4)
        obs_frame.grid(row=_r, column=0, sticky="ew", pady=(0, 4))
        _r += 1
        self.txt_obs = tk.Text(obs_frame, height=2, wrap="word")
        self.txt_obs.configure(
            bg=QDV_COLORS.get("input_inset", QDV_COLORS["input_bg"]),
            fg=QDV_COLORS["fg"],
            insertbackground=QDV_COLORS["accent"],
            relief="flat",
            highlightthickness=0,
        )
        self.txt_obs.pack(fill="x")

        # Botones: Guardar, Cerrar, Historial
        act = ttk.Frame(outer)
        act.grid(row=_r, column=0, sticky="ew", pady=(0, 4))
        _r += 1
        self.lbl_status = ttk.Label(act, text="", font=("Segoe UI", 10))
        self.lbl_status.pack(side="left", fill="x", expand=True)
        self.btn_guardar = ttk.Button(act, text="Guardar", command=self.on_save)
        self.btn_guardar.pack(side="right", padx=(8, 0))
        self.btn_historial = ttk.Button(act, text="Historial", command=self.on_historial)
        self.btn_historial.pack(side="right", padx=(8, 0))
        ttk.Button(act, text="Cerrar", command=self.destroy).pack(side="right")

        # Tabla de registros del día: ocupa todo el espacio vertical restante (prioridad operativa)
        bottom_box = ttk.LabelFrame(outer, text="Registros del día (solo hoy) - Doble clic para editar", padding=4)
        bottom_box.grid(row=_r, column=0, sticky="nsew")
        outer.rowconfigure(_r, weight=1)
        outer.columnconfigure(0, weight=1)

        bottom_box.rowconfigure(1, weight=1)
        bottom_box.columnconfigure(0, weight=1)

        toolbar = ttk.Frame(bottom_box)
        toolbar.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 4))

        ttk.Label(toolbar, text="Tip: doble clic sobre una fila para editar cualquier parámetro.", font=("Segoe UI", 8))\
            .pack(side="left")
        self.btn_borrar = ttk.Button(toolbar, text="Borrar seleccionado", command=self.on_delete_selected)
        self.btn_borrar.pack(side="right")

        tv_wrap = ttk.Frame(bottom_box)
        tv_wrap.grid(row=1, column=0, columnspan=2, sticky="nsew")
        tv_wrap.rowconfigure(0, weight=1)
        tv_wrap.columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(tv_wrap, columns=(), show="headings", style="Treeview")
        vsb = ttk.Scrollbar(tv_wrap, orient="vertical", command=self.tree.yview)
        hsb_tree = ttk.Scrollbar(tv_wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb_tree.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb_tree.grid(row=1, column=0, sticky="ew")

        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self.tree.bind("<Double-1>", self._on_tree_double_click)

    def _on_tree_select(self, _event):
        sel = self.tree.selection()
        if not sel:
            self._selected_row_id = None
            return
        vals = self.tree.item(sel[0], "values")
        if not vals:
            self._selected_row_id = None
            return
        try:
            self._selected_row_id = int(vals[0])
        except Exception:
            self._selected_row_id = None

    def _on_tree_double_click(self, _event):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], "values")
        if not vals:
            return
        try:
            rid = int(vals[0])
        except Exception:
            return

        dlg = EditRegistroDialog(self, self.db, rid)
        self.wait_window(dlg)
        self.operadores = self.db.fetch_operadores()
        self.cbo_operador["values"] = self.operadores
        self._load_day_table()

    def _clear_voltage_fields(self):
        block = getattr(self, "voltage_block", None) or getattr(self, "voltage_container", None)
        if block is not None:
            for w in block.winfo_children():
                w.destroy()
        self.voltage_vars = []
        if getattr(self, "var_voltaje_total", None) is not None:
            self.var_voltaje_total.set("")

    def _validate_voltages_rule(self, n: int) -> Tuple[bool, str]:
        voltajes = []
        for i, var in enumerate(self.voltage_vars):
            s = var.get().strip()
            if not s:
                return False, "Complete los voltajes."
            try:
                v = float(s)
            except Exception:
                return False, "Voltajes: solo números y punto (.)"
            voltajes.append(v)

        # Si hay más de 1 celda, aplicar rango por celda
        if n > 1:
            for i, v in enumerate(voltajes):
                if v < VOLTAGE_MIN or v > VOLTAGE_MAX:
                    return False, f"ADVERTENCIA: Voltaje de celda {i+1} fuera de rango ({VOLTAGE_MIN} a {VOLTAGE_MAX})."

        return True, "OK"


    def _update_voltaje_total(self):
        total = 0.0
        for var in self.voltage_vars:
            s = var.get().strip()
            if s:
                try:
                    total += float(s)
                except ValueError:
                    pass
        self.var_voltaje_total.set(fmt_num(total) if total else "")

    def _rebuild_voltage_fields(self):
        raw = self.var_celdas.get().strip()
        if raw == "" or (not raw.isdigit()):
            self._clear_voltage_fields()
            self._refresh_save_state()
            return

        n = int(raw)
        if n <= 0 or n > 20:
            self._clear_voltage_fields()
            self._refresh_save_state()
            return

        if len(self.voltage_vars) == n:
            self._refresh_save_state()
            return

        self._clear_voltage_fields()
        block = getattr(self, "voltage_block", None) or getattr(self, "voltage_container", None)
        if block is None:
            self._refresh_save_state()
            return
        pad_c = 1
        w_cel = 6
        for i in range(n):
            var = tk.StringVar()
            var.trace_add("write", lambda *_: self._refresh_save_state())
            var.trace_add("write", lambda *_: self._update_voltaje_total())
            self.voltage_vars.append(var)
            lbl = ttk.Label(block, text=str(i + 1), anchor="center", width=w_cel)
            lbl.grid(row=0, column=i, padx=pad_c, pady=1, sticky="ew")
            ent = ttk.Entry(block, textvariable=var, width=w_cel, validate="key", validatecommand=self.vcmd_float)
            ent.grid(row=1, column=i, padx=pad_c, pady=1, sticky="ew")
        block.columnconfigure(tuple(range(n)), weight=0, uniform="v")
        self._update_voltaje_total()
        self._refresh_save_state()

    def _wire_validation(self):
        vars_to_trace = [
            self.var_electrolizador, self.var_celdas, self.var_turno,
            self.var_amperaje, self.var_caudal_agua, self.var_caudal_salmuera,
            self.var_hipo_conc, self.var_hipo_exceso_soda,
            self.var_sal_temp, self.var_sal_conc, self.var_sal_ph,
            self.var_soda_conc, self.var_declor_ph,
            self.var_operador, self.var_lote
        ]
        for v in vars_to_trace:
            v.trace_add("write", lambda *_: self._refresh_save_state())

        self.var_celdas.trace_add("write", lambda *_: self._rebuild_voltage_fields())
        self.txt_obs.bind("<KeyRelease>", lambda _e: self._refresh_save_state())
        self.var_motivo_atraso.trace_add("write", lambda *_: self._refresh_save_state())

    def _validate_all(self):
        n = 0
        try: n = int(self.var_celdas.get().strip())
        except: return False, "Cantidad de celdas inválida."

        if not is_int_ok(self.var_electrolizador.get()):
            return False, "Electrolizador debe ser numérico entero."
        if not is_int_ok(self.var_celdas.get()):
            return False, "Cantidad de celdas debe ser numérica entera."

        # Regla especial:
        # Si cantidad_celdas == 1, se permite solo 2 veces seguidas por electrolizador.
        if n == 1:
            elect = int(self.var_electrolizador.get().strip())
            consec = self._count_consecutive_single_cell(elect)
            if consec >= 2:
                return False, "ADVERTENCIA: Para este electrolizador ya hubo 2 registros seguidos con 1 celda. En la tercera carga, la cantidad de celdas debe ser distinta de 1."

        n = int(self.var_celdas.get().strip())
        if n <= 0:
            return False, "Cantidad de celdas debe ser mayor que 0."
        if n > 20:
            return False, "Cantidad de celdas no puede ser mayor que 20."
        if len(self.voltage_vars) != n:
            return False, "Complete los voltajes según la cantidad de celdas."

        if self.var_turno.get().strip() not in ("M", "T", "N"):
            return False, "Turno es obligatorio (M, T o N)."

        # Validación de voltajes
        voltajes = []
        for i, var in enumerate(self.voltage_vars):
            s = var.get().strip()
            if not s:
                return False, "Complete los voltajes."
            try:
                v = float(s)
            except Exception:
                return False, "Voltajes: solo números y punto (.)"
            voltajes.append(v)

        # Regla: rango por celda SOLO si hay más de 1 celda
        # Si n == 1, es Voltaje total (manual) y puede ser > 4.5
        if n > 1:
            for i, v in enumerate(voltajes):
                if v < VOLTAGE_MIN or v > VOLTAGE_MAX:
                    return False, f"ADVERTENCIA: Voltaje de celda {i+1} fuera de rango ({VOLTAGE_MIN} a {VOLTAGE_MAX})."


        checks = [
            ("Amperaje", self.var_amperaje.get()),
            ("Caudal agua (L/h)", self.var_caudal_agua.get()),
            ("Caudal salmuera (L/h)", self.var_caudal_salmuera.get()),
            ("Hipoclorito - Concentración", self.var_hipo_conc.get()),
            ("Hipoclorito - Exceso de soda", self.var_hipo_exceso_soda.get()),
            ("Salmuera salida - Temperatura", self.var_sal_temp.get()),
            ("Salmuera salida - Concentración", self.var_sal_conc.get()),
            ("Salmuera salida - pH", self.var_sal_ph.get()),
            ("Soda salida - Concentración", self.var_soda_conc.get()),
            ("Declorinación - pH", self.var_declor_ph.get()),
        ]
        for label, val in checks:
            if not is_float_ok(val):
                return False, f"{label} es obligatorio y debe ser numérico (usar punto)."

        q_agua = float(self.var_caudal_agua.get().strip())
        q_sal = float(self.var_caudal_salmuera.get().strip())

        # Regla: Salmuera SIEMPRE mayor que Agua
        if q_sal <= q_agua:
            return False, "ADVERTENCIA: Caudal de salmuera DEBE ser mayor al caudal de agua."


        op = self.var_operador.get().strip()
        if not op:
            return False, "Operador es obligatorio."
        if op not in self.operadores:
            return False, "Operador inválido. Seleccione uno del desplegable o agregue uno nuevo con +."
        lote = self.var_lote.get().strip()
        if not lote:
            return False, "Lote es obligatorio."

        fe = self._form_electrolizador_int_optional()
        if fe is not None and self._raw_seconds_until_due(fe) <= 0:
            if not self.var_motivo_atraso.get().strip():
                return False, "ADVERTENCIA: El análisis está vencido. Indique el motivo del atraso."

        return True, "OK"

    def _count_consecutive_single_cell(self, electrolizador: int) -> int:
        """
        Cuenta cuántos registros consecutivos más recientes tienen cantidad_celdas = 1
        para el electrolizador indicado.
        """
        try:
            rows = self.db.fetch_last_salmuera_by_electrolizador(electrolizador, limit=10)
        except Exception:
            return 0

        count = 0
        for r in rows:
            try:
                if int(r.get("cantidad_celdas", 0)) == 1:
                    count += 1
                else:
                    break
            except Exception:
                break
        return count


    def _refresh_save_state(self):
        ok, msg = self._validate_all()

        # Mostrar mensaje rojo grande si no está OK (si usás un label de estado)
        if hasattr(self, "lbl_status") and self.lbl_status is not None:
            if ok:
                self.lbl_status.config(text="")
            else:
                self.lbl_status.config(text=msg, foreground="red", font=("Segoe UI", 12, "bold"))

        # Habilitar / deshabilitar Guardar
        if hasattr(self, "btn_guardar") and self.btn_guardar is not None:
            self.btn_guardar.configure(state=("normal" if ok else "disabled"))


    def _build_tree_columns_for_day(self, registros: List[Dict[str, Any]]):
        max_celdas = 0
        for r in registros:
            try:
                max_celdas = max(max_celdas, int(r["cantidad_celdas"]))
            except Exception:
                pass

        cols = [
            ("id", "ID", 60),
            ("fecha", "Fecha", 90),
            ("hora", "Hora", 70),
            ("elect", "Elect.", 70),
            ("celdas", "Celdas", 70),
            ("turno", "Turno", 70),
        ]
        for i in range(1, max_celdas + 1):
            cols.append((f"v{i}", f"V{i}", 75))
        cols += [
            ("vtotal", "V Total", 80),
            ("amper", "Amperaje", 85),
            ("qagua", "Q Agua (L/h)", 100),
            ("qsal", "Q Salmuera (L/h)", 115),
            ("hipoc", "Hipo Conc", 85),
            ("hipexs", "Hipo Exceso Soda", 130),
            ("stemp", "Sal T°", 75),
            ("sconc", "Sal Conc", 85),
            ("sph", "Sal pH", 75),
            ("sodac", "Soda Conc", 95),
            ("dph", "Declor pH", 85),
            ("oper", "Operador", 140),
            ("lote", "Lote", 100),
            ("obs", "Observaciones", 260),
            ("creado", "Creado", 160),
        ]

        self.tree_cols = [c[0] for c in cols]
        self.tree.configure(columns=self.tree_cols)

        for col_key, col_title, col_w in cols:
            self.tree.heading(col_key, text=col_title)
            anchor = "w" if col_key in ("oper", "obs", "creado") else "center"
            self.tree.column(col_key, width=col_w, anchor=anchor, stretch=False)

    def _load_day_table(self):
        registros = self.db.fetch_salmuera_by_date(self.fixed_date_iso)
        self._build_tree_columns_for_day(registros)

        for item in self.tree.get_children():
            self.tree.delete(item)

        for r in registros:
            volts = r["voltajes_celdas"] or []
            row = [
                r["id"],
                iso_to_ddmmyyyy(r["fecha_iso"]),
                r["hora_hm"],
                r["electrolizador"],
                r["cantidad_celdas"],
                r["turno"],
            ]

            v_cols = [c for c in self.tree_cols if c.startswith("v") and c[1:].isdigit()]
            vmax = len(v_cols)
            for i in range(vmax):
                row.append(fmt_num(volts[i]) if i < len(volts) else "")

            row += [
                fmt_num(r["voltaje_total"]),
                fmt_num(r["amperaje"]),
                fmt_num(r["caudal_agua_l_h"]),
                fmt_num(r["caudal_salmuera_l_h"]),
                fmt_num(r["hipo_conc"]),
                fmt_num(r["hipo_exceso_soda"]),
                fmt_num(r["sal_temp"]),
                fmt_num(r["sal_conc"]),
                fmt_num(r["sal_ph"]),
                fmt_num(r["soda_conc"]),
                fmt_num(r["declor_ph"]),
                r["operador"],
                r.get("lote", ""),
                (r.get("observaciones", "") or ""),
                r["created_at_iso"],
            ]
            self.tree.insert("", "end", values=row)

        self._selected_row_id = None

    def on_save(self):
        import traceback
        try:
            # 1) Validar
            ok, msg = self._validate_all()
            if not ok:
                messagebox.showerror("No se puede guardar", msg)
                return

            # 2) Armar datos mínimos y guardar
            now = datetime.now()
            n = int(self.var_celdas.get().strip())

            voltajes = []
            for var in self.voltage_vars:
                voltajes.append(float(var.get().strip()))

            data = {
                "fecha_iso": self.fixed_date_iso,
                "hora_hm": now.strftime("%H:%M"),
                "electrolizador": int(self.var_electrolizador.get().strip()),
                "cantidad_celdas": n,
                "turno": self.var_turno.get().strip(),

                "voltajes_celdas": voltajes,
                "voltaje_total": float(sum(voltajes)),

                "amperaje": float(self.var_amperaje.get().strip()),
                "caudal_agua_l_h": float(self.var_caudal_agua.get().strip()),
                "caudal_salmuera_l_h": float(self.var_caudal_salmuera.get().strip()),

                "hipo_conc": float(self.var_hipo_conc.get().strip()),
                "hipo_exceso_soda": float(self.var_hipo_exceso_soda.get().strip()),

                "sal_temp": float(self.var_sal_temp.get().strip()),
                "sal_conc": float(self.var_sal_conc.get().strip()),
                "sal_ph": float(self.var_sal_ph.get().strip()),

                "soda_conc": float(self.var_soda_conc.get().strip()),
                "declor_ph": float(self.var_declor_ph.get().strip()),

                "operador": self.var_operador.get().strip(),
                "lote": self.var_lote.get().strip(),
                "observaciones": self.txt_obs.get("1.0", "end").strip(),
                "atraso_motivo": self.var_motivo_atraso.get().strip(),

                "created_at_iso": now.isoformat(timespec="seconds"),
            }

            # Insertar una sola vez
            self.db.insert_salmuera(data)

            # Refrescar tabla
            self._load_day_table()
            self._clear_form()
            self._set_default_operador_lote()
            self._refresh_save_state()

            self._rebuild_electro_timer_panel()
            self._start_timer()

        except Exception as e:
            messagebox.showerror("Error al guardar", f"{e}\n\n{traceback.format_exc()}")


    def _clear_form(self):
        self.var_electrolizador.set("")
        self.var_celdas.set("")
        self.var_turno.set("")
        self._clear_voltage_fields()

        self.var_amperaje.set("")
        self.var_caudal_agua.set("")
        self.var_caudal_salmuera.set("")

        self.var_hipo_conc.set("")
        self.var_hipo_exceso_soda.set("")

        self.var_sal_temp.set("")
        self.var_sal_conc.set("")
        self.var_sal_ph.set("")

        self.var_soda_conc.set("")
        self.var_declor_ph.set("")
        self.var_operador.set("")
        self.var_lote.set("")
        self.txt_obs.delete("1.0", "end")

        self._refresh_save_state()

    def on_delete_selected(self):
        if self._selected_row_id is None:
            messagebox.showwarning("Atención", "Seleccione un registro en la tabla para borrarlo.")
            return

        dlg = CodigoSeguridadDialog(self, title="Borrar registro - Código de seguridad")
        self.wait_window(dlg)
        if dlg.result is None:
            return

        if dlg.result != SECURITY_DELETE_CODE:
            messagebox.showerror("Código incorrecto", "Código de seguridad inválido. No se borró el registro.")
            return

        if not messagebox.askyesno("Confirmar", f"¿Confirma borrar el registro ID {self._selected_row_id}?"):
            return

        try:
            self.db.delete_salmuera_by_id(self._selected_row_id)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo borrar el registro.\n\nDetalle: {e}")
            return

        self._load_day_table()
        self._rebuild_electro_timer_panel()
        self._start_timer()
        messagebox.showinfo("OK", f"Registro ID {self._selected_row_id} eliminado.")

    def on_historial(self):
        if not OPENPYXL_OK:
            messagebox.showerror("Falta dependencia", "Para exportar a Excel instalá:\n\npip install openpyxl")
            return

        hoy_ddmmyyyy = datetime.now().strftime("%d/%m/%Y")
        dlg = HistorialDialog(self, default_desde=hoy_ddmmyyyy, default_hasta=hoy_ddmmyyyy)
        self.wait_window(dlg)
        if not dlg.result:
            return

        desde_s, hasta_s = dlg.result
        d1 = parse_date_ddmmyyyy(desde_s)
        d2 = parse_date_ddmmyyyy(hasta_s)
        if not d1 or not d2:
            messagebox.showerror("Error", "Rango de fechas inválido.")
            return

        desde_iso = date_to_iso(d1)
        hasta_iso = date_to_iso(d2)

        try:
            registros = self.db.fetch_salmuera_range(desde_iso, hasta_iso)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el historial.\n\nDetalle: {e}")
            return

        if not registros:
            messagebox.showinfo("Sin datos", "No hay registros en el rango seleccionado.")
            return

        default_name = f"historial_salmuera_{desde_iso}_a_{hasta_iso}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Guardar reporte Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")]
        )
        if not save_path:
            return

        try:
            self._export_excel(registros, save_path)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el Excel.\n\nDetalle: {e}")
            return

        messagebox.showinfo("OK", f"Reporte generado:\n{save_path}")

    def _export_excel(self, registros: List[Dict[str, Any]], path: str):
        max_celdas = 0
        for r in registros:
            try:
                max_celdas = max(max_celdas, int(r["cantidad_celdas"]))
            except Exception:
                pass

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Salmuera"

        headers = ["ID", "Fecha", "Hora", "Electrolizador", "Cantidad Celdas", "Turno"]
        for i in range(1, max_celdas + 1):
            headers.append(f"V{i}")
        headers.append("Voltaje Total")
        headers += [
            "Amperaje", "Caudal Agua (L/h)", "Caudal Salmuera (L/h)",
            "Hipoclorito - Concentración", "Hipoclorito - Exceso Soda",
            "Salmuera Salida - Temperatura", "Salmuera Salida - Concentración", "Salmuera Salida - pH",
            "Soda Salida - Concentración", "Declorinación - pH",
            "Operador", "Lote", "Observaciones", "Motivo atraso", "Creado (ISO)"
        ]
        ws.append(headers)

        for r in registros:
            row = [
                r["id"],
                iso_to_ddmmyyyy(r["fecha_iso"]),
                r["hora_hm"],
                r["electrolizador"],
                r["cantidad_celdas"],
                r["turno"],
            ]
            volts = r["voltajes_celdas"] or []
            for i in range(max_celdas):
                row.append(volts[i] if i < len(volts) else "")
            row.append(r["voltaje_total"])

            row += [
                r["amperaje"],
                r["caudal_agua_l_h"],
                r["caudal_salmuera_l_h"],
                r["hipo_conc"],
                r["hipo_exceso_soda"],
                r["sal_temp"],
                r["sal_conc"],
                r["sal_ph"],
                r["soda_conc"],
                r["declor_ph"],
                r["operador"],
                r.get("lote", ""),
                r.get("observaciones", ""),
                r.get("atraso_motivo", ""),
                r["created_at_iso"]
            ]
            ws.append(row)

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter]:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)

        wb.save(path)