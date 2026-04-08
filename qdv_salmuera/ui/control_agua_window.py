from __future__ import annotations

import sqlite3
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from qdv_salmuera.data.db import DB
from qdv_salmuera.config.settings import SECURITY_DELETE_CODE
from qdv_salmuera.utils.validators import validate_float, validate_int, is_int_ok, is_float_ok
from qdv_salmuera.utils.validators import fmt_num
from qdv_salmuera.utils.dates import iso_to_ddmmyyyy, parse_date_ddmmyyyy, date_to_iso
from qdv_salmuera.ui.widgets import ScrollableFrame
from qdv_salmuera.ui.dialogs import HistorialDialog, CodigoSeguridadDialog
from qdv_salmuera.ui.produccion_ui_helpers import add_operador_via_dialog, bind_excel_field_navigation
from qdv_salmuera.ui.theme import QDV_COLORS
from qdv_salmuera.ui.module_timer import PersistentModuleTimer, TimerConfig
from qdv_salmuera.ui.module_labels import module_label
from qdv_salmuera.utils.module_defaults import build_daily_lot, get_current_username


def _compute_turno_from_hour(hhmm: str) -> str:
    """
    Turnos:
    - Mañana: 00:00–08:00  -> "M"
    - Tarde:  08:00–16:00  -> "T"
    - Noche:  16:00–00:00  -> "N"
    """
    try:
        h = int((hhmm or "").split(":")[0])
    except Exception:
        h = 0
    if 0 <= h < 8:
        return "M"
    if 8 <= h < 16:
        return "T"
    return "N"


class ControlAguaWindow(tk.Toplevel):
    COLUMN_STATES = ("En operación", "Regenerada", "Por regenerar")

    def __init__(self, master, db: DB):
        super().__init__(master)
        self.db = db

        self.title(f"Química del Valle - Producción - {module_label('agua')}")
        self.geometry("1200x800")
        self.minsize(860, 600)

        now = datetime.now()
        self.fixed_date_ddmmyyyy = now.strftime("%d/%m/%Y")
        self.fixed_date_iso = now.strftime("%Y-%m-%d")

        self.var_fecha = tk.StringVar(value=self.fixed_date_ddmmyyyy)
        self.var_hora = tk.StringVar(value=now.strftime("%H:%M"))
        self.var_turno = tk.StringVar(value=_compute_turno_from_hour(self.var_hora.get()))

        self.operadores = self.db.fetch_operadores()
        self.var_operador = tk.StringVar()
        self.cbo_operador: Optional[ttk.Combobox] = None

        self.var_lote = tk.StringVar()
        self.var_num_columna = tk.StringVar()
        self.var_temperatura = tk.StringVar()
        self.var_dureza = tk.StringVar()

        self.txt_obs: Optional[tk.Text] = None

        self.vcmd_float = (self.register(validate_float), "%P")
        self.vcmd_int = (self.register(validate_int), "%P")

        self.btn_guardar: Optional[ttk.Button] = None
        self.btn_historial: Optional[ttk.Button] = None
        self.btn_borrar: Optional[ttk.Button] = None
        self.lbl_status: Optional[ttk.Label] = None

        self.tree: Optional[ttk.Treeview] = None
        self.tree_cols: List[str] = []
        self._selected_row_id: Optional[int] = None
        self.column_rows: Dict[int, Dict[str, Any]] = {}

        # Timer Agua (8h)
        self.var_timer = tk.StringVar(value="08:00:00")
        self.lbl_timer = None
        self._timer = None

        self._build_ui()
        self._set_default_operador_lote()
        self._wire_validation()
        self._refresh_save_state()
        self._load_day_table()

        self._init_timer()

    def _clear_row(self) -> None:
        """Limpia solo los campos editables (fecha/hora/turno quedan automáticos)."""
        self.var_num_columna.set("")
        self.var_temperatura.set("")
        self.var_dureza.set("")
        if self.txt_obs is not None:
            self.txt_obs.delete("1.0", "end")
        self._set_default_operador_lote()

    def _set_default_operador_lote(self) -> None:
        username = get_current_username(self).strip()
        if username:
            if username not in self.operadores:
                try:
                    self.db.add_operador(username)
                except sqlite3.IntegrityError:
                    pass
                self.operadores = self.db.fetch_operadores()
                if self.cbo_operador is not None:
                    self.cbo_operador["values"] = self.operadores
            self.var_operador.set(username)
        next_correlative = self.db.get_daily_sample_count("agua", self.fixed_date_iso) + 1
        self.var_lote.set(build_daily_lot(next_correlative, self.fixed_date_iso))

    def _init_timer(self) -> None:
        def _fetch_last_iso() -> Optional[str]:
            last = self.db.fetch_last_agua()
            return (last or {}).get("created_at_iso")

        def _on_overdue(overdue: bool) -> None:
            if self.lbl_timer is not None:
                self.lbl_timer.configure(foreground=("red" if overdue else QDV_COLORS["fg"]))

        self._timer = PersistentModuleTimer(
            owner=self,
            config=TimerConfig(interval_seconds=8 * 60 * 60, default_label="08:00:00"),
            fetch_last_created_at_iso=_fetch_last_iso,
            out_var=self.var_timer,
            on_overdue_change=_on_overdue,
        )
        self._timer.start()

    def _build_ui(self) -> None:
        outer = ttk.Frame(self, padding=16)
        outer.pack(fill="both", expand=True)

        # Advertencias no bloqueantes (estilo centralizado en ui/theme.py)
        self._warn_entry_style = "QDV.Warn.TEntry"

        header = ttk.Frame(outer)
        header.pack(fill="x", pady=(0, 8))
        ttk.Label(header, text=module_label("agua"), font=("Segoe UI", 18, "bold")).pack(side="left")
        ttk.Label(header, text=f"Día: {self.fixed_date_ddmmyyyy}", font=("Segoe UI", 11)).pack(side="right")

        # Cronómetro
        timer_box = ttk.LabelFrame(outer, text="Cronómetro de análisis (cada 8 h)", padding=8)
        timer_box.pack(fill="x", pady=(0, 10))
        trow = ttk.Frame(timer_box)
        trow.pack(fill="x")
        ttk.Label(trow, text="Tiempo restante:").pack(side="left")
        self.lbl_timer = ttk.Label(trow, textvariable=self.var_timer, font=("Segoe UI", 14, "bold"))
        self.lbl_timer.pack(side="left", padx=(8, 0))

        sc = ScrollableFrame(outer)
        sc.pack(fill="both", expand=True)
        body = sc.inner

        # Form (una fila tipo Excel)
        reg = ttk.LabelFrame(body, text=module_label("agua"), padding=12)
        reg.pack(fill="x", pady=(0, 12))

        # Columnas (UNA sola fuente de verdad: widths en caracteres)
        cols: List[Tuple[str, str, int]] = [
            ("fecha", "Fecha", 10),
            ("hora", "Hora", 8),
            ("operador", "Operador", 15),
            ("op_add", "", 3),
            ("lote", "Lote", 12),
            ("columna", "Nro col.", 8),
            ("temperatura", "Temp", 8),
            ("dureza", "Dureza (max. 1 ppm)", 16),
            ("obs", "Observaciones", 30),
        ]

        # Planilla con scroll horizontal (encabezados + fila de carga)
        sheet_wrap = ttk.Frame(reg)
        sheet_wrap.pack(fill="x")

        sheet_canvas = tk.Canvas(
            sheet_wrap,
            background=QDV_COLORS.get("card_elevated", QDV_COLORS["card"]),
            highlightthickness=0,
            bd=0,
        )
        sheet_hsb = ttk.Scrollbar(sheet_wrap, orient="horizontal", command=sheet_canvas.xview)
        sheet_canvas.configure(xscrollcommand=sheet_hsb.set)

        sheet_canvas.pack(side="top", fill="x", expand=False)
        sheet_hsb.pack(side="bottom", fill="x")

        sheet = ttk.Frame(sheet_canvas)
        _sheet_win = sheet_canvas.create_window((0, 0), window=sheet, anchor="nw")

        def _sync_sheet_scroll(_event=None) -> None:
            sheet_canvas.configure(scrollregion=sheet_canvas.bbox("all"))

        def _sync_sheet_width(event) -> None:
            sheet_canvas.itemconfigure(_sheet_win, height=event.height)

        sheet.bind("<Configure>", _sync_sheet_scroll)
        sheet_canvas.bind("<Configure>", _sync_sheet_width)

        for i, (_k, title, wch) in enumerate(cols):
            ttk.Label(
                sheet,
                text=title,
                width=wch,
                anchor="center",
                font=("Segoe UI", 9, "bold"),
                foreground="#334155",
            ).grid(row=0, column=i, padx=(2, 2), pady=(0, 6), sticky="w")

        ttk.Entry(sheet, textvariable=self.var_fecha, state="readonly", width=cols[0][2]).grid(row=1, column=0, padx=(2, 2), sticky="w")
        ttk.Entry(sheet, textvariable=self.var_hora, state="readonly", width=cols[1][2]).grid(row=1, column=1, padx=(2, 2), sticky="w")

        self.cbo_operador = ttk.Combobox(
            sheet,
            textvariable=self.var_operador,
            values=self.operadores,
            state="readonly",
            width=cols[2][2],
        )
        self.cbo_operador.grid(row=1, column=2, padx=(2, 2), sticky="w")
        ttk.Button(sheet, text="+", width=cols[3][2], command=self._add_operador).grid(row=1, column=3, padx=(2, 2), sticky="w")

        self.ent_lote = ttk.Entry(sheet, textvariable=self.var_lote, width=cols[4][2], state="readonly")
        self.ent_lote.grid(row=1, column=4, padx=(2, 2), sticky="w")

        self.ent_columna = ttk.Entry(
            sheet,
            textvariable=self.var_num_columna,
            validate="key",
            validatecommand=self.vcmd_int,
            width=cols[5][2],
        )
        self.ent_columna.grid(row=1, column=5, padx=(2, 2), sticky="w")

        self.ent_temperatura = ttk.Entry(
            sheet,
            textvariable=self.var_temperatura,
            validate="key",
            validatecommand=self.vcmd_float,
            width=cols[6][2],
        )
        self.ent_temperatura.grid(row=1, column=6, padx=(2, 2), sticky="w")

        self.ent_dureza = ttk.Entry(
            sheet,
            textvariable=self.var_dureza,
            validate="key",
            validatecommand=self.vcmd_float,
            width=cols[7][2],
        )
        self.ent_dureza.grid(row=1, column=7, padx=(2, 2), sticky="w")

        self.txt_obs = tk.Text(sheet, height=2, width=cols[8][2], wrap="word")
        self.txt_obs.configure(
            bg=QDV_COLORS.get("input_inset", QDV_COLORS["input_bg"]),
            fg=QDV_COLORS["fg"],
            insertbackground=QDV_COLORS["accent"],
            relief="flat",
            highlightthickness=0,
            padx=2,
            pady=1,
        )
        self.txt_obs.grid(row=1, column=8, padx=(2, 2), sticky="w")

        # Navegación por teclado tipo Excel
        focus_order = [
            self.cbo_operador,
            self.ent_lote,
            self.ent_columna,
            self.ent_temperatura,
            self.ent_dureza,
            self.txt_obs,
        ]
        bind_excel_field_navigation(focus_order, self.on_save)

        # Acciones
        act = ttk.Frame(body)
        act.pack(fill="x", pady=(0, 10))
        self.lbl_status = ttk.Label(act, text="", font=("Segoe UI", 10))
        self.lbl_status.pack(side="left", fill="x", expand=True)

        self.btn_guardar = ttk.Button(act, text="Guardar", command=self.on_save)
        self.btn_guardar.pack(side="right", padx=(8, 0))

        self.btn_historial = ttk.Button(act, text="Historial", command=self.on_historial)
        self.btn_historial.pack(side="right", padx=(8, 0))

        ttk.Button(act, text="Cerrar", command=self.destroy).pack(side="right")

        self._build_columnas_intercambio_panel(body)
        self.refresh_columnas_intercambio_panel()

        bottom = ttk.LabelFrame(body, text="Registros del día", padding=12)
        bottom.pack(fill="both", expand=True)

        toolbar = ttk.Frame(bottom)
        toolbar.pack(fill="x", pady=(0, 8))
        ttk.Label(toolbar, text="Seleccione una fila para borrar.", font=("Segoe UI", 9), foreground=QDV_COLORS["muted"]).pack(side="left")
        self.btn_borrar = ttk.Button(toolbar, text="Borrar seleccionado", command=self.on_delete_selected)
        self.btn_borrar.pack(side="right")

        tv_wrap = ttk.Frame(bottom)
        tv_wrap.pack(fill="both", expand=True)

        self.tree_cols = [
            "id",
            "fecha",
            "hora",
            "operador",
            "lote",
            "columna",
            "temperatura",
            "dureza",
            "observaciones",
            "creado",
        ]

        self.tree = ttk.Treeview(tv_wrap, columns=self.tree_cols, show="headings", style="Treeview")
        vsb = ttk.Scrollbar(tv_wrap, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(bottom, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        widths = {
            "id": 60,
            "fecha": 90,
            "hora": 70,
            "operador": 130,
            "lote": 120,
            "columna": 90,
            "temperatura": 100,
            "dureza": 90,
            "observaciones": 260,
            "creado": 160,
        }

        for col_key in self.tree_cols:
            title = col_key
            anchor = "center"
            if col_key in ("operador", "lote", "observaciones", "creado"):
                anchor = "w"
            if col_key == "columna":
                title = "Columna"
            if col_key == "temperatura":
                title = "Temperatura"
            if col_key == "dureza":
                title = "Dureza"

            self.tree.heading(col_key, text=title)
            self.tree.column(col_key, width=widths.get(col_key, 120), anchor=anchor, stretch=False)

        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)

    def _wire_validation(self) -> None:
        vars_to_trace = [
            self.var_operador,
            self.var_lote,
            self.var_num_columna,
            self.var_temperatura,
            self.var_dureza,
        ]
        for v in vars_to_trace:
            v.trace_add("write", lambda *_: self._refresh_save_state())

        if self.txt_obs is not None:
            self.txt_obs.bind("<KeyRelease>", lambda _e: self._refresh_save_state())

    def _build_columnas_intercambio_panel(self, parent: ttk.Frame) -> None:
        panel = ttk.LabelFrame(parent, text="ESTADO DE COLUMNAS DE INTERCAMBIO IÓNICO", padding=10)
        panel.pack(fill="x", pady=(0, 12))
        self.column_panel = panel

        headers = [
            ("col", "Columna", 12),
            ("state", "Estado", 16),
            ("freg", "Fecha última regeneración", 18),
            ("hreg", "Hora última regeneración", 16),
            ("ds", "Dureza salida (ppm)", 16),
            ("dp", "Dureza post regeneración (ppm)", 24),
            ("obs", "Observaciones", 26),
            ("act", "Acción", 14),
            ("tag", "Indicador", 14),
        ]
        for i, (_k, txt, w) in enumerate(headers):
            ttk.Label(panel, text=txt, width=w, anchor="center", font=("Segoe UI", 9, "bold")).grid(
                row=0, column=i, padx=3, pady=(0, 6), sticky="w"
            )

        for idx, columna in enumerate((1, 2, 3), start=1):
            var_estado = tk.StringVar(value="En operación")
            var_fecha = tk.StringVar(value="")
            var_hora = tk.StringVar(value="")
            var_dureza_salida = tk.StringVar(value="")
            var_dureza_post = tk.StringVar(value="")
            var_obs = tk.StringVar(value="")

            ttk.Label(panel, text=f"Columna {columna}", width=12).grid(row=idx, column=0, padx=3, pady=3, sticky="w")
            cbo_estado = ttk.Combobox(
                panel,
                textvariable=var_estado,
                values=list(self.COLUMN_STATES),
                state="readonly",
                width=16,
            )
            cbo_estado.grid(row=idx, column=1, padx=3, pady=3, sticky="w")
            cbo_estado.bind("<<ComboboxSelected>>", lambda _e, c=columna: self._on_estado_columna_changed(c))

            ttk.Entry(panel, textvariable=var_fecha, state="readonly", width=18).grid(row=idx, column=2, padx=3, pady=3, sticky="w")
            ttk.Entry(panel, textvariable=var_hora, state="readonly", width=16).grid(row=idx, column=3, padx=3, pady=3, sticky="w")

            ent_salida = ttk.Entry(
                panel,
                textvariable=var_dureza_salida,
                validate="key",
                validatecommand=self.vcmd_float,
                width=16,
            )
            ent_salida.grid(row=idx, column=4, padx=3, pady=3, sticky="w")

            ent_post = ttk.Entry(
                panel,
                textvariable=var_dureza_post,
                validate="key",
                validatecommand=self.vcmd_float,
                width=24,
            )
            ent_post.grid(row=idx, column=5, padx=3, pady=3, sticky="w")

            ttk.Entry(panel, textvariable=var_obs, width=26).grid(row=idx, column=6, padx=3, pady=3, sticky="w")
            ttk.Button(panel, text="Guardar estado", command=lambda c=columna: self._save_estado_columna(c)).grid(
                row=idx, column=7, padx=3, pady=3, sticky="w"
            )

            lbl_tag = tk.Label(
                panel,
                text="En operación",
                width=14,
                bg="#D1FAE5",
                fg="#065F46",
                relief="groove",
                bd=1,
                padx=4,
                pady=2,
            )
            lbl_tag.grid(row=idx, column=8, padx=3, pady=3, sticky="w")

            self.column_rows[columna] = {
                "var_estado": var_estado,
                "var_fecha": var_fecha,
                "var_hora": var_hora,
                "var_dureza_salida": var_dureza_salida,
                "var_dureza_post": var_dureza_post,
                "var_obs": var_obs,
                "ent_dureza_salida": ent_salida,
                "ent_dureza_post": ent_post,
                "lbl_tag": lbl_tag,
            }

    def validate_dureza_ppm(self, value: str, required: bool = False) -> Tuple[bool, Optional[float]]:
        v = (value or "").strip()
        if not v:
            if required:
                return False, None
            return True, None
        if not is_float_ok(v):
            return False, None
        try:
            return True, float(v.replace(",", "."))
        except Exception:
            return False, None

    def _state_visuals(self, estado: str) -> Tuple[str, str]:
        if estado == "Regenerada":
            return "#DBEAFE", "#1E3A8A"
        if estado == "Por regenerar":
            return "#FEF3C7", "#92400E"
        return "#D1FAE5", "#065F46"

    def _apply_estado_visual(self, columna: int) -> None:
        row = self.column_rows[columna]
        estado = row["var_estado"].get().strip() or "En operación"
        bg, fg = self._state_visuals(estado)
        row["lbl_tag"].configure(text=estado, bg=bg, fg=fg)
        if estado == "Regenerada":
            row["ent_dureza_post"].configure(style="TEntry")

    def _on_estado_columna_changed(self, columna: int) -> None:
        self._apply_estado_visual(columna)

    def refresh_columnas_intercambio_panel(self) -> None:
        latest = self.db.get_latest_estado_columnas()
        for columna in (1, 2, 3):
            row = self.column_rows[columna]
            data = latest.get(columna, {})
            row["var_estado"].set(data.get("estado", "En operación"))
            row["var_fecha"].set(data.get("fecha_regeneracion", ""))
            row["var_hora"].set(data.get("hora_regeneracion", ""))
            ds = data.get("dureza_salida_ppm")
            dp = data.get("dureza_post_regeneracion_ppm")
            row["var_dureza_salida"].set("" if ds is None else fmt_num(ds))
            row["var_dureza_post"].set("" if dp is None else fmt_num(dp))
            row["var_obs"].set(data.get("observaciones", ""))
            self._apply_estado_visual(columna)

    def _save_estado_columna(self, columna: int) -> None:
        row = self.column_rows[columna]
        estado = (row["var_estado"].get() or "").strip()
        if estado not in self.COLUMN_STATES:
            messagebox.showerror("Estado de columnas", f"Estado inválido para columna {columna}.")
            return

        req_reg = estado == "Regenerada"
        ok_s, dureza_salida = self.validate_dureza_ppm(row["var_dureza_salida"].get(), required=req_reg)
        ok_p, dureza_post = self.validate_dureza_ppm(row["var_dureza_post"].get(), required=req_reg)
        if not ok_s:
            messagebox.showerror("Estado de columnas", f"Columna {columna}: Dureza salida inválida.")
            return
        if not ok_p:
            messagebox.showerror("Estado de columnas", f"Columna {columna}: Dureza post regeneración inválida.")
            return

        fecha_reg = (row["var_fecha"].get() or "").strip() or None
        hora_reg = (row["var_hora"].get() or "").strip() or None

        if estado == "Regenerada":
            now = datetime.now()
            fecha_reg = now.strftime("%d/%m/%Y")
            hora_reg = now.strftime("%H:%M")
            # Reflejar inmediatamente en UI (aunque falle refresco por DB/historial)
            row["var_fecha"].set(fecha_reg)
            row["var_hora"].set(hora_reg)

        try:
            self.db.save_estado_columna(
                columna_numero=columna,
                estado=estado,
                fecha_regeneracion=fecha_reg,
                hora_regeneracion=hora_reg,
                dureza_salida_ppm=dureza_salida,
                dureza_post_regeneracion_ppm=dureza_post,
                observaciones=row["var_obs"].get(),
            )
            self.refresh_columnas_intercambio_panel()
        except Exception as e:
            messagebox.showerror("Estado de columnas", f"No se pudo guardar columna {columna}.\n\nDetalle: {e}")

    def _add_operador(self) -> None:
        name = add_operador_via_dialog(self, self.db)
        if name is None:
            return
        self.operadores = self.db.fetch_operadores()
        if self.cbo_operador is not None:
            self.cbo_operador["values"] = self.operadores
        self.var_operador.set(name)

    def _validate_all(self) -> Tuple[bool, str]:
        op = (self.var_operador.get() or "").strip()
        if not op:
            return False, "Operador es obligatorio."
        if op not in self.operadores:
            return False, "Operador inválido. Seleccione uno del desplegable o agregue uno nuevo con +."

        lote = (self.var_lote.get() or "").strip()
        if not lote:
            return False, "Lote es obligatorio."

        # Turno se guarda para trazabilidad, pero NO limita la cantidad de registros.
        turno = (self.var_turno.get() or "").strip()
        if not turno:
            return False, "Turno es obligatorio."

        if not is_int_ok(self.var_num_columna.get()):
            return False, "Número de columna es obligatorio y debe ser entero."
        if not is_float_ok(self.var_temperatura.get()):
            return False, "Temperatura es obligatoria y debe ser numérica."
        if not is_float_ok(self.var_dureza.get()):
            return False, "Dureza es obligatoria y debe ser numérica."

        return True, "OK"

    def _apply_recommended_warnings(self) -> List[str]:
        msgs: List[str] = []

        def _warn_entry(entry: Optional[ttk.Entry], warn: bool) -> None:
            if entry is None:
                return
            try:
                entry.configure(style=(self._warn_entry_style if warn else "TEntry"))
            except Exception:
                pass

        dureza_s = (self.var_dureza.get() or "").strip()
        dureza_warn = False
        if is_float_ok(dureza_s):
            try:
                dureza = float(dureza_s.replace(",", "."))
                dureza_warn = dureza > 1
            except Exception:
                dureza_warn = False

        _warn_entry(getattr(self, "ent_dureza", None), dureza_warn)
        if dureza_warn:
            msgs.append("Dureza por encima del rango recomendado (max. 1 ppm).")

        return msgs

    def _refresh_save_state(self) -> None:
        if self.btn_guardar is None or self.lbl_status is None:
            return

        ok, msg = self._validate_all()
        if ok:
            warn_msgs = self._apply_recommended_warnings()
            if warn_msgs:
                self.lbl_status.config(
                    text=" | ".join(warn_msgs),
                    foreground=QDV_COLORS.get("warning", "#F59E0B"),
                    font=("Segoe UI", 10, "bold"),
                )
            else:
                self.lbl_status.config(text="")
        else:
            self.lbl_status.config(text=msg, foreground="red", font=("Segoe UI", 12, "bold"))
        self.btn_guardar.configure(state=("normal" if ok else "disabled"))

    def _load_day_table(self) -> None:
        if self.tree is None:
            return
        registros = self.db.fetch_agua_by_date(self.fixed_date_iso)
        for item in self.tree.get_children():
            self.tree.delete(item)
        for r in registros:
            row = [
                r["id"],
                iso_to_ddmmyyyy(r["fecha_iso"]),
                r["hora_hm"],
                r["operador"],
                r["lote"],
                r["numero_columna"],
                fmt_num(r["temperatura"]),
                fmt_num(r["dureza"]),
                (r.get("observaciones", "") or ""),
                r["created_at_iso"],
            ]
            self.tree.insert("", "end", values=row)
        self._selected_row_id = None

    def _on_tree_select(self, _event) -> None:
        if self.tree is None:
            return
        sel = self.tree.selection()
        if not sel:
            self._selected_row_id = None
            return
        vals = self.tree.item(sel[0], "values")
        if not vals:
            self._selected_row_id = None
            return
        try:
            self._selected_row_id = int(vals[0])
        except Exception:
            self._selected_row_id = None

    def on_delete_selected(self) -> None:
        if self._selected_row_id is None:
            messagebox.showwarning("Atención", "Seleccione un registro en la tabla para borrarlo.")
            return

        dlg = CodigoSeguridadDialog(self, title="Borrar registro - Código de seguridad")
        self.wait_window(dlg)
        if dlg.result is None:
            return
        if dlg.result != SECURITY_DELETE_CODE:
            messagebox.showerror("Código incorrecto", "Código de seguridad inválido. No se borró el registro.")
            return

        if not messagebox.askyesno("Confirmar", f"¿Confirma borrar el registro ID {self._selected_row_id}?"):
            return

        try:
            self.db.delete_agua_by_id(self._selected_row_id)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo borrar el registro.\n\nDetalle: {e}")
            return

        self._load_day_table()
        self._refresh_save_state()
        messagebox.showinfo("OK", f"Registro ID {self._selected_row_id} eliminado.")

    def on_save(self) -> None:
        try:
            ok, msg = self._validate_all()
            if not ok:
                messagebox.showerror("No se puede guardar", msg)
                return

            now = datetime.now()
            self.var_hora.set(now.strftime("%H:%M"))
            # Si el usuario deja la ventana abierta y cruza de turno/hora, recalcular turno
            self.var_turno.set(_compute_turno_from_hour(self.var_hora.get()))

            # Revalidar con el turno recalculado (para no saltarnos la regla 1 por turno).
            ok, msg = self._validate_all()
            if not ok:
                messagebox.showerror("No se puede guardar", msg)
                return

            data = {
                "fecha_iso": self.fixed_date_iso,
                "hora_hm": now.strftime("%H:%M"),
                "turno": self.var_turno.get().strip(),
                "operador": self.var_operador.get().strip(),
                "lote": self.var_lote.get().strip(),
                "numero_columna": int(self.var_num_columna.get().strip()),
                "temperatura": float(self.var_temperatura.get().strip().replace(",", ".")),
                "dureza": float(self.var_dureza.get().strip().replace(",", ".")),
                "observaciones": self.txt_obs.get("1.0", "end").strip() if self.txt_obs is not None else "",
                "created_at_iso": now.isoformat(timespec="seconds"),
            }

            self.db.insert_agua(data)

            self._load_day_table()
            self._clear_row()
            self._refresh_save_state()
            self.after(50, lambda: self.cbo_operador.focus_set())
            if self._timer is not None:
                self._timer.reset_from_created_at_iso(data["created_at_iso"])
            # Sin popup de éxito: carga rápida continua.
        except Exception as e:
            messagebox.showerror("Error al guardar", f"{e}")

    def destroy(self):
        try:
            if self._timer is not None:
                self._timer.stop()
        except Exception:
            pass
        super().destroy()

    def on_historial(self) -> None:
        if not OPENPYXL_OK:
            messagebox.showerror("Falta dependencia", "Para exportar a Excel instalá:\n\npip install openpyxl")
            return

        hoy_ddmmyyyy = datetime.now().strftime("%d/%m/%Y")
        dlg = HistorialDialog(self, default_desde=hoy_ddmmyyyy, default_hasta=hoy_ddmmyyyy)
        self.wait_window(dlg)
        if not dlg.result:
            return

        desde_s, hasta_s = dlg.result
        d1 = parse_date_ddmmyyyy(desde_s)
        d2 = parse_date_ddmmyyyy(hasta_s)
        if not d1 or not d2:
            messagebox.showerror("Error", "Rango de fechas inválido.")
            return

        desde_iso = date_to_iso(d1)
        hasta_iso = date_to_iso(d2)

        try:
            registros = self.db.fetch_agua_range(desde_iso, hasta_iso)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el historial.\n\nDetalle: {e}")
            return

        if not registros:
            messagebox.showinfo("Sin datos", "No hay registros en el rango seleccionado.")
            return

        default_name = f"historial_agua_{desde_iso}_a_{hasta_iso}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Guardar reporte Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not save_path:
            return

        try:
            self._export_excel(registros, save_path)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el Excel.\n\nDetalle: {e}")
            return

        messagebox.showinfo("OK", f"Reporte generado:\n{save_path}")

    def _export_excel(self, registros: List[Dict[str, Any]], path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Agua"

        headers = [
            "ID",
            "Fecha",
            "Hora",
            "Turno",
            "Operador",
            "Lote",
            "Columna",
            "Temperatura",
            "Dureza",
            "Observaciones",
            "Creado (ISO)",
        ]
        ws.append(headers)

        for r in registros:
            ws.append([
                r["id"],
                iso_to_ddmmyyyy(r["fecha_iso"]),
                r["hora_hm"],
                r["turno"],
                r["operador"],
                r["lote"],
                r["numero_columna"],
                r["temperatura"],
                r["dureza"],
                r.get("observaciones", ""),
                r["created_at_iso"],
            ])

        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter]:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)

        wb.save(path)

