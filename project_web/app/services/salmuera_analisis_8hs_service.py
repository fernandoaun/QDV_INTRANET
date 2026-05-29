from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from flask import current_app, has_request_context
from sqlalchemy import select

from app.extensions import db
from app.models import SalmueraAnalisis8hs
from app.services.upload_paths import resolve_under_upload_roots, uploads_workspace_root
from app.web.modules.produccion.operativa_context import compute_turno_from_hour

ANALISIS_8HS_INTERVAL_SECONDS = 8 * 60 * 60
ATTACHMENT_FIELDS = {
    "dureza": {
        "column": "file_dureza_path",
        "file_input": "file_dureza",
        "label": "Dureza de salmuera",
    },
    "cloro_libre": {
        "column": "file_cloro_libre_path",
        "file_input": "file_cloro_libre",
        "label": "Cloro libre en salmuera",
    },
}


def _parse_float_required(raw: str | None, label: str) -> float:
    s = (raw or "").strip().replace(",", ".")
    if not s:
        raise ValueError(f"{label} es obligatorio.")
    try:
        v = float(s)
    except ValueError as exc:
        raise ValueError(f"{label} debe ser numérico.") from exc
    if v != v:
        raise ValueError(f"{label} debe ser numérico.")
    return v


def _attachment_relative_path(row_id: int, field: str, stored_filename: str) -> Path:
    return Path("salmuera_analisis_8hs") / str(int(row_id)) / field / stored_filename


def _attachment_storage_dir(row_id: int, field: str) -> Path:
    p = uploads_workspace_root() / "salmuera_analisis_8hs" / str(int(row_id)) / field
    p.mkdir(parents=True, exist_ok=True)
    return p


def _safe_original_name(filename: str) -> str:
    base = Path(filename).name.strip()
    return (base or "analisis.pdf")[:200]


def _validate_attachment_upload(fs: Any, max_bytes: int) -> tuple[str, str]:
    if fs is None or not getattr(fs, "filename", None):
        raise ValueError("Seleccioná un archivo PDF.")
    original = _safe_original_name(str(fs.filename or ""))
    ext = Path(original).suffix.lower()
    if ext != ".pdf":
        raise ValueError("Solo se permiten archivos PDF.")
    if max_bytes <= 0:
        max_bytes = 15 * 1024 * 1024
    fs.seek(0)
    blob = fs.read(max_bytes + 1)
    fs.seek(0)
    if len(blob) > max_bytes:
        raise ValueError("El archivo PDF supera el tamaño máximo permitido.")
    content_type = (getattr(fs, "content_type", None) or "").lower()
    head = fs.read(16)
    fs.seek(0)
    if content_type and "pdf" not in content_type:
        raise ValueError("El archivo no es un PDF válido (tipo MIME).")
    if len(head) < 4 or head[:4] != b"%PDF":
        raise ValueError("El contenido no es un PDF válido.")
    return original, ".pdf"


def attachment_resolve_path(row: SalmueraAnalisis8hs, field: str) -> Path | None:
    meta = ATTACHMENT_FIELDS.get(field)
    if meta is None:
        raise ValueError("Campo de adjunto inválido.")
    raw = (getattr(row, str(meta["column"])) or "").strip()
    if not raw:
        return None
    return resolve_under_upload_roots(Path(raw))


def attachment_exists(row: SalmueraAnalisis8hs, field: str) -> bool:
    return attachment_resolve_path(row, field) is not None


def save_attachment(row: SalmueraAnalisis8hs, field: str, fs: Any) -> None:
    meta = ATTACHMENT_FIELDS.get(field)
    if meta is None:
        raise ValueError("Campo de adjunto inválido.")
    if fs is None or not getattr(fs, "filename", None):
        return
    mx = 15 * 1024 * 1024
    if has_request_context():
        try:
            mx = int(current_app.config.get("SALMUERA_ANALISIS_8HS_PDF_MAX_BYTES") or mx)
        except (TypeError, ValueError):
            mx = 15 * 1024 * 1024
    _original, ext = _validate_attachment_upload(fs, mx)
    old = attachment_resolve_path(row, field)
    stored = f"{uuid4().hex}{ext}"
    dest = _attachment_storage_dir(int(row.id), field) / stored
    rel = _attachment_relative_path(int(row.id), field, stored)
    try:
        fs.save(str(dest))
    except Exception:
        try:
            dest.unlink(missing_ok=True)
        except OSError:
            pass
        raise
    if old is not None:
        try:
            old.unlink(missing_ok=True)
        except OSError:
            pass
    setattr(row, str(meta["column"]), rel.as_posix())


def delete_attachment(row: SalmueraAnalisis8hs, field: str) -> bool:
    meta = ATTACHMENT_FIELDS.get(field)
    if meta is None:
        raise ValueError("Campo de adjunto inválido.")
    fp = attachment_resolve_path(row, field)
    if fp is not None:
        try:
            fp.unlink(missing_ok=True)
        except OSError:
            pass
    had_value = bool((getattr(row, str(meta["column"])) or "").strip())
    setattr(row, str(meta["column"]), None)
    return had_value


def row_to_dict(row: SalmueraAnalisis8hs) -> dict[str, Any]:
    dureza_has_path = bool((row.file_dureza_path or "").strip())
    cloro_has_path = bool((row.file_cloro_libre_path or "").strip())
    dureza_exists = attachment_exists(row, "dureza")
    cloro_exists = attachment_exists(row, "cloro_libre")
    return {
        "id": int(row.id),
        "fecha": row.fecha,
        "hora": row.hora,
        "fecha_hora_iso": row.fecha_hora_iso,
        "turno": row.turno,
        "operador": row.operador,
        "dureza_salmuera": float(row.dureza_salmuera),
        "cloro_libre_salmuera": float(row.cloro_libre_salmuera),
        "observaciones": row.observaciones or "",
        "file_dureza_path": row.file_dureza_path or "",
        "file_cloro_libre_path": row.file_cloro_libre_path or "",
        "file_dureza_present": dureza_has_path and dureza_exists,
        "file_cloro_libre_present": cloro_has_path and cloro_exists,
        "file_dureza_missing": dureza_has_path and not dureza_exists,
        "file_cloro_libre_missing": cloro_has_path and not cloro_exists,
        "created_at_iso": row.created_at_iso,
    }


def create_from_form(form: Any, *, now: datetime, operador: str, files: Any | None = None) -> SalmueraAnalisis8hs:
    dureza = _parse_float_required(form.get("dureza_salmuera"), "Dureza de salmuera")
    cloro = _parse_float_required(form.get("cloro_libre_salmuera"), "Cloro libre en salmuera")
    fecha = now.strftime("%Y-%m-%d")
    hora = now.strftime("%H:%M")
    iso = now.isoformat(timespec="seconds")
    row = SalmueraAnalisis8hs(
        fecha=fecha,
        hora=hora,
        fecha_hora_iso=iso,
        turno=compute_turno_from_hour(hora),
        operador=(operador or "").strip(),
        dureza_salmuera=dureza,
        cloro_libre_salmuera=cloro,
        observaciones=(form.get("observaciones") or "").strip() or None,
        created_at_iso=iso,
    )
    db.session.add(row)
    db.session.flush()
    if files is not None:
        for field, meta in ATTACHMENT_FIELDS.items():
            save_attachment(row, field, files.get(str(meta["file_input"])))
    return row


def latest_row() -> SalmueraAnalisis8hs | None:
    return db.session.scalar(
        select(SalmueraAnalisis8hs).order_by(SalmueraAnalisis8hs.fecha_hora_iso.desc(), SalmueraAnalisis8hs.id.desc()).limit(1)
    )


def build_status(now: datetime) -> dict[str, Any]:
    from app.services import plant_stop_service as ps

    row = latest_row()
    if row is None:
        plant_stop = ps.analisis8_plant_stop_overlay(
            last_fecha_hora_iso=None,
            interval_sec=int(ANALISIS_8HS_INTERVAL_SECONDS),
        )
        if plant_stop.get("active"):
            frozen = int(plant_stop.get("frozen_remaining_sec") or 0)
            return {
                "has_records": False,
                "last": None,
                "next_due_iso": None,
                "remaining_seconds": frozen,
                "is_due": False,
                "message": f"Parada de planta desde {plant_stop.get('started_at_iso') or '—'}.",
                "plant_stop": plant_stop,
            }
        return {
            "has_records": False,
            "last": None,
            "next_due_iso": None,
            "remaining_seconds": None,
            "is_due": True,
            "message": "No hay análisis registrados. Realizar primer análisis.",
            "plant_stop": plant_stop,
        }
    try:
        last_dt = datetime.fromisoformat(str(row.fecha_hora_iso))
    except ValueError:
        last_dt = now
    anchor_iso = str(row.fecha_hora_iso or row.created_at_iso or "")
    pause_extra = ps.pause_seconds_after_anchor(anchor_iso, ps.CIRCUIT_REACTOR, now_iso=now.isoformat(timespec="seconds"))
    next_due = last_dt + timedelta(seconds=ANALISIS_8HS_INTERVAL_SECONDS) + timedelta(seconds=pause_extra)
    remaining = int((next_due - now).total_seconds())
    plant_stop = ps.analisis8_plant_stop_overlay(
        last_fecha_hora_iso=anchor_iso,
        interval_sec=int(ANALISIS_8HS_INTERVAL_SECONDS),
    )
    if plant_stop.get("active"):
        frozen = int(plant_stop.get("frozen_remaining_sec") or max(0, remaining))
        return {
            "has_records": True,
            "last": row_to_dict(row),
            "next_due_iso": next_due.isoformat(timespec="seconds"),
            "remaining_seconds": frozen,
            "is_due": False,
            "message": f"Parada de planta: cronómetro detenido desde {plant_stop.get('started_at_iso') or '—'}.",
            "plant_stop": plant_stop,
        }
    return {
        "has_records": True,
        "last": row_to_dict(row),
        "next_due_iso": next_due.isoformat(timespec="seconds"),
        "remaining_seconds": remaining,
        "is_due": remaining <= 0,
        "message": (
            "Análisis de salmuera vencido. Registrar dureza y cloro libre."
            if remaining <= 0
            else "Próximo análisis programado."
        ),
        "plant_stop": plant_stop,
    }


def filtered_rows(fecha_desde: str = "", fecha_hasta: str = "", *, limit: int | None = 1000) -> list[SalmueraAnalisis8hs]:
    desde = (fecha_desde or "").strip()
    hasta = (fecha_hasta or "").strip()
    if desde and hasta and desde > hasta:
        desde, hasta = hasta, desde
    q = select(SalmueraAnalisis8hs)
    if desde:
        q = q.where(SalmueraAnalisis8hs.fecha >= desde)
    if hasta:
        q = q.where(SalmueraAnalisis8hs.fecha <= hasta)
    q = q.order_by(SalmueraAnalisis8hs.fecha_hora_iso.desc(), SalmueraAnalisis8hs.id.desc())
    if limit is not None:
        q = q.limit(int(limit))
    return list(db.session.scalars(q).all())


def export_excel(fecha_desde: str = "", fecha_hasta: str = "") -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    headers = [
        "Fecha",
        "Hora",
        "Turno",
        "Operador",
        "Dureza de salmuera",
        "Cloro libre en salmuera",
        "Archivo dureza",
        "Archivo cloro libre",
        "Observaciones",
        "Fecha/hora ISO",
        "Creado",
    ]
    rows = filtered_rows(fecha_desde, fecha_hasta, limit=None)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Analisis salmuera 8hs"
    bold = Font(bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start=2):
        values = [
            row.fecha,
            row.hora,
            row.turno,
            row.operador,
            float(row.dureza_salmuera),
            float(row.cloro_libre_salmuera),
            row.file_dureza_path or "",
            row.file_cloro_libre_path or "",
            row.observaciones or "",
            row.fecha_hora_iso,
            row.created_at_iso,
        ]
        for c_idx, value in enumerate(values, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value).alignment = Alignment(vertical="top", wrap_text=False)
    max_row = 1 + len(rows)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row}"
    for c_idx in range(1, len(headers) + 1):
        letter = get_column_letter(c_idx)
        max_len = 10
        for r_idx in range(1, min(max_row, 500) + 1):
            v = ws.cell(row=r_idx, column=c_idx).value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[letter].width = min(max_len + 2, 55)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
