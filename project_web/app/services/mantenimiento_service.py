from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from statistics import mean, pstdev
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.orm import selectinload
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from app.auth_utils import user_display_name
from app.extensions import db
from app.models import (
    Equipo,
    MaintenanceAttachment,
    MaintenanceComponent,
    MaintenanceFailure,
    MaintenanceOrder,
    MaintenanceOrderResource,
    MaintenancePlan,
    MaintenancePrediction,
    MaintenanceResource,
    User,
)
from app.services.upload_paths import resolve_under_upload_roots, uploads_workspace_root
from app.utils.datetime_operacion import now_operacion_local_iso_seconds

EQUIPO_ESTADOS: dict[str, str] = {
    "operativo": "Operativo",
    "fuera_de_servicio": "Fuera de servicio",
    "en_mantenimiento": "En mantenimiento",
    "dado_de_baja": "Dado de baja",
}

FAILURE_ESTADOS: dict[str, str] = {
    "reportado": "Reportado",
    "en_analisis": "En análisis",
    "programado": "Programado",
    "en_ejecucion": "En ejecución",
    "finalizado": "Finalizado",
}

CRITICIDADES: dict[str, str] = {
    "baja": "Baja",
    "media": "Media",
    "alta": "Alta",
    "critica": "Crítica",
}

TIPOS_MANTENIMIENTO: dict[str, str] = {
    "preventivo": "Preventivo",
    "correctivo": "Correctivo",
    "predictivo": "Predictivo",
}

ORDER_ESTADOS: dict[str, str] = {
    "pendiente": "Pendiente",
    "programado": "Programado",
    "en_ejecucion": "En ejecución",
    "finalizado": "Finalizado",
    "cancelado": "Cancelado",
    "reprogramado": "Reprogramado",
}

PRIORIDADES: dict[str, str] = {
    "baja": "Baja",
    "media": "Media",
    "alta": "Alta",
    "critica": "Crítica",
}

FRECUENCIA_PERIODOS: dict[str, str] = {
    "mensual": "Mensual",
    "trimestral": "Trimestral",
    "semestral": "Semestral",
    "anual": "Anual",
}

RESOURCE_CATEGORIAS: dict[str, str] = {
    "personal": "Personal requerido",
    "repuesto": "Repuestos",
    "herramienta": "Herramientas",
    "insumo": "Insumos",
    "epp": "EPP",
    "servicio_externo": "Servicios externos",
}

PREDICTION_ESTADOS: dict[str, str] = {
    "sugerida": "Sugerida",
    "programada": "Programada",
    "descartada": "Descartada",
}


@dataclass(frozen=True)
class CorrectivoFiltros:
    desde: str = ""
    hasta: str = ""
    equipo_id: int | None = None
    estado: str = ""
    responsable: str = ""
    criticidad: str = ""


@dataclass(frozen=True)
class OrdenFiltros:
    desde: str = ""
    hasta: str = ""
    equipo_id: int | None = None
    tipo_mantenimiento: str = ""
    estado: str = ""
    responsable: str = ""
    criticidad: str = ""


def labels_context() -> dict[str, dict[str, str]]:
    return {
        "equipo_estados": EQUIPO_ESTADOS,
        "failure_estados": FAILURE_ESTADOS,
        "criticidades": CRITICIDADES,
        "tipos_mantenimiento": TIPOS_MANTENIMIENTO,
        "order_estados": ORDER_ESTADOS,
        "prioridades": PRIORIDADES,
        "frecuencia_periodos": FRECUENCIA_PERIODOS,
        "resource_categorias": RESOURCE_CATEGORIAS,
        "prediction_estados": PREDICTION_ESTADOS,
    }


def _clean(raw: str | None) -> str:
    return (raw or "").strip()


def _none_if_empty(raw: str | None) -> str | None:
    s = _clean(raw)
    return s or None


def _valid_or_default(raw: str | None, allowed: dict[str, str], default: str) -> str:
    s = _clean(raw)
    return s if s in allowed else default


def _parse_int(raw: str | None) -> int | None:
    s = _clean(raw)
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def _parse_float(raw: str | None) -> float | None:
    s = _clean(raw)
    if not s:
        return None
    try:
        return max(float(s.replace(",", ".")), 0.0)
    except ValueError:
        return None


def _normalize_date(raw: str | None) -> str | None:
    s = _clean(raw)
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10]).isoformat()
    except ValueError:
        return None


def _normalize_datetime_local(raw: str | None) -> str:
    s = _clean(raw)
    if not s:
        return now_operacion_local_iso_seconds()
    if "T" not in s:
        return f"{s}T00:00:00"
    if len(s) == 16:
        return f"{s}:00"
    return s[:32]


def _parse_iso_dt(raw: str | None) -> datetime | None:
    s = _clean(raw)
    if not s:
        return None
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        return None


def input_datetime_value(raw: str | None) -> str:
    s = _clean(raw)
    return s[:16] if len(s) >= 16 else s


def current_datetime_input_value() -> str:
    return input_datetime_value(now_operacion_local_iso_seconds())


def format_dt(raw: str | None) -> str:
    s = _clean(raw)
    if not s:
        return ""
    return s.replace("T", " ")[:16]


def list_equipos() -> list[Equipo]:
    return list(db.session.scalars(select(Equipo).order_by(Equipo.nombre_equipo.asc(), Equipo.id.asc())).all())


def list_equipos_activos() -> list[Equipo]:
    return list(
        db.session.scalars(
            select(Equipo).where(Equipo.activo.is_(True)).order_by(Equipo.nombre_equipo.asc(), Equipo.id.asc())
        ).all()
    )


def list_components_for_equipo(equipo_id: int) -> list[MaintenanceComponent]:
    return list(
        db.session.scalars(
            select(MaintenanceComponent)
            .where(MaintenanceComponent.equipo_principal_id == equipo_id)
            .order_by(MaintenanceComponent.nombre.asc(), MaintenanceComponent.id.asc())
        ).all()
    )


def list_users_for_responsable() -> list[User]:
    return list(db.session.scalars(select(User).where(User.activo.is_(True)).order_by(User.username.asc())).all())


def update_equipo_from_form(equipo: Equipo, form: Any) -> None:
    equipo.codigo_interno = _none_if_empty(form.get("codigo_interno"))
    equipo.nombre_equipo = _clean(form.get("nombre_equipo")) or equipo.nombre_equipo
    equipo.descripcion = _clean(form.get("descripcion"))
    equipo.tipo_equipo = _none_if_empty(form.get("tipo_equipo"))
    equipo.area_sector = _none_if_empty(form.get("area_sector"))
    equipo.equipo_principal_id = _parse_int(form.get("equipo_principal_id"))
    if equipo.equipo_principal_id == equipo.id:
        equipo.equipo_principal_id = None
    equipo.marca = _none_if_empty(form.get("marca"))
    equipo.modelo = _none_if_empty(form.get("modelo"))
    equipo.numero_serie = _none_if_empty(form.get("numero_serie"))
    equipo.fecha_alta = _none_if_empty(form.get("fecha_alta"))
    equipo.estado = _valid_or_default(form.get("estado"), EQUIPO_ESTADOS, "operativo")
    equipo.observaciones = _none_if_empty(form.get("observaciones"))
    equipo.activo = equipo.estado != "dado_de_baja"


def create_component_from_form(equipo: Equipo, form: Any) -> MaintenanceComponent:
    now = now_operacion_local_iso_seconds()
    component = MaintenanceComponent(
        equipo_principal_id=int(equipo.id),
        codigo_interno=_none_if_empty(form.get("codigo_interno")),
        nombre=_clean(form.get("nombre")),
        tipo_componente=_none_if_empty(form.get("tipo_componente")),
        marca=_none_if_empty(form.get("marca")),
        modelo=_none_if_empty(form.get("modelo")),
        numero_serie=_none_if_empty(form.get("numero_serie")),
        estado=_valid_or_default(form.get("estado"), EQUIPO_ESTADOS, "operativo"),
        observaciones=_none_if_empty(form.get("observaciones")),
        created_at_iso=now,
        updated_at_iso=now,
    )
    if not component.nombre:
        raise ValueError("El nombre del componente es obligatorio.")
    db.session.add(component)
    return component


def parse_correctivo_filtros(args: Any) -> CorrectivoFiltros:
    return CorrectivoFiltros(
        desde=_clean(args.get("desde")),
        hasta=_clean(args.get("hasta")),
        equipo_id=_parse_int(args.get("equipo_id")),
        estado=_clean(args.get("estado")),
        responsable=_clean(args.get("responsable")),
        criticidad=_clean(args.get("criticidad")),
    )


def list_failures(filtros: CorrectivoFiltros | None = None, *, limit: int | None = 500) -> list[MaintenanceFailure]:
    filtros = filtros or CorrectivoFiltros()
    stmt = (
        select(MaintenanceFailure)
        .options(
            selectinload(MaintenanceFailure.equipo),
            selectinload(MaintenanceFailure.component),
            selectinload(MaintenanceFailure.attachments),
        )
        .order_by(MaintenanceFailure.detected_at_iso.desc(), MaintenanceFailure.id.desc())
    )
    if filtros.desde:
        stmt = stmt.where(MaintenanceFailure.detected_at_iso >= f"{filtros.desde}T00:00:00")
    if filtros.hasta:
        stmt = stmt.where(MaintenanceFailure.detected_at_iso <= f"{filtros.hasta}T23:59:59")
    if filtros.equipo_id:
        stmt = stmt.where(MaintenanceFailure.equipo_id == filtros.equipo_id)
    if filtros.estado in FAILURE_ESTADOS:
        stmt = stmt.where(MaintenanceFailure.estado == filtros.estado)
    if filtros.criticidad in CRITICIDADES:
        stmt = stmt.where(MaintenanceFailure.criticidad == filtros.criticidad)
    if filtros.responsable:
        stmt = stmt.where(MaintenanceFailure.responsable_trabajo.ilike(f"%{filtros.responsable}%"))
    if limit is not None:
        stmt = stmt.limit(limit)
    return list(db.session.scalars(stmt).all())


def create_failure_from_form(form: Any, user: User | None) -> MaintenanceFailure:
    equipo_id = _parse_int(form.get("equipo_id"))
    if equipo_id is None:
        raise ValueError("Seleccioná el equipo principal.")
    equipo = db.session.get(Equipo, equipo_id)
    if equipo is None:
        raise ValueError("El equipo seleccionado no existe.")
    component_id = _parse_int(form.get("component_id"))
    if component_id is not None:
        component = db.session.get(MaintenanceComponent, component_id)
        if component is None or int(component.equipo_principal_id) != int(equipo.id):
            raise ValueError("El componente seleccionado no corresponde al equipo.")

    now = now_operacion_local_iso_seconds()
    failure = MaintenanceFailure(
        detected_at_iso=_normalize_datetime_local(form.get("detected_at_iso")),
        equipo_id=int(equipo.id),
        component_id=component_id,
        reported_by_user_id=int(user.id) if user is not None else None,
        reported_by_display=user_display_name(user),
        descripcion_falla=_clean(form.get("descripcion_falla")),
        sintoma_observado=_none_if_empty(form.get("sintoma_observado")),
        causa_probable=_none_if_empty(form.get("causa_probable")),
        criticidad=_valid_or_default(form.get("criticidad"), CRITICIDADES, "media"),
        estado="reportado",
        responsable_trabajo=_none_if_empty(form.get("responsable_trabajo")),
        observaciones=_none_if_empty(form.get("observaciones")),
        created_at_iso=now,
        updated_at_iso=now,
    )
    if not failure.descripcion_falla:
        raise ValueError("La descripción de la falla es obligatoria.")
    db.session.add(failure)
    return failure


def update_failure_from_form(failure: MaintenanceFailure, form: Any) -> None:
    failure.estado = _valid_or_default(form.get("estado"), FAILURE_ESTADOS, failure.estado or "reportado")
    failure.criticidad = _valid_or_default(form.get("criticidad"), CRITICIDADES, failure.criticidad or "media")
    failure.descripcion_falla = _clean(form.get("descripcion_falla")) or failure.descripcion_falla
    failure.sintoma_observado = _none_if_empty(form.get("sintoma_observado"))
    failure.causa_probable = _none_if_empty(form.get("causa_probable"))
    failure.causa_real = _none_if_empty(form.get("causa_real"))
    failure.accion_realizada = _none_if_empty(form.get("accion_realizada"))
    failure.repuestos_utilizados = _none_if_empty(form.get("repuestos_utilizados"))
    failure.recursos_utilizados = _none_if_empty(form.get("recursos_utilizados"))
    failure.responsable_trabajo = _none_if_empty(form.get("responsable_trabajo"))
    failure.observaciones = _none_if_empty(form.get("observaciones"))
    failure.closed_at_iso = _normalize_datetime_local(form.get("closed_at_iso")) if form.get("closed_at_iso") else None
    manual_downtime = _none_if_empty(form.get("tiempo_fuera_servicio_horas"))
    if manual_downtime is not None:
        try:
            failure.tiempo_fuera_servicio_horas = max(float(manual_downtime.replace(",", ".")), 0.0)
        except ValueError:
            failure.tiempo_fuera_servicio_horas = None
    elif failure.closed_at_iso:
        start = _parse_iso_dt(failure.detected_at_iso)
        end = _parse_iso_dt(failure.closed_at_iso)
        if start is not None and end is not None and end >= start:
            failure.tiempo_fuera_servicio_horas = round((end - start).total_seconds() / 3600, 2)
    if failure.estado == "finalizado" and not failure.closed_at_iso:
        failure.closed_at_iso = now_operacion_local_iso_seconds()
    failure.updated_at_iso = now_operacion_local_iso_seconds()


def save_failure_attachment(failure: MaintenanceFailure, fs: FileStorage | None, user: User | None) -> MaintenanceAttachment | None:
    if fs is None or not fs.filename:
        return None
    safe = secure_filename(fs.filename)
    if not safe:
        raise ValueError("El nombre del archivo adjunto no es válido.")
    now = now_operacion_local_iso_seconds()
    rel_dir = Path("maintenance") / "failures" / str(failure.id)
    target_dir = uploads_workspace_root() / rel_dir
    target_dir.mkdir(parents=True, exist_ok=True)
    stored_name = f"{now.replace(':', '').replace('-', '').replace('T', '_')}_{safe}"
    rel_path = rel_dir / stored_name
    fs.save(target_dir / stored_name)
    attachment = MaintenanceAttachment(
        failure_id=int(failure.id),
        equipo_id=int(failure.equipo_id),
        component_id=int(failure.component_id) if failure.component_id is not None else None,
        original_filename=fs.filename[:256],
        stored_path=rel_path.as_posix(),
        content_type=fs.mimetype,
        uploaded_by_user_id=int(user.id) if user is not None else None,
        created_at_iso=now,
    )
    db.session.add(attachment)
    return attachment


def resolve_attachment_path(attachment: MaintenanceAttachment) -> Path | None:
    return resolve_under_upload_roots(Path(attachment.stored_path))


def plan_status(plan: MaintenancePlan, today: date | None = None) -> str:
    if not plan.activo:
        return "inactivo"
    d = _normalize_date(plan.proxima_fecha)
    if d is None:
        return "sin_fecha"
    today = today or date.today()
    due = date.fromisoformat(d)
    if due < today:
        return "vencido"
    if due <= today + timedelta(days=7):
        return "proximo"
    return "al_dia"


def _component_belongs_to_equipo(component_id: int | None, equipo_id: int) -> bool:
    if component_id is None:
        return True
    c = db.session.get(MaintenanceComponent, component_id)
    return c is not None and int(c.equipo_principal_id) == int(equipo_id)


def list_plans() -> list[MaintenancePlan]:
    return list(
        db.session.scalars(
            select(MaintenancePlan)
            .options(selectinload(MaintenancePlan.equipo), selectinload(MaintenancePlan.component))
            .order_by(MaintenancePlan.activo.desc(), MaintenancePlan.proxima_fecha.asc(), MaintenancePlan.id.desc())
        ).all()
    )


def _apply_plan_form(plan: MaintenancePlan, form: Any) -> None:
    equipo_id = _parse_int(form.get("equipo_id"))
    if equipo_id is None or db.session.get(Equipo, equipo_id) is None:
        raise ValueError("Seleccioná un equipo válido.")
    component_id = _parse_int(form.get("component_id"))
    if not _component_belongs_to_equipo(component_id, equipo_id):
        raise ValueError("El componente seleccionado no corresponde al equipo.")
    plan.equipo_id = equipo_id
    plan.component_id = component_id
    plan.tipo_mantenimiento = _valid_or_default(form.get("tipo_mantenimiento"), TIPOS_MANTENIMIENTO, "preventivo")
    plan.nombre = _clean(form.get("nombre"))
    if not plan.nombre:
        raise ValueError("El nombre del plan es obligatorio.")
    plan.frecuencia_dias = _parse_int(form.get("frecuencia_dias"))
    plan.frecuencia_horas_uso = _parse_float(form.get("frecuencia_horas_uso"))
    periodo = _clean(form.get("frecuencia_periodo"))
    plan.frecuencia_periodo = periodo if periodo in FRECUENCIA_PERIODOS else None
    plan.proxima_fecha = _normalize_date(form.get("proxima_fecha"))
    plan.responsable = _none_if_empty(form.get("responsable"))
    plan.duracion_estimada_horas = _parse_float(form.get("duracion_estimada_horas"))
    plan.tareas = _none_if_empty(form.get("tareas"))
    plan.recursos_necesarios = _none_if_empty(form.get("recursos_necesarios"))
    plan.repuestos_necesarios = _none_if_empty(form.get("repuestos_necesarios"))
    plan.herramientas_necesarias = _none_if_empty(form.get("herramientas_necesarias"))
    plan.epp_necesarios = _none_if_empty(form.get("epp_necesarios"))
    plan.observaciones = _none_if_empty(form.get("observaciones"))
    plan.activo = form.get("activo", "1") == "1"
    plan.updated_at_iso = now_operacion_local_iso_seconds()


def create_plan_from_form(form: Any) -> MaintenancePlan:
    now = now_operacion_local_iso_seconds()
    plan = MaintenancePlan(created_at_iso=now, updated_at_iso=now, equipo_id=0, nombre="")
    _apply_plan_form(plan, form)
    db.session.add(plan)
    return plan


def update_plan_from_form(plan: MaintenancePlan, form: Any) -> None:
    _apply_plan_form(plan, form)


def parse_order_filtros(args: Any) -> OrdenFiltros:
    return OrdenFiltros(
        desde=_clean(args.get("desde")),
        hasta=_clean(args.get("hasta")),
        equipo_id=_parse_int(args.get("equipo_id")),
        tipo_mantenimiento=_clean(args.get("tipo_mantenimiento")),
        estado=_clean(args.get("estado")),
        responsable=_clean(args.get("responsable")),
        criticidad=_clean(args.get("criticidad")),
    )


def list_orders(filtros: OrdenFiltros | None = None, *, limit: int | None = 500) -> list[MaintenanceOrder]:
    filtros = filtros or OrdenFiltros()
    stmt = (
        select(MaintenanceOrder)
        .options(
            selectinload(MaintenanceOrder.equipo),
            selectinload(MaintenanceOrder.component),
            selectinload(MaintenanceOrder.plan),
            selectinload(MaintenanceOrder.order_resources),
        )
        .order_by(MaintenanceOrder.fecha_programada.asc(), MaintenanceOrder.id.desc())
    )
    if filtros.desde:
        stmt = stmt.where(MaintenanceOrder.fecha_programada >= filtros.desde)
    if filtros.hasta:
        stmt = stmt.where(MaintenanceOrder.fecha_programada <= filtros.hasta)
    if filtros.equipo_id:
        stmt = stmt.where(MaintenanceOrder.equipo_id == filtros.equipo_id)
    if filtros.tipo_mantenimiento in TIPOS_MANTENIMIENTO:
        stmt = stmt.where(MaintenanceOrder.tipo_mantenimiento == filtros.tipo_mantenimiento)
    if filtros.estado in ORDER_ESTADOS:
        stmt = stmt.where(MaintenanceOrder.estado == filtros.estado)
    if filtros.criticidad in CRITICIDADES:
        stmt = stmt.where(MaintenanceOrder.criticidad == filtros.criticidad)
    if filtros.responsable:
        stmt = stmt.where(MaintenanceOrder.responsable.ilike(f"%{filtros.responsable}%"))
    if limit is not None:
        stmt = stmt.limit(limit)
    return list(db.session.scalars(stmt).all())


def suggested_resources(equipo_id: int, component_id: int | None, tipo_mantenimiento: str) -> list[MaintenanceResource]:
    stmt = (
        select(MaintenanceResource)
        .where(
            MaintenanceResource.activo.is_(True),
            MaintenanceResource.tipo_mantenimiento == tipo_mantenimiento,
            (MaintenanceResource.equipo_id.is_(None)) | (MaintenanceResource.equipo_id == equipo_id),
        )
        .order_by(MaintenanceResource.categoria.asc(), MaintenanceResource.nombre.asc())
    )
    rows = list(db.session.scalars(stmt).all())
    return [r for r in rows if r.component_id is None or (component_id is not None and int(r.component_id) == int(component_id))]


def copy_suggested_resources_to_order(order: MaintenanceOrder) -> None:
    for r in suggested_resources(int(order.equipo_id), order.component_id, order.tipo_mantenimiento):
        db.session.add(
            MaintenanceOrderResource(
                order_id=int(order.id),
                resource_id=int(r.id),
                categoria=r.categoria,
                nombre=r.nombre,
                cantidad=r.cantidad,
                unidad=r.unidad,
                tiempo_estimado_horas=r.tiempo_estimado_horas,
                observaciones=r.observaciones,
                created_at_iso=now_operacion_local_iso_seconds(),
            )
        )


def _apply_order_form(order: MaintenanceOrder, form: Any) -> None:
    equipo_id = _parse_int(form.get("equipo_id"))
    if equipo_id is None or db.session.get(Equipo, equipo_id) is None:
        raise ValueError("Seleccioná un equipo válido.")
    component_id = _parse_int(form.get("component_id"))
    if not _component_belongs_to_equipo(component_id, equipo_id):
        raise ValueError("El componente seleccionado no corresponde al equipo.")
    fecha = _normalize_date(form.get("fecha_programada"))
    if fecha is None:
        raise ValueError("Indicá una fecha programada válida.")
    order.equipo_id = equipo_id
    order.component_id = component_id
    order.tipo_mantenimiento = _valid_or_default(form.get("tipo_mantenimiento"), TIPOS_MANTENIMIENTO, "preventivo")
    order.fecha_programada = fecha
    order.prioridad = _valid_or_default(form.get("prioridad"), PRIORIDADES, "media")
    order.criticidad = _valid_or_default(form.get("criticidad"), CRITICIDADES, "media")
    order.responsable = _none_if_empty(form.get("responsable"))
    order.estado = _valid_or_default(form.get("estado"), ORDER_ESTADOS, "programado")
    order.tareas = _none_if_empty(form.get("tareas"))
    order.recursos_necesarios = _none_if_empty(form.get("recursos_necesarios"))
    order.repuestos_necesarios = _none_if_empty(form.get("repuestos_necesarios"))
    order.herramientas_necesarias = _none_if_empty(form.get("herramientas_necesarias"))
    order.epp_necesarios = _none_if_empty(form.get("epp_necesarios"))
    order.tiempo_estimado_horas = _parse_float(form.get("tiempo_estimado_horas"))
    order.observaciones = _none_if_empty(form.get("observaciones"))
    order.resultado = _none_if_empty(form.get("resultado"))
    if order.estado == "en_ejecucion" and not order.executed_at_iso:
        order.executed_at_iso = now_operacion_local_iso_seconds()
    if order.estado == "finalizado" and not order.closed_at_iso:
        order.closed_at_iso = now_operacion_local_iso_seconds()
    order.updated_at_iso = now_operacion_local_iso_seconds()


def _form_getlist(form: Any, key: str) -> list[str]:
    if hasattr(form, "getlist"):
        return list(form.getlist(key))
    value = form.get(key)
    if value is None:
        return []
    return value if isinstance(value, list) else [value]


def _apply_order_resources_from_form(order: MaintenanceOrder, form: Any) -> None:
    nombres = _form_getlist(form, "resource_nombre")
    categorias = _form_getlist(form, "resource_categoria")
    cantidades = _form_getlist(form, "resource_cantidad")
    unidades = _form_getlist(form, "resource_unidad")
    tiempos = _form_getlist(form, "resource_tiempo_estimado_horas")
    observaciones = _form_getlist(form, "resource_observaciones")

    db.session.query(MaintenanceOrderResource).filter_by(order_id=int(order.id)).delete()
    now = now_operacion_local_iso_seconds()
    for idx, raw_nombre in enumerate(nombres):
        nombre = _clean(raw_nombre)
        if not nombre:
            continue
        db.session.add(
            MaintenanceOrderResource(
                order_id=int(order.id),
                resource_id=None,
                categoria=_valid_or_default(categorias[idx] if idx < len(categorias) else None, RESOURCE_CATEGORIAS, "personal"),
                nombre=nombre,
                cantidad=_parse_float(cantidades[idx] if idx < len(cantidades) else None),
                unidad=_none_if_empty(unidades[idx] if idx < len(unidades) else None),
                tiempo_estimado_horas=_parse_float(tiempos[idx] if idx < len(tiempos) else None),
                observaciones=_none_if_empty(observaciones[idx] if idx < len(observaciones) else None),
                created_at_iso=now,
            )
        )


def create_order_from_form(form: Any) -> MaintenanceOrder:
    now = now_operacion_local_iso_seconds()
    order = MaintenanceOrder(created_at_iso=now, updated_at_iso=now, equipo_id=0, fecha_programada=date.today().isoformat())
    _apply_order_form(order, form)
    db.session.add(order)
    db.session.flush()
    _apply_order_resources_from_form(order, form)
    return order


def create_order_from_plan(plan: MaintenancePlan, fecha_programada: str | None = None) -> MaintenanceOrder:
    now = now_operacion_local_iso_seconds()
    order = MaintenanceOrder(
        plan_id=int(plan.id),
        equipo_id=int(plan.equipo_id),
        component_id=int(plan.component_id) if plan.component_id is not None else None,
        tipo_mantenimiento=plan.tipo_mantenimiento or "preventivo",
        fecha_programada=_normalize_date(fecha_programada) or plan.proxima_fecha or date.today().isoformat(),
        prioridad="media",
        criticidad="media",
        responsable=plan.responsable,
        estado="programado",
        tareas=plan.tareas,
        tiempo_estimado_horas=plan.duracion_estimada_horas,
        observaciones=plan.observaciones,
        created_at_iso=now,
        updated_at_iso=now,
    )
    db.session.add(order)
    db.session.flush()
    return order


def update_order_from_form(order: MaintenanceOrder, form: Any) -> None:
    _apply_order_form(order, form)
    _apply_order_resources_from_form(order, form)
    if order.estado == "finalizado" and order.plan_id:
        plan = db.session.get(MaintenancePlan, order.plan_id)
        if plan is not None and plan.frecuencia_dias and order.fecha_programada:
            try:
                plan.proxima_fecha = (date.fromisoformat(order.fecha_programada) + timedelta(days=int(plan.frecuencia_dias))).isoformat()
                plan.updated_at_iso = now_operacion_local_iso_seconds()
            except ValueError:
                pass


def list_resources() -> list[MaintenanceResource]:
    return list(
        db.session.scalars(
            select(MaintenanceResource)
            .options(selectinload(MaintenanceResource.equipo), selectinload(MaintenanceResource.component))
            .order_by(MaintenanceResource.activo.desc(), MaintenanceResource.categoria.asc(), MaintenanceResource.nombre.asc())
        ).all()
    )


def create_resource_from_form(form: Any) -> MaintenanceResource:
    equipo_id = _parse_int(form.get("equipo_id"))
    component_id = _parse_int(form.get("component_id"))
    if equipo_id is not None and db.session.get(Equipo, equipo_id) is None:
        raise ValueError("El equipo seleccionado no existe.")
    if equipo_id is not None and not _component_belongs_to_equipo(component_id, equipo_id):
        raise ValueError("El componente seleccionado no corresponde al equipo.")
    now = now_operacion_local_iso_seconds()
    resource = MaintenanceResource(
        equipo_id=equipo_id,
        component_id=component_id,
        tipo_mantenimiento=_valid_or_default(form.get("tipo_mantenimiento"), TIPOS_MANTENIMIENTO, "preventivo"),
        categoria=_valid_or_default(form.get("categoria"), RESOURCE_CATEGORIAS, "personal"),
        nombre=_clean(form.get("nombre")),
        cantidad=_parse_float(form.get("cantidad")),
        unidad=_none_if_empty(form.get("unidad")),
        tiempo_estimado_horas=_parse_float(form.get("tiempo_estimado_horas")),
        observaciones=_none_if_empty(form.get("observaciones")),
        activo=form.get("activo", "1") == "1",
        created_at_iso=now,
        updated_at_iso=now,
    )
    if not resource.nombre:
        raise ValueError("El nombre del recurso es obligatorio.")
    db.session.add(resource)
    return resource


def _failure_group_key(failure: MaintenanceFailure) -> str:
    raw = failure.causa_real or failure.causa_probable or failure.sintoma_observado or failure.descripcion_falla
    normalized = " ".join(_clean(raw).lower().split())[:180] or "falla"
    return f"{failure.equipo_id}:{failure.component_id or 0}:{normalized}"


def refresh_predictions() -> int:
    rows = list(
        db.session.scalars(
            select(MaintenanceFailure)
            .where(MaintenanceFailure.estado == "finalizado")
            .order_by(MaintenanceFailure.equipo_id.asc(), MaintenanceFailure.component_id.asc(), MaintenanceFailure.detected_at_iso.asc())
        ).all()
    )
    grouped: dict[str, list[MaintenanceFailure]] = {}
    for row in rows:
        grouped.setdefault(_failure_group_key(row), []).append(row)
    count = 0
    now = now_operacion_local_iso_seconds()
    for source_key, failures in grouped.items():
        if len(failures) < 3:
            continue
        dates: list[date] = []
        for f in failures:
            d = _normalize_date(f.detected_at_iso)
            if d is not None:
                dates.append(date.fromisoformat(d))
        dates = sorted(set(dates))
        if len(dates) < 3:
            continue
        intervals = [(b - a).days for a, b in zip(dates, dates[1:]) if (b - a).days > 0]
        if len(intervals) < 2:
            continue
        avg = round(mean(intervals), 1)
        deviation = pstdev(intervals) if len(intervals) > 1 else 0.0
        confidence = "alta" if len(dates) >= 5 and deviation <= avg * 0.25 else "media" if len(dates) >= 4 else "bajo"
        next_date = dates[-1] + timedelta(days=max(int(round(avg)), 1))
        sample = failures[-1]
        tipo = sample.causa_real or sample.causa_probable or sample.sintoma_observado or sample.descripcion_falla
        prediction = db.session.scalar(select(MaintenancePrediction).where(MaintenancePrediction.source_key == source_key))
        if prediction is None:
            prediction = MaintenancePrediction(source_key=source_key, created_at_iso=now)
            db.session.add(prediction)
        prediction.equipo_id = int(sample.equipo_id)
        prediction.component_id = int(sample.component_id) if sample.component_id is not None else None
        prediction.tipo_falla = _clean(tipo)[:256] or "Falla repetitiva"
        prediction.cantidad_fallas = len(dates)
        prediction.promedio_dias_entre_fallas = avg
        prediction.ultima_fecha_falla = dates[-1].isoformat()
        prediction.fecha_estimada_proxima = next_date.isoformat()
        prediction.nivel_confianza = confidence
        prediction.recomendacion = f"Revisar antes de {next_date.isoformat()} por patrón histórico cada {avg:g} días."
        if prediction.estado not in ("programada", "descartada"):
            prediction.estado = "sugerida"
        prediction.updated_at_iso = now
        count += 1
    return count


def list_predictions() -> list[MaintenancePrediction]:
    return list(
        db.session.scalars(
            select(MaintenancePrediction)
            .options(selectinload(MaintenancePrediction.equipo), selectinload(MaintenancePrediction.component))
            .order_by(MaintenancePrediction.fecha_estimada_proxima.asc(), MaintenancePrediction.id.desc())
        ).all()
    )


def create_order_from_prediction(prediction: MaintenancePrediction, fecha_programada: str | None = None) -> MaintenanceOrder:
    now = now_operacion_local_iso_seconds()
    order = MaintenanceOrder(
        prediction_id=int(prediction.id),
        equipo_id=int(prediction.equipo_id),
        component_id=int(prediction.component_id) if prediction.component_id is not None else None,
        tipo_mantenimiento="predictivo",
        fecha_programada=_normalize_date(fecha_programada) or prediction.fecha_estimada_proxima or date.today().isoformat(),
        prioridad="alta",
        criticidad="alta" if prediction.nivel_confianza == "alta" else "media",
        estado="programado",
        tareas=prediction.recomendacion,
        observaciones=f"Sugerencia predictiva: {prediction.tipo_falla}",
        created_at_iso=now,
        updated_at_iso=now,
    )
    db.session.add(order)
    db.session.flush()
    prediction.estado = "programada"
    prediction.updated_at_iso = now
    return order


def dashboard_counts() -> dict[str, int]:
    today = date.today().isoformat()
    next_week = (date.today() + timedelta(days=7)).isoformat()
    abiertos = db.session.scalar(
        select(func.count()).select_from(MaintenanceFailure).where(MaintenanceFailure.estado != "finalizado")
    )
    criticos = db.session.scalar(
        select(func.count())
        .select_from(MaintenanceFailure)
        .where(MaintenanceFailure.estado != "finalizado", MaintenanceFailure.criticidad == "critica")
    )
    equipos_mantenimiento = db.session.scalar(
        select(func.count()).select_from(Equipo).where(Equipo.estado == "en_mantenimiento")
    )
    componentes = db.session.scalar(select(func.count()).select_from(MaintenanceComponent))
    vencidos = db.session.scalar(
        select(func.count())
        .select_from(MaintenanceOrder)
        .where(MaintenanceOrder.fecha_programada < today, MaintenanceOrder.estado.notin_(("finalizado", "cancelado")))
    )
    proximos = db.session.scalar(
        select(func.count())
        .select_from(MaintenanceOrder)
        .where(
            MaintenanceOrder.fecha_programada >= today,
            MaintenanceOrder.fecha_programada <= next_week,
            MaintenanceOrder.estado.notin_(("finalizado", "cancelado")),
        )
    )
    correctivos_mes = db.session.scalar(
        select(func.count()).select_from(MaintenanceFailure).where(MaintenanceFailure.detected_at_iso >= today[:8] + "01")
    )
    preventivos_mes = db.session.scalar(
        select(func.count())
        .select_from(MaintenanceOrder)
        .where(MaintenanceOrder.tipo_mantenimiento == "preventivo", MaintenanceOrder.fecha_programada >= today[:8] + "01")
    )
    downtime = db.session.scalar(select(func.coalesce(func.sum(MaintenanceFailure.tiempo_fuera_servicio_horas), 0.0)))
    predictivas = db.session.scalar(
        select(func.count()).select_from(MaintenancePrediction).where(MaintenancePrediction.estado == "sugerida")
    )
    return {
        "correctivos_abiertos": int(abiertos or 0),
        "correctivos_criticos": int(criticos or 0),
        "equipos_en_mantenimiento": int(equipos_mantenimiento or 0),
        "componentes": int(componentes or 0),
        "mantenimientos_vencidos": int(vencidos or 0),
        "mantenimientos_proximos": int(proximos or 0),
        "correctivos_mes": int(correctivos_mes or 0),
        "preventivos_mes": int(preventivos_mes or 0),
        "tiempo_fuera_servicio_horas": int(round(float(downtime or 0.0))),
        "predictivas_sugeridas": int(predictivas or 0),
    }


def top_failures_by_equipo(limit: int = 5) -> list[tuple[str, int]]:
    rows = db.session.execute(
        select(Equipo.nombre_equipo, func.count(MaintenanceFailure.id))
        .join(MaintenanceFailure, MaintenanceFailure.equipo_id == Equipo.id)
        .group_by(Equipo.nombre_equipo)
        .order_by(func.count(MaintenanceFailure.id).desc())
        .limit(limit)
    ).all()
    return [(str(name), int(count)) for name, count in rows]


def top_failures_by_component(limit: int = 5) -> list[tuple[str, int]]:
    rows = db.session.execute(
        select(MaintenanceComponent.nombre, func.count(MaintenanceFailure.id))
        .join(MaintenanceFailure, MaintenanceFailure.component_id == MaintenanceComponent.id)
        .group_by(MaintenanceComponent.nombre)
        .order_by(func.count(MaintenanceFailure.id).desc())
        .limit(limit)
    ).all()
    return [(str(name), int(count)) for name, count in rows]


def order_counts_by_estado() -> list[tuple[str, int]]:
    rows = db.session.execute(
        select(MaintenanceOrder.estado, func.count(MaintenanceOrder.id))
        .group_by(MaintenanceOrder.estado)
        .order_by(func.count(MaintenanceOrder.id).desc())
    ).all()
    return [(ORDER_ESTADOS.get(str(k), str(k)), int(v)) for k, v in rows]


def order_counts_by_criticidad() -> list[tuple[str, int]]:
    rows = db.session.execute(
        select(MaintenanceOrder.criticidad, func.count(MaintenanceOrder.id))
        .group_by(MaintenanceOrder.criticidad)
        .order_by(func.count(MaintenanceOrder.id).desc())
    ).all()
    return [(CRITICIDADES.get(str(k), str(k)), int(v)) for k, v in rows]


def export_correctivos_xlsx(rows: list[MaintenanceFailure]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Correctivos"
    headers = [
        "ID",
        "Detectado",
        "Equipo",
        "Componente",
        "Reporta",
        "Descripción",
        "Síntoma",
        "Causa probable",
        "Causa real",
        "Criticidad",
        "Estado",
        "Responsable",
        "Cierre",
        "Horas fuera de servicio",
        "Acción realizada",
        "Repuestos",
        "Recursos",
        "Observaciones",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r in rows:
        ws.append(
            [
                r.id,
                format_dt(r.detected_at_iso),
                r.equipo.nombre_equipo if r.equipo else r.equipo_id,
                r.component.nombre if r.component else "",
                r.reported_by_display,
                r.descripcion_falla,
                r.sintoma_observado,
                r.causa_probable,
                r.causa_real,
                CRITICIDADES.get(r.criticidad, r.criticidad),
                FAILURE_ESTADOS.get(r.estado, r.estado),
                r.responsable_trabajo,
                format_dt(r.closed_at_iso),
                r.tiempo_fuera_servicio_horas,
                r.accion_realizada,
                r.repuestos_utilizados,
                r.recursos_utilizados,
                r.observaciones,
            ]
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
    for idx in range(1, len(headers) + 1):
        max_len = 10
        for row in ws.iter_rows(min_col=idx, max_col=idx, max_row=min(ws.max_row, 300)):
            val = row[0].value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 55)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def export_orders_xlsx(rows: list[MaintenanceOrder]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Mantenimientos"
    headers = [
        "ID",
        "Fecha programada",
        "Tipo",
        "Equipo",
        "Componente",
        "Estado",
        "Prioridad",
        "Criticidad",
        "Responsable",
        "Tareas",
        "Recursos",
        "Repuestos",
        "Herramientas",
        "EPP",
        "Tiempo estimado",
        "Resultado",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r in rows:
        ws.append(
            [
                r.id,
                r.fecha_programada,
                TIPOS_MANTENIMIENTO.get(r.tipo_mantenimiento, r.tipo_mantenimiento),
                r.equipo.nombre_equipo if r.equipo else r.equipo_id,
                r.component.nombre if r.component else "",
                ORDER_ESTADOS.get(r.estado, r.estado),
                PRIORIDADES.get(r.prioridad, r.prioridad),
                CRITICIDADES.get(r.criticidad, r.criticidad),
                r.responsable,
                r.tareas,
                r.recursos_necesarios,
                r.repuestos_necesarios,
                r.herramientas_necesarias,
                r.epp_necesarios,
                r.tiempo_estimado_horas,
                r.resultado,
            ]
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
    for idx in range(1, len(headers) + 1):
        max_len = 10
        for row in ws.iter_rows(min_col=idx, max_col=idx, max_row=min(ws.max_row, 300)):
            val = row[0].value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 55)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
