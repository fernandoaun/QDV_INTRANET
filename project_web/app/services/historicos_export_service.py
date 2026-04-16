"""Exportación de históricos operativos a Excel (.xlsx), una hoja por módulo."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Any, Callable

from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.auth_utils import (
    user_can,
    user_can_access_entregas_hub,
    user_can_access_production_hub,
    user_can_view_stock_consumos,
    user_can_view_stock_historial,
    user_can_view_stock_ingreso_categoria,
    user_display_name,
)
from app.extensions import db
from app.models import (
    AguaRegistro,
    BolsonRegistro,
    ConsumoStock,
    Entrega,
    IngresoStock,
    ReactorRegistro,
    SalmueraRegistro,
    ShiftHandover,
    User,
)
from app.web.modules.produccion.salmuera_helpers import salmuera_row_to_dict

MAX_ROWS_PER_SHEET = 80_000

DateRangeValidate = tuple[bool, str | None, date | None, date | None]


def parse_and_validate_rango_fechas(desde_raw: str, hasta_raw: str) -> DateRangeValidate:
    """Devuelve (ok, mensaje_error, desde, hasta). Ambos extremos inclusive en filtros."""
    ds = (desde_raw or "").strip()
    hs = (hasta_raw or "").strip()
    if not ds or not hs:
        return False, "Indicá fecha desde y fecha hasta.", None, None
    try:
        d0 = date.fromisoformat(ds)
        d1 = date.fromisoformat(hs)
    except ValueError:
        return False, "Las fechas deben tener formato AAAA-MM-DD.", None, None
    if d1 < d0:
        return False, "La fecha hasta no puede ser anterior a la fecha desde.", None, None
    return True, None, d0, d1


def _iso_bounds(d0: date, d1: date) -> tuple[str, str]:
    """Límites string comparables con campos ISO guardados en BD (naive local)."""
    return (f"{d0.isoformat()}T00:00:00", f"{d1.isoformat()}T23:59:59")


def _safe_sheet_title(name: str) -> str:
    s = re.sub(r"[" + re.escape("[]:*?/\\") + r"]", "_", (name or "").strip())[:31]
    return s or "Hoja"


def _autosize_and_freeze(ws, ncols: int, nrows: int) -> None:
    from openpyxl.utils import get_column_letter

    if nrows < 1:
        return
    ws.freeze_panes = "A2"
    last_col = get_column_letter(ncols)
    ws.auto_filter.ref = f"A1:{last_col}{nrows}"
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        maxlen = 10
        for r in range(1, min(nrows, 500) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                maxlen = max(maxlen, min(len(str(v)), 60))
        ws.column_dimensions[col_letter].width = min(maxlen + 2, 55)


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    from openpyxl.styles import Alignment, Font

    bold = Font(bold=True)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r_i, row in enumerate(rows, start=2):
        for c_i, val in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    _autosize_and_freeze(ws, len(headers), 1 + len(rows))


def _fetch_salmuera(d0: date, d1: date) -> list[list[Any]]:
    rows = db.session.scalars(
        select(SalmueraRegistro)
        .where(SalmueraRegistro.fecha_iso >= d0.isoformat(), SalmueraRegistro.fecha_iso <= d1.isoformat())
        .order_by(SalmueraRegistro.fecha_iso.asc(), SalmueraRegistro.id.asc())
        .limit(MAX_ROWS_PER_SHEET)
    ).all()
    out: list[list[Any]] = []
    for r in rows:
        d = salmuera_row_to_dict(r)
        vj = d.get("voltajes_celdas") or []
        v_str = ", ".join(str(x) for x in vj) if isinstance(vj, list) else str(r.voltajes_json or "")
        warns = d.get("warnings") or []
        w_str = "; ".join(str(x) for x in warns) if warns else ""
        out.append(
            [
                d.get("id"),
                d.get("fecha_iso"),
                d.get("hora_hm"),
                d.get("created_at_iso"),
                d.get("electrolizador"),
                d.get("turno"),
                d.get("operador"),
                d.get("lote"),
                d.get("cantidad_celdas"),
                v_str,
                d.get("voltaje_total"),
                d.get("amperaje"),
                d.get("caudal_agua_l_h"),
                d.get("caudal_salmuera_l_h"),
                d.get("hipo_conc"),
                d.get("hipo_exceso_soda"),
                d.get("sal_temp"),
                d.get("sal_conc"),
                d.get("sal_ph"),
                d.get("soda_conc"),
                d.get("declor_ph"),
                w_str,
                d.get("observaciones"),
                d.get("atraso_motivo"),
            ]
        )
    return out


def _fetch_reactor(d0: date, d1: date) -> list[list[Any]]:
    rows = db.session.scalars(
        select(ReactorRegistro)
        .where(ReactorRegistro.fecha_iso >= d0.isoformat(), ReactorRegistro.fecha_iso <= d1.isoformat())
        .order_by(ReactorRegistro.fecha_iso.asc(), ReactorRegistro.id.asc())
        .limit(MAX_ROWS_PER_SHEET)
    ).all()
    return [
        [
            r.id,
            r.fecha_iso,
            r.hora_hm,
            r.created_at_iso,
            r.operador,
            r.lote,
            r.ph,
            r.temperatura,
            r.densidad,
            r.concentracion_tabla,
            r.exceso_naoh,
            r.exceso_na2co3,
            r.observaciones,
        ]
        for r in rows
    ]


def _fetch_agua(d0: date, d1: date) -> list[list[Any]]:
    rows = db.session.scalars(
        select(AguaRegistro)
        .where(AguaRegistro.fecha_iso >= d0.isoformat(), AguaRegistro.fecha_iso <= d1.isoformat())
        .order_by(AguaRegistro.fecha_iso.asc(), AguaRegistro.id.asc())
        .limit(MAX_ROWS_PER_SHEET)
    ).all()
    return [
        [
            r.id,
            r.fecha_iso,
            r.hora_hm,
            r.created_at_iso,
            r.turno,
            r.operador,
            r.lote,
            r.numero_columna,
            r.temperatura,
            r.dureza,
            r.observaciones,
        ]
        for r in rows
    ]


def _fetch_bolson(d0: date, d1: date) -> list[list[Any]]:
    rows = db.session.scalars(
        select(BolsonRegistro)
        .where(BolsonRegistro.fecha_iso >= d0.isoformat(), BolsonRegistro.fecha_iso <= d1.isoformat())
        .order_by(BolsonRegistro.fecha_iso.asc(), BolsonRegistro.id.asc())
        .limit(MAX_ROWS_PER_SHEET)
    ).all()
    return [[r.id, r.fecha_iso, r.hora_hm, r.created_at_iso] for r in rows]


def _fetch_entregas(d0: date, d1: date) -> list[list[Any]]:
    rows = db.session.scalars(
        select(Entrega)
        .where(Entrega.fecha_prevista >= d0.isoformat(), Entrega.fecha_prevista <= d1.isoformat())
        .options(
            selectinload(Entrega.cliente_row),
            selectinload(Entrega.lugar_row),
            selectinload(Entrega.producto_terminado),
            selectinload(Entrega.chofer_row),
        )
        .order_by(Entrega.fecha_prevista.asc(), Entrega.id.asc())
        .limit(MAX_ROWS_PER_SHEET)
    ).all()
    out: list[list[Any]] = []
    for e in rows:
        cli = e.cliente_row.nombre if e.cliente_row else e.cliente
        lug = e.lugar_row.nombre if e.lugar_row else e.lugar_entrega
        prod = e.producto_terminado.nombre if e.producto_terminado else e.producto
        ch = e.chofer_row.nombre if e.chofer_row else (e.chofer_previsto or "")
        out.append(
            [
                e.id,
                e.estado,
                cli,
                lug,
                prod,
                e.cantidad,
                (e.unidad or "").strip() or "L",
                e.fecha_prevista,
                ch,
                e.observaciones,
                e.created_at_iso,
                e.updated_at_iso,
                e.cargada_at_iso,
                e.entregada_at_iso,
                e.entregada_chofer_nombre,
                e.entregada_lugar,
                e.entregada_dia_semana,
                e.stock_categoria,
                e.stock_marca,
                e.consumo_stock_id,
            ]
        )
    return out


def _fetch_ingresos_stock(d0: date, d1: date) -> list[list[Any]]:
    d_s, d_e = d0.isoformat(), d1.isoformat()
    rows = db.session.scalars(
        select(IngresoStock)
        .where(IngresoStock.fecha >= d_s, IngresoStock.fecha <= d_e)
        .order_by(IngresoStock.fecha.asc(), IngresoStock.id.asc())
        .limit(MAX_ROWS_PER_SHEET)
    ).all()
    return [
        [
            r.id,
            r.categoria,
            r.producto,
            r.marca,
            r.lote,
            r.cantidad,
            r.unidad,
            r.fecha,
            r.hora,
            r.vencimiento,
            r.operador,
            r.proveedor,
            r.observaciones,
            r.created_at_iso,
            r.cargado_por_user_id,
        ]
        for r in rows
    ]


def _fetch_consumos_stock(d0: date, d1: date) -> list[list[Any]]:
    d_s, d_e = d0.isoformat(), d1.isoformat()
    rows = db.session.scalars(
        select(ConsumoStock)
        .where(ConsumoStock.fecha >= d_s, ConsumoStock.fecha <= d_e)
        .order_by(ConsumoStock.fecha.asc(), ConsumoStock.id.asc())
        .limit(MAX_ROWS_PER_SHEET)
    ).all()
    return [
        [
            r.id,
            r.categoria,
            r.producto,
            r.marca,
            r.cantidad,
            r.fecha,
            r.hora,
            r.operador,
            r.observaciones,
            r.equipo_id,
            r.ingreso_stock_id,
            r.created_at_iso,
        ]
        for r in rows
    ]


def _fetch_cambio_turno(d0: date, d1: date) -> list[list[Any]]:
    lo, hi = _iso_bounds(d0, d1)
    rows = db.session.scalars(
        select(ShiftHandover)
        .where(ShiftHandover.handed_over_at_iso >= lo, ShiftHandover.handed_over_at_iso <= hi)
        .order_by(ShiftHandover.handed_over_at_iso.asc(), ShiftHandover.id.asc())
        .limit(MAX_ROWS_PER_SHEET)
    ).all()
    out: list[list[Any]] = []
    for h in rows:
        out_u = db.session.get(User, h.outgoing_user_id)
        in_u = db.session.get(User, h.incoming_user_id) if h.incoming_user_id else None
        out.append(
            [
                h.id,
                h.status,
                h.shift_started_at_iso,
                h.handed_over_at_iso,
                h.received_at_iso,
                user_display_name(out_u) or (out_u.username if out_u else ""),
                user_display_name(in_u) or (in_u.username if in_u else ""),
                h.hypochlorite_stock_liters,
                h.closing_notes,
                h.reception_status,
                h.reception_notes,
                h.created_at_iso,
                h.updated_at_iso,
            ]
        )
    return out


@dataclass(frozen=True)
class ExportModuleDef:
    key: str
    label: str
    sheet_title: str
    perm: Callable[[Any], bool]


def export_module_definitions() -> list[ExportModuleDef]:
    """Definición estable de módulos exportables (permisos de vista)."""

    def _can_stock_ing(u: Any) -> bool:
        return bool(
            user_can_view_stock_historial(u)
            or user_can_view_stock_ingreso_categoria(u, "materia_prima")
            or user_can_view_stock_ingreso_categoria(u, "laboratorio")
        )

    def _can_stock_con(u: Any) -> bool:
        return bool(user_can_view_stock_historial(u) or user_can_view_stock_consumos(u))

    def _can_shift(u: Any) -> bool:
        if u is None:
            return False
        return bool(getattr(u, "is_admin", False) or user_can_access_production_hub(u))

    return [
        ExportModuleDef(
            "salmuera",
            "Control de hipoclorito (registros por fecha de planilla)",
            "Hipoclorito",
            lambda u: user_can(u, "salmuera"),
        ),
        ExportModuleDef(
            "reactor",
            "Circuito de salmuera — reactor (registros por fecha)",
            "Reactor_salmuera",
            lambda u: user_can(u, "reactor"),
        ),
        ExportModuleDef(
            "agua",
            "Circuito de agua (registros por fecha)",
            "Agua",
            lambda u: user_can(u, "agua"),
        ),
        ExportModuleDef(
            "bolson",
            "Registro bolsón / fechas operativas asociadas",
            "Bolson",
            lambda u: user_can(u, "bolson_registro"),
        ),
        ExportModuleDef(
            "entregas",
            "Entregas (filtradas por fecha prevista; inclusive extremos)",
            "Entregas",
            user_can_access_entregas_hub,
        ),
        ExportModuleDef(
            "stock_ingresos",
            "Stock — ingresos (fecha del ingreso)",
            "Stock_ingresos",
            _can_stock_ing,
        ),
        ExportModuleDef(
            "stock_consumos",
            "Stock — consumos (fecha del consumo)",
            "Stock_consumos",
            _can_stock_con,
        ),
        ExportModuleDef(
            "cambio_turno",
            "Cambio de turno (filtrado por fecha/hora de entrega del turno)",
            "Cambio_turno",
            _can_shift,
        ),
    ]


def allowed_export_keys_for_user(user: Any) -> list[str]:
    return [m.key for m in export_module_definitions() if m.perm(user)]


HEADERS: dict[str, list[str]] = {
    "salmuera": [
        "ID",
        "Fecha planilla",
        "Hora (HM)",
        "Creado (ISO)",
        "Electrolizador",
        "Turno",
        "Operador",
        "Lote",
        "Celdas (N)",
        "Voltajes celdas",
        "V total",
        "Amperaje",
        "Caudal agua L/h",
        "Caudal salmuera L/h",
        "Hipo conc",
        "Hipo exceso soda",
        "Sal temp",
        "Sal conc",
        "Sal pH",
        "Soda conc",
        "Declor pH",
        "Avisos",
        "Observaciones",
        "Motivo atraso",
    ],
    "reactor": [
        "ID",
        "Fecha",
        "Hora",
        "Creado (ISO)",
        "Operador",
        "Lote",
        "pH",
        "Temperatura",
        "Densidad",
        "Conc. tabla",
        "Exceso NaOH",
        "Exceso Na2CO3",
        "Observaciones",
    ],
    "agua": [
        "ID",
        "Fecha",
        "Hora",
        "Creado (ISO)",
        "Turno",
        "Operador",
        "Lote",
        "N° columna",
        "Temperatura",
        "Dureza",
        "Observaciones",
    ],
    "bolson": ["ID", "Fecha", "Hora", "Creado (ISO)"],
    "entregas": [
        "ID",
        "Estado",
        "Cliente",
        "Lugar entrega",
        "Producto",
        "Cantidad",
        "Unidad",
        "Fecha prevista",
        "Chofer (previsto / catálogo)",
        "Observaciones",
        "Creado (ISO)",
        "Actualizado (ISO)",
        "Cargada (ISO)",
        "Entregada (ISO)",
        "Chofer entrega (confirmación)",
        "Lugar entrega (confirmación)",
        "Día semana entrega",
        "Stock categoría",
        "Stock marca",
        "Consumo stock ID",
    ],
    "stock_ingresos": [
        "ID",
        "Categoría",
        "Producto",
        "Marca",
        "Lote",
        "Cantidad",
        "Unidad",
        "Fecha",
        "Hora",
        "Vencimiento",
        "Operador",
        "Proveedor",
        "Observaciones",
        "Creado (ISO)",
        "Usuario carga ID",
    ],
    "stock_consumos": [
        "ID",
        "Categoría",
        "Producto",
        "Marca",
        "Cantidad",
        "Fecha",
        "Hora",
        "Operador",
        "Observaciones",
        "Equipo ID",
        "Ingreso stock ID",
        "Creado (ISO)",
    ],
    "cambio_turno": [
        "ID",
        "Estado",
        "Inicio turno (ISO)",
        "Entrega turno (ISO)",
        "Recepción (ISO)",
        "Operador saliente",
        "Operador entrante",
        "Stock hipoclorito (L declarado)",
        "Notas cierre",
        "Estado recepción",
        "Notas recepción",
        "Creado (ISO)",
        "Actualizado (ISO)",
    ],
}

FETCHERS: dict[str, Callable[[date, date], list[list[Any]]]] = {
    "salmuera": _fetch_salmuera,
    "reactor": _fetch_reactor,
    "agua": _fetch_agua,
    "bolson": _fetch_bolson,
    "entregas": _fetch_entregas,
    "stock_ingresos": _fetch_ingresos_stock,
    "stock_consumos": _fetch_consumos_stock,
    "cambio_turno": _fetch_cambio_turno,
}


def build_historicos_workbook(selected_keys: list[str], d0: date, d1: date) -> tuple[BytesIO | None, str | None]:
    """
    Genera un único .xlsx con una hoja por clave en `selected_keys`.
    Retorna (buffer, error_msg). Si error_msg no es None, buffer es None.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        return None, (
            "En el servidor no está instalada la librería openpyxl (necesaria para Excel). "
            "El administrador debe asegurarse de que el deploy ejecute: pip install -r requirements.txt"
        )

    keys = [k for k in selected_keys if k in FETCHERS]
    if not keys:
        return None, "No hay módulos válidos para exportar."

    wb = Workbook()
    first = True
    total_rows = 0
    for key in keys:
        fetch = FETCHERS[key]
        data = fetch(d0, d1)
        headers = HEADERS[key]
        title = _safe_sheet_title(
            next((m.sheet_title for m in export_module_definitions() if m.key == key), key)
        )
        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title=title)
        _write_sheet(ws, headers, data)
        total_rows += len(data)

    if total_rows == 0:
        return None, (
            "No hay registros en el período elegido para los módulos seleccionados "
            f"({d0.isoformat()} a {d1.isoformat()}, extremos inclusive)."
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, None


def export_download_filename(d0: date, d1: date) -> str:
    return f"historicos_sistema_{d0.isoformat()}_a_{d1.isoformat()}.xlsx"
