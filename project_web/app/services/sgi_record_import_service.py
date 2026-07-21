"""Importación segura de Word/Excel hacia esquemas de registro digital."""
from __future__ import annotations

import base64
import hashlib
import html
import io
import mimetypes
import re
import uuid
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from app.models.sgi import (
    ASSOC_IMPORTED_EXCEL,
    ASSOC_IMPORTED_WORD,
    RECORD_ALLOWED_EXTENSIONS,
    RECORD_MAX_UPLOAD_BYTES,
)

_W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_V_NS = "urn:schemas-microsoft-com:vml"
_W = f"{{{_W_NS}}}"
_R = f"{{{_R_NS}}}"
_A = f"{{{_A_NS}}}"
_V = f"{{{_V_NS}}}"
_HEADER_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/header"
)
_FOOTER_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/footer"
)
_IMAGE_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
)

_SAFE_FORMULA_FUNCS = {
    "SUM",
    "SUMA",
    "AVERAGE",
    "PROMEDIO",
    "MIN",
    "MINIMO",
    "MÍNIMO",
    "MAX",
    "MAXIMO",
    "MÁXIMO",
    "COUNT",
    "CONTAR",
    "IF",
    "SI",
    "COUNTIF",
    "CONTAR.SI",
    "SUMIF",
    "SUMAR.SI",
}

_MIME_BY_EXT = {
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
}


class ImportSecurityError(ValueError):
    """Archivo rechazado por reglas de seguridad o formato."""


def _slug(text: str, fallback: str = "campo") -> str:
    raw = (text or "").strip().lower()
    raw = re.sub(r"[^\w\s\-]", "", raw, flags=re.UNICODE)
    raw = re.sub(r"[\s\-]+", "_", raw).strip("_")
    return (raw[:48] or fallback)


def _field(
    *,
    name: str,
    label: str,
    ftype: str,
    order: int,
    section: str = "Datos generales",
    required: bool = False,
    options: list[str] | None = None,
    default_value: Any = None,
    placeholder: str = "",
    validation: dict | None = None,
    formula: str | None = None,
    columns: list[dict] | None = None,
) -> dict[str, Any]:
    out: dict[str, Any] = {
        "id": str(uuid.uuid4()),
        "name": name,
        "label": label,
        "type": ftype,
        "required": bool(required),
        "order": order,
        "section": section,
        "defaultValue": default_value,
        "placeholder": placeholder,
        "validation": validation or {},
        "options": options or [],
        "permissions": {"editableBy": [], "visibleTo": []},
        "mode": "editable",
    }
    if formula:
        out["formula"] = formula
    if columns is not None:
        out["columns"] = columns
    return out


def infer_field_type(label: str) -> str:
    t = (label or "").lower()
    if any(k in t for k in ("fecha y hora", "datetime", "timestamp")):
        return "datetime"
    if any(k in t for k in ("fecha", "date")):
        return "date"
    if any(k in t for k in ("hora", "time")):
        return "time"
    if any(k in t for k in ("correo", "email", "e-mail")):
        return "email"
    if any(k in t for k in ("teléfono", "telefono", "phone", "celular")):
        return "phone"
    if any(k in t for k in ("porcentaje", "%", "percent")):
        return "percent"
    if any(k in t for k in ("moneda", "$", "importe", "precio", "costo")):
        return "currency"
    if any(k in t for k in ("sí/no", "si/no", "yes/no", "conforme")):
        return "yes_no"
    if any(k in t for k in ("observacion", "observación", "comentario", "notas")):
        return "textarea"
    if any(k in t for k in ("firma", "signature")):
        return "signature"
    if any(k in t for k in ("responsable", "usuario", "operador")):
        return "system_user"
    if any(k in t for k in ("área", "area", "sector")):
        return "area"
    if any(k in t for k in ("equipo",)):
        return "equipment"
    if any(k in t for k in ("cantidad", "número", "numero", "valor", "total")):
        return "decimal"
    return "text"


def validate_upload(file: FileStorage) -> tuple[str, str]:
    """Valida extensión/tamaño básico. Devuelve (ext, original_name)."""
    if file is None or not getattr(file, "filename", None):
        raise ImportSecurityError("No se recibió ningún archivo.")
    original = (file.filename or "").strip()
    if not original:
        raise ImportSecurityError("Nombre de archivo vacío.")
    # Path traversal / nombres peligrosos
    if ".." in original or "/" in original.replace("\\", "/") or "\\" in original:
        # secure_filename will strip path; still reject obvious traversal in raw name segments
        base = Path(original.replace("\\", "/")).name
        if base != original and (".." in original or original.startswith(("/", "\\"))):
            raise ImportSecurityError("Nombre de archivo no permitido.")
        original = base
    ext = Path(original).suffix.lower()
    if ext not in RECORD_ALLOWED_EXTENSIONS:
        raise ImportSecurityError(
            f"Extensión no permitida. Use: {', '.join(RECORD_ALLOWED_EXTENSIONS)}."
        )
    # Peek size if possible
    stream = file.stream
    pos = stream.tell()
    stream.seek(0, io.SEEK_END)
    size = stream.tell()
    stream.seek(pos)
    if size > RECORD_MAX_UPLOAD_BYTES:
        raise ImportSecurityError(
            f"El archivo supera el tamaño permitido ({RECORD_MAX_UPLOAD_BYTES // (1024 * 1024)} MB)."
        )
    if size == 0:
        raise ImportSecurityError("El archivo está vacío o corrupto.")
    return ext, original


def read_file_bytes(file: FileStorage) -> bytes:
    file.stream.seek(0)
    data = file.stream.read()
    file.stream.seek(0)
    if not data:
        raise ImportSecurityError("El archivo está vacío o corrupto.")
    if len(data) > RECORD_MAX_UPLOAD_BYTES:
        raise ImportSecurityError(
            f"El archivo supera el tamaño permitido ({RECORD_MAX_UPLOAD_BYTES // (1024 * 1024)} MB)."
        )
    return data


def assert_safe_office_bytes(data: bytes, ext: str) -> None:
    """Rechaza ejecutables, macros OOXML y ZIP corruptos."""
    if data[:2] == b"MZ":
        raise ImportSecurityError("No se aceptan archivos ejecutables.")
    if ext in (".docx", ".xlsx"):
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                names = [n.replace("\\", "/") for n in zf.namelist()]
        except zipfile.BadZipFile as exc:
            raise ImportSecurityError("El archivo está corrupto o no es un Office válido.") from exc
        lower = [n.lower() for n in names]
        if any("vbaProject" in n or n.endswith(".bin") and "macro" in n for n in lower):
            raise ImportSecurityError("El archivo contiene macros. No se permiten.")
        if any(n.startswith("xl/vba") or n.startswith("word/vba") for n in lower):
            raise ImportSecurityError("El archivo contiene macros. No se permiten.")
        # Contraseña en OOXML suele cifrar el paquete (EncryptedPackage)
        if any(n.endswith("encryptedpackage") or n.endswith("encryptioninfo") for n in lower):
            raise ImportSecurityError("El archivo está protegido por contraseña.")
    if ext == ".xls":
        # OLE compound: D0 CF 11 E0
        if not data.startswith(b"\xd0\xcf\x11\xe0"):
            raise ImportSecurityError("El archivo .xls no es válido.")
        # Heurística macros VBA
        if b"VBA" in data[:8000] or b"_VBA_PROJECT" in data:
            raise ImportSecurityError("El archivo contiene macros. No se permiten.")


def file_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def safe_storage_name(original: str, content_hash: str) -> str:
    stem = secure_filename(Path(original).stem) or "archivo"
    ext = Path(original).suffix.lower()
    return f"{stem[:80]}_{content_hash[:12]}{ext}"


def origin_type_for_ext(ext: str) -> str:
    if ext == ".docx":
        return ASSOC_IMPORTED_WORD
    return ASSOC_IMPORTED_EXCEL


# --- Word ---

def _docx_paragraph_texts(root: ET.Element) -> list[str]:
    texts: list[str] = []
    for p in root.iter(f"{_W}p"):
        parts = [t.text or "" for t in p.iter(f"{_W}t") if t.text]
        line = "".join(parts).strip()
        if line:
            texts.append(line)
    return texts


def _docx_tables(root: ET.Element) -> list[list[list[str]]]:
    tables: list[list[list[str]]] = []
    for tbl in root.iter(f"{_W}tbl"):
        rows: list[list[str]] = []
        for tr in tbl.findall(f"{_W}tr"):
            cells: list[str] = []
            for tc in tr.findall(f"{_W}tc"):
                parts = [t.text or "" for t in tc.iter(f"{_W}t") if t.text]
                cells.append(" ".join(parts).strip())
            if any(cells):
                rows.append(cells)
        if rows:
            tables.append(rows)
    return tables


def _input_html(name: str, *, input_type: str = "text", extra_class: str = "") -> str:
    cls = f"sgi-rec-inline {extra_class}".strip()
    safe = html.escape(name, quote=True)
    if input_type == "textarea":
        return f'<textarea class="{cls}" data-sgi-field="{safe}" rows="3"></textarea>'
    if input_type == "yes_no":
        return (
            f'<select class="{cls}" data-sgi-field="{safe}">'
            f'<option value="">—</option><option value="Sí">Sí</option><option value="No">No</option>'
            f"</select>"
        )
    return f'<input type="{html.escape(input_type)}" class="{cls}" data-sgi-field="{safe}" />'


def _zip_norm(name: str) -> str:
    return name.replace("\\", "/").lstrip("/")


def _resolve_zip_path(base_dir: str, target: str) -> str:
    target = _zip_norm(target)
    if target.startswith("/"):
        return target.lstrip("/")
    base = Path(_zip_norm(base_dir))
    return _zip_norm(str((base / target).as_posix()))


def _parse_rels(zf: zipfile.ZipFile, rels_path: str) -> dict[str, dict[str, str]]:
    """Mapa rId -> {target, type} desde un .rels OOXML."""
    out: dict[str, dict[str, str]] = {}
    try:
        root = ET.fromstring(zf.read(rels_path))
    except KeyError:
        return out
    for rel in root:
        rid = rel.get("Id") or ""
        target = rel.get("Target") or ""
        rtype = rel.get("Type") or ""
        if rid and target:
            out[rid] = {"target": target, "type": rtype}
    return out


def _media_data_uri(zf: zipfile.ZipFile, media_path: str) -> str | None:
    path = _zip_norm(media_path)
    try:
        raw = zf.read(path)
    except KeyError:
        # Algunas rutas vienen con word/ ya incluido o no
        alt = path if path.startswith("word/") else f"word/{path}"
        try:
            raw = zf.read(alt)
            path = alt
        except KeyError:
            return None
    if not raw or len(raw) > 2_000_000:
        return None
    mime, _ = mimetypes.guess_type(path)
    if not mime:
        lower = path.lower()
        if lower.endswith(".png"):
            mime = "image/png"
        elif lower.endswith((".jpg", ".jpeg")):
            mime = "image/jpeg"
        elif lower.endswith(".gif"):
            mime = "image/gif"
        elif lower.endswith(".wmf"):
            mime = "image/x-wmf"
        elif lower.endswith(".emf"):
            mime = "image/x-emf"
        else:
            mime = "application/octet-stream"
    if mime.startswith("image/x-"):
        # Navegadores no muestran WMF/EMF; omitir
        return None
    b64 = base64.b64encode(raw).decode("ascii")
    return f"data:{mime};base64,{b64}"


def _w_val(el: ET.Element | None) -> str:
    if el is None:
        return ""
    return (el.get(f"{_W}val") or el.get("val") or "").strip()


def _tc_merge_info(tc: ET.Element) -> tuple[str, int]:
    """Devuelve (vmerge_state, colspan). vmerge: '', 'restart', 'continue'."""
    colspan = 1
    vmerge = ""
    tc_pr = tc.find(f"{_W}tcPr")
    if tc_pr is None:
        return vmerge, colspan
    grid_span = tc_pr.find(f"{_W}gridSpan")
    if grid_span is not None:
        try:
            colspan = max(1, int(_w_val(grid_span) or "1"))
        except ValueError:
            colspan = 1
    v_merge = tc_pr.find(f"{_W}vMerge")
    if v_merge is not None:
        raw = _w_val(v_merge).lower()
        vmerge = "restart" if raw == "restart" else "continue"
    return vmerge, colspan


def _ooxml_paragraph_html(p: ET.Element, zf: zipfile.ZipFile, image_map: dict[str, str]) -> str:
    bits: list[str] = []
    # Imágenes DrawingML / VML dentro del párrafo
    for blip in p.iter(f"{_A}blip"):
        rid = blip.get(f"{_R}embed") or blip.get(f"{_R}link") or ""
        src = image_map.get(rid)
        if src:
            bits.append(f'<img class="sgi-rec-doc-logo" src="{html.escape(src, quote=True)}" alt="" />')
    for imagedata in p.iter(f"{_V}imagedata"):
        rid = imagedata.get(f"{_R}id") or imagedata.get("r:id") or ""
        src = image_map.get(rid)
        if src:
            bits.append(f'<img class="sgi-rec-doc-logo" src="{html.escape(src, quote=True)}" alt="" />')
    texts = [t.text or "" for t in p.iter(f"{_W}t") if t.text]
    line = "".join(texts).strip()
    if line:
        bits.append(html.escape(line))
    if not bits:
        return ""
    return f"<p>{''.join(bits)}</p>"


def _tc_inner_html(tc: ET.Element, zf: zipfile.ZipFile, image_map: dict[str, str]) -> str:
    inner: list[str] = []
    for child in list(tc):
        if child.tag == f"{_W}p":
            ph = _ooxml_paragraph_html(child, zf, image_map)
            if ph:
                inner.append(ph[3:-4] if ph.startswith("<p>") and ph.endswith("</p>") else ph)
        elif child.tag == f"{_W}tbl":
            inner.append(_ooxml_table_html(child, zf, image_map))
    if not any("sgi-rec-doc-logo" in x for x in inner):
        for blip in tc.iter(f"{_A}blip"):
            rid = blip.get(f"{_R}embed") or ""
            src = image_map.get(rid)
            if src:
                inner.insert(
                    0,
                    f'<img class="sgi-rec-doc-logo" src="{html.escape(src, quote=True)}" alt="" />',
                )
    return "".join(inner) or "&nbsp;"


def _ooxml_table_html(tbl: ET.Element, zf: zipfile.ZipFile, image_map: dict[str, str]) -> str:
    trs = tbl.findall(f"{_W}tr")
    # Precalcular rowspan real para vMerge restart
    merge_state: list[list[tuple[str, int]]] = []
    for tr in trs:
        merge_state.append([_tc_merge_info(tc) for tc in tr.findall(f"{_W}tc")])

    rowspan_map: dict[tuple[int, int], int] = {}
    for ri, row in enumerate(merge_state):
        for ci, (vmerge, _colspan) in enumerate(row):
            if vmerge != "restart":
                continue
            span = 1
            for rj in range(ri + 1, len(merge_state)):
                # misma columna lógica aproximada por índice de celda
                if ci >= len(merge_state[rj]):
                    break
                if merge_state[rj][ci][0] == "continue":
                    span += 1
                else:
                    break
            rowspan_map[(ri, ci)] = span

    rows_html: list[str] = []
    for ri, tr in enumerate(trs):
        cells_html: list[str] = []
        for ci, tc in enumerate(tr.findall(f"{_W}tc")):
            vmerge, colspan = merge_state[ri][ci]
            if vmerge == "continue":
                continue
            attrs = ""
            if colspan > 1:
                attrs += f' colspan="{colspan}"'
            rs = rowspan_map.get((ri, ci), 1)
            if rs > 1:
                attrs += f' rowspan="{rs}"'
            cells_html.append(f"<td{attrs}>{_tc_inner_html(tc, zf, image_map)}</td>")
        if cells_html:
            rows_html.append(f"<tr>{''.join(cells_html)}</tr>")
    if not rows_html:
        return ""
    return f'<table class="sgi-rec-doc-table sgi-rec-doc-header-table">{"".join(rows_html)}</table>'


def _ooxml_part_to_html(zf: zipfile.ZipFile, part_path: str) -> str:
    """Convierte header/footer/body OOXML a HTML conservando tablas e imágenes."""
    path = _zip_norm(part_path)
    try:
        xml = zf.read(path)
    except KeyError:
        return ""
    root = ET.fromstring(xml)
    part_dir = str(Path(path).parent.as_posix())
    rels_name = f"{part_dir}/_rels/{Path(path).name}.rels"
    rels = _parse_rels(zf, rels_name)
    image_map: dict[str, str] = {}
    for rid, meta in rels.items():
        if _IMAGE_REL_TYPE in (meta.get("type") or "") or "/image" in (meta.get("type") or ""):
            media = _resolve_zip_path(part_dir, meta["target"])
            uri = _media_data_uri(zf, media)
            if uri:
                image_map[rid] = uri

    chunks: list[str] = []
    # Hijos directos del hdr/ftr/body (no iter profundo para no duplicar tablas anidadas)
    container = root
    for candidate in (f"{_W}hdr", f"{_W}ftr", f"{_W}body"):
        found = root.find(candidate)
        if found is not None:
            container = found
            break

    for child in list(container):
        if child.tag == f"{_W}tbl":
            th = _ooxml_table_html(child, zf, image_map)
            if th:
                chunks.append(th)
        elif child.tag == f"{_W}p":
            ph = _ooxml_paragraph_html(child, zf, image_map)
            if ph:
                chunks.append(ph)
    return "\n".join(chunks)


def _docx_section_parts_html(data: bytes, rel_type: str, css_class: str) -> str:
    """Extrae encabezados o pies del Word (word/header*.xml / footer*.xml)."""
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            doc_rels = _parse_rels(zf, "word/_rels/document.xml.rels")
            parts: list[str] = []
            seen: set[str] = set()
            for meta in doc_rels.values():
                if (meta.get("type") or "") != rel_type:
                    continue
                target = _resolve_zip_path("word", meta["target"])
                if target in seen:
                    continue
                seen.add(target)
                html_part = _ooxml_part_to_html(zf, target)
                if html_part.strip():
                    parts.append(html_part)
            if not parts:
                return ""
            # Preferir el primer header/footer distinto (default / first / even suelen repetir)
            body = parts[0]
            return f'<div class="{css_class}">{body}</div>'
    except Exception:
        return ""


def _docx_header_html(data: bytes) -> str:
    return _docx_section_parts_html(data, _HEADER_REL_TYPE, "sgi-rec-doc-header")


def _docx_footer_html(data: bytes) -> str:
    return _docx_section_parts_html(data, _FOOTER_REL_TYPE, "sgi-rec-doc-footer")


def _docx_bytes_to_html(data: bytes) -> str:
    try:
        import mammoth

        result = mammoth.convert_to_html(
            io.BytesIO(data),
            convert_image=mammoth.images.data_uri,
        )
        body = (result.value or "").strip()
        if body:
            return body
    except Exception:
        pass
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            return _ooxml_part_to_html(zf, "word/document.xml") or "<p></p>"
    except Exception:
        return "<p></p>"


def _field_input_from_meta(name: str, ftype: str) -> str:
    if ftype in ("textarea", "observations"):
        return _input_html(name, input_type="textarea")
    if ftype == "yes_no":
        return _input_html(name, input_type="yes_no")
    if ftype == "date":
        return _input_html(name, input_type="date")
    if ftype == "datetime":
        return _input_html(name, input_type="datetime-local")
    if ftype == "time":
        return _input_html(name, input_type="time")
    if ftype in ("integer", "decimal", "percent", "currency"):
        return _input_html(name, input_type="number")
    return _input_html(name, input_type="text")


def _wrap_field_control(control_html: str) -> str:
    """Envuelve el control para que siempre se vea como caja editable."""
    return f'<span class="sgi-rec-field-wrap">{control_html}</span>'


def _norm_label_key(text: str) -> str:
    t = (text or "").strip().lower().rstrip(":").strip()
    t = re.sub(r"\s+", " ", t)
    return t


def _td_visible_text(inner_html: str) -> str:
    """Texto visible de una celda (ignora tags, nbsp, subrayados vacíos)."""
    text = re.sub(r"(?is)<br\s*/?>", " ", inner_html or "")
    text = re.sub(r"(?is)<[^>]+>", " ", text)
    text = html.unescape(text)
    text = text.replace("\xa0", " ").replace("\u200b", "")
    # Guiones/ Underscores / puntos de relleno no cuentan como contenido
    text = re.sub(r"[\s_\.\-–—·•□☐]+", "", text)
    return text.strip()


_TD_CELL_RE = re.compile(r"<td([^>]*)>((?:(?!</?td\b).)*)</td>", re.I | re.S)


def _fill_blank_table_cells(html_in: str, make_control) -> str:
    """Pone input en toda celda sin texto real (p. ej. SECTOR/FIRMA con subrayado vacío)."""

    def repl(m: re.Match[str]) -> str:
        attrs = m.group(1) or ""
        inner = m.group(2) or ""
        if "data-sgi-field" in inner or "sgi-rec-field-wrap" in inner:
            return m.group(0)
        if re.search(r"<t[dh]\b", inner, re.I):
            return m.group(0)
        visible = _td_visible_text(inner)
        if visible:
            return m.group(0)
        return f"<td{attrs}>{make_control()}</td>"

    return _TD_CELL_RE.sub(repl, html_in)


def _inject_fields_into_docx_html(raw_html: str, fields: list[dict[str, Any]]) -> str:
    """Inserta inputs junto a cada etiqueta del Word (prioridad: label → celda vacía)."""
    text_fields = [f for f in fields if f.get("type") not in ("editable_table", "calculated")]
    by_label: dict[str, dict[str, Any]] = {}
    for f in text_fields:
        key = _norm_label_key(str(f.get("label") or ""))
        if key and key not in by_label:
            by_label[key] = f

    placed: set[str] = set()
    auto_i = 0

    def mark_from(snippet: str) -> str:
        for m in re.finditer(r'data-sgi-field="([^"]+)"', snippet):
            placed.add(m.group(1))
        return snippet

    def control_for_label(label_text: str) -> str:
        nonlocal auto_i
        key = _norm_label_key(label_text)
        f = by_label.get(key)
        if f is None:
            for lab, cand in by_label.items():
                if cand["name"] in placed:
                    continue
                if key and (key in lab or lab in key):
                    f = cand
                    break
        if f is not None and f["name"] not in placed:
            placed.add(f["name"])
            return mark_from(
                _wrap_field_control(_field_input_from_meta(f["name"], f.get("type") or "text"))
            )
        auto_i += 1
        name = f"campo_auto_{auto_i}"
        placed.add(name)
        return mark_from(_wrap_field_control(_input_html(name, input_type="text")))

    def auto_control(prefix: str = "celda") -> str:
        nonlocal auto_i
        auto_i += 1
        name = f"{prefix}_{auto_i}"
        placed.add(name)
        return mark_from(_wrap_field_control(_input_html(name, input_type="text")))

    out = raw_html

    # 1) Primero: cada etiqueta conocida (Fecha, Lugar, Tema…) → input al lado
    for lab_key, f in sorted(by_label.items(), key=lambda kv: -len(kv[0])):
        if f["name"] in placed:
            continue
        label = (f.get("label") or "").strip()
        if len(label) < 2:
            continue
        esc = re.escape(label)
        pattern = re.compile(
            rf"({esc}\s*:?\s*)(?!(?:(?:&nbsp;|&#160;|\s|\xa0)|</?(?:span|strong|em|b|i|u)[^>]*>)*"
            rf"<(?:input|textarea|select|span)\b)",
            re.I,
        )

        def _after_label(m: re.Match[str], _f: dict[str, Any] = f) -> str:
            if _f["name"] in placed:
                return m.group(0)
            placed.add(_f["name"])
            ctrl = _wrap_field_control(
                _field_input_from_meta(_f["name"], _f.get("type") or "text")
            )
            return m.group(1) + mark_from(ctrl)

        new_out, n = pattern.subn(_after_label, out, count=1)
        if n:
            out = new_out

    # 2) "Algo:" al final de párrafo/celda (sin input aún) → input genérico o de campo
    def label_colon_tail(m: re.Match[str]) -> str:
        label_txt = re.sub(r"<[^>]+>", "", m.group("label"))
        return f'{m.group("label")}{m.group("colon")}{control_for_label(label_txt)}{m.group("end")}'

    out = re.sub(
        r"(?P<label>(?:^|>)[^<>]{0,80}?)"
        r"(?P<colon>:\s*)"
        r"(?P<blank>(?:&nbsp;|&#160;|\s|\xa0)*)"
        r"(?P<end></(?:p|span|strong|em|td|th|div|li)>)",
        label_colon_tail,
        out,
        flags=re.I,
    )

    # 3) Blancos tipo ___ / □
    def repl_blank(_m: re.Match[str]) -> str:
        return auto_control("campo_auto")

    out = re.sub(r"_{3,}|\.{4,}|□|☐|\[\s*\]", repl_blank, out)

    # 4) Celdas sin texto real (incluye <u></u>, spans vacíos, etc.) — no usan campos con nombre
    out = _fill_blank_table_cells(out, lambda: auto_control("celda"))

    # 5) Campos con nombre que no encontraron etiqueta
    missing = [
        f
        for f in text_fields
        if f.get("name") and f["name"] not in placed and f'data-sgi-field="{f["name"]}"' not in out
    ]
    if missing:
        rows = []
        for f in missing:
            label = html.escape(f.get("label") or f["name"])
            rows.append(
                "<tr>"
                f'<td class="sgi-rec-label">{label}</td>'
                f"<td>{_wrap_field_control(_field_input_from_meta(f['name'], f.get('type') or 'text'))}</td>"
                "</tr>"
            )
            placed.add(f["name"])
        out += (
            '<div class="sgi-rec-extra-fields">'
            '<table class="sgi-rec-doc-table"><tbody>'
            + "".join(rows)
            + "</tbody></table></div>"
        )

    for f in fields:
        if f.get("type") != "editable_table":
            continue
        name = html.escape(f["name"], quote=True)
        label = html.escape(f.get("label") or "Tabla")
        cols = f.get("columns") or []
        heads = "".join(f"<th>{html.escape(c.get('label') or c.get('key') or '')}</th>" for c in cols)
        marker = f'data-sgi-table="{name}"'
        if marker in out:
            continue
        out += (
            f'<div class="sgi-rec-table-block" data-sgi-table="{name}">'
            f"<p><strong>{label}</strong></p>"
            f'<table class="sgi-rec-doc-table"><thead><tr>{heads}</tr></thead>'
            f'<tbody data-sgi-table-body="{name}"></tbody></table>'
            f'<button type="button" class="sgi-rec-add-row sgi-proc-no-print" data-sgi-table-add="{name}">+ Fila</button>'
            f"</div>"
        )
    return f'<!--sgi-layout-v:4-->\n{out}'



def build_docx_layout(data: bytes, fields: list[dict[str, Any]]) -> str:
    """Layout fiel al adjunto: encabezado Word + cuerpo + pie."""
    header = _docx_header_html(data)
    body = _inject_fields_into_docx_html(_docx_bytes_to_html(data), fields)
    footer = _docx_footer_html(data)
    parts = [p for p in (header, body, footer) if p and p.strip()]
    return "\n".join(parts) if parts else "<p></p>"


def build_xlsx_layout(data: bytes, fields: list[dict[str, Any]]) -> str:
    """Reproduce hojas Excel como tablas documentales editables (no formulario Bootstrap)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "<p>No se pudo generar la vista del Excel.</p>"

    try:
        wb = load_workbook(io.BytesIO(data), data_only=False, read_only=True)
    except Exception:
        return "<p>No se pudo generar la vista del Excel.</p>"

    table_fields = {f["name"]: f for f in fields if f.get("type") == "editable_table"}
    text_fields = [f for f in fields if f.get("type") not in ("editable_table", "calculated")]
    text_by_label = {(f.get("label") or "").strip().lower(): f for f in text_fields}

    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f'<h2 class="sgi-rec-sheet-title">{html.escape(sheet_name)}</h2>')
            rows = []
            for row in ws.iter_rows(max_row=60, max_col=20, values_only=False):
                vals = list(row)
                if any(c.value is not None for c in vals):
                    rows.append(vals)
                if len(rows) >= 40:
                    break
            if not rows:
                continue

            # Si hay tabla editable asociada a esta hoja, renderizarla como grilla documental
            tname = f"tabla_{_slug(sheet_name)}"
            tf = table_fields.get(tname)
            if tf and len(rows) >= 1:
                cols = tf.get("columns") or []
                heads = "".join(f"<th>{html.escape(c.get('label') or '')}</th>" for c in cols)
                parts.append(
                    f'<div class="sgi-rec-table-block" data-sgi-table="{html.escape(tname, quote=True)}">'
                    f'<table class="sgi-rec-doc-table sgi-rec-excel-table"><thead><tr>{heads}</tr></thead>'
                    f'<tbody data-sgi-table-body="{html.escape(tname, quote=True)}"></tbody></table>'
                    f'<button type="button" class="sgi-rec-add-row sgi-proc-no-print" data-sgi-table-add="{html.escape(tname, quote=True)}">+ Fila</button>'
                    f"</div>"
                )
                continue

            # Grilla celda a celda conservando valores fijos y editando vacíos / etiquetas
            parts.append('<table class="sgi-rec-doc-table sgi-rec-excel-table">')
            for ri, row in enumerate(rows):
                parts.append("<tr>")
                skip_next: set[int] = set()
                for ci, cell in enumerate(row):
                    if ci in skip_next:
                        continue
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        parts.append(f'<td class="sgi-rec-formula">{html.escape(val)}</td>')
                        continue
                    if val is not None and str(val).strip() != "":
                        label = str(val).strip().rstrip(":")
                        nxt = row[ci + 1].value if ci + 1 < len(row) else "x"
                        if ci + 1 < len(row) and (nxt is None or nxt == ""):
                            fld = text_by_label.get(label.lower())
                            name = fld["name"] if fld else _slug(label)
                            ftype = (fld or {}).get("type") or infer_field_type(label)
                            parts.append(f'<td class="sgi-rec-label">{html.escape(label)}</td>')
                            itype = "yes_no" if ftype == "yes_no" else ("date" if ftype == "date" else "text")
                            if ftype in ("textarea", "observations"):
                                itype = "textarea"
                            parts.append(f"<td>{_input_html(name, input_type=itype)}</td>")
                            skip_next.add(ci + 1)
                            continue
                        parts.append(f"<td>{html.escape(str(val))}</td>")
                    else:
                        parts.append(f"<td>{_input_html(f'c_{ri}_{ci}')}</td>")
                parts.append("</tr>")
            parts.append("</table>")
    finally:
        wb.close()

    return "\n".join(parts) if parts else "<p></p>"


def analyze_docx(data: bytes) -> dict[str, Any]:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            xml = zf.read("word/document.xml")
    except (KeyError, zipfile.BadZipFile) as exc:
        raise ImportSecurityError("No se pudo leer el archivo Word.") from exc

    root = ET.fromstring(xml)
    paragraphs = _docx_paragraph_texts(root)
    tables = _docx_tables(root)

    sections: list[dict[str, Any]] = [{"id": "sec_general", "title": "Datos generales", "order": 0}]
    fields: list[dict[str, Any]] = []
    warnings: list[str] = []
    order = 1
    used_names: set[str] = set()

    blank_re = re.compile(r"_{3,}|:{1}\s*$|\.{3,}")
    label_value_re = re.compile(r"^([^:]{2,80})\s*:\s*(.*)$")
    yes_no_re = re.compile(r"\b(s[ií]\s*/\s*no|yes\s*/\s*no)\b", re.I)

    current_section = "Datos generales"
    for line in paragraphs:
        # Títulos cortos en mayúsculas → sección
        if len(line) <= 80 and line.isupper() and not blank_re.search(line):
            current_section = line.title()
            sections.append({"id": f"sec_{_slug(line)}", "title": current_section, "order": len(sections)})
            continue

        m = label_value_re.match(line)
        if m:
            label = m.group(1).strip()
            rest = (m.group(2) or "").strip()
            name = _slug(label)
            if name in used_names:
                name = f"{name}_{order}"
            used_names.add(name)
            ftype = "yes_no" if yes_no_re.search(line) else infer_field_type(label)
            if not rest or blank_re.search(rest) or rest in ("☐", "□", "[ ]"):
                fields.append(
                    _field(name=name, label=label, ftype=ftype, order=order, section=current_section)
                )
                order += 1
            continue

        if blank_re.search(line) or "☐" in line or "[ ]" in line:
            label = blank_re.sub("", line).strip(" :.-") or f"Campo {order}"
            name = _slug(label)
            if name in used_names:
                name = f"{name}_{order}"
            used_names.add(name)
            ftype = "yes_no" if yes_no_re.search(line) or "☐" in line else infer_field_type(label)
            fields.append(_field(name=name, label=label, ftype=ftype, order=order, section=current_section))
            order += 1

    for ti, table in enumerate(tables):
        if not table:
            continue
        header = table[0]
        if len(table) >= 2 and all(header):
            cols = []
            for ci, h in enumerate(header):
                key = _slug(h or f"col_{ci+1}", fallback=f"col_{ci+1}")
                cols.append({"key": key, "label": h or f"Columna {ci+1}", "type": infer_field_type(h), "required": False})
            name = f"tabla_{ti+1}"
            fields.append(
                _field(
                    name=name,
                    label=f"Tabla {ti + 1}",
                    ftype="editable_table",
                    order=order,
                    section=current_section,
                    columns=cols,
                )
            )
            order += 1
        else:
            for ri, row in enumerate(table):
                if not row:
                    continue
                label = row[0] if row[0] else f"Fila {ri+1}"
                name = _slug(label)
                if name in used_names:
                    name = f"{name}_{order}"
                used_names.add(name)
                fields.append(
                    _field(
                        name=name,
                        label=label,
                        ftype=infer_field_type(label),
                        order=order,
                        section=current_section,
                    )
                )
                order += 1

    if not fields:
        warnings.append("No se encontraron campos editables. Se creó un campo de observaciones.")
        fields.append(
            _field(
                name="observaciones",
                label="Observaciones",
                ftype="textarea",
                order=1,
                section="Datos generales",
            )
        )

    title = paragraphs[0][:120] if paragraphs else "Registro importado"
    layout_html = build_docx_layout(data, fields)
    return {
        "detectedType": "word",
        "confidence": 0.75 if len(fields) > 1 else 0.45,
        "suggestedName": title,
        "suggestedCode": "",
        "warnings": warnings,
        "sections": sections,
        "fields": fields,
        "tables": [f for f in fields if f["type"] == "editable_table"],
        "formulas": [],
        "layoutHtml": layout_html,
        "layoutMode": "document",
    }


# --- Excel ---

def _convert_formula(raw: str) -> tuple[str | None, str | None]:
    """Traduce fórmulas simples a DSL seguro. Sin eval()."""
    if not raw or not str(raw).startswith("="):
        return None, None
    expr = str(raw)[1:].strip().upper()
    # Referencias externas / peligrosas
    if "!" in expr and ("[" in expr or "HTTP" in expr or "\\\\" in expr):
        return None, "Fórmula con referencia externa no soportada."
    if any(x in expr for x in ("CMD|", "EXEC(", "HYPERLINK(", "WEBSERVICE(", "CALL(")):
        return None, "Fórmula insegura rechazada."
    m = re.match(r"^([A-ZÁÉÍÓÚÑ\.]+)\((.*)\)$", expr)
    if not m:
        if re.fullmatch(r"[\d\.\+\-\*/\(\)\sA-Z]+", expr):
            return f"math:{expr}", None
        return None, f"Fórmula no convertida: {raw}"
    func, args = m.group(1), m.group(2)
    func_norm = func.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    if func not in _SAFE_FORMULA_FUNCS and func_norm not in _SAFE_FORMULA_FUNCS:
        return None, f"Función no soportada: {func}"
    return f"{func_norm}({args})", None


def analyze_xlsx(data: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportSecurityError("openpyxl no está disponible en el servidor.") from exc

    try:
        wb = load_workbook(io.BytesIO(data), data_only=False, read_only=True)
    except Exception as exc:
        msg = str(exc).lower()
        if "password" in msg or "encrypted" in msg:
            raise ImportSecurityError("El archivo Excel está protegido por contraseña.") from exc
        raise ImportSecurityError("No se pudo leer el archivo Excel.") from exc

    sections: list[dict[str, Any]] = []
    fields: list[dict[str, Any]] = []
    formulas: list[dict[str, Any]] = []
    warnings: list[str] = []
    order = 1
    used_names: set[str] = set()
    sheet_names = list(wb.sheetnames)
    suggested = sheet_names[0] if sheet_names else "Registro Excel"

    try:
        for sheet_idx, sheet_name in enumerate(sheet_names):
            ws = wb[sheet_name]
            section_title = sheet_name
            sections.append({"id": f"sec_{_slug(sheet_name)}", "title": section_title, "order": sheet_idx})

            rows = []
            for row in ws.iter_rows(max_row=80, max_col=40, values_only=False):
                values = list(row)
                if any(c.value is not None for c in values):
                    rows.append(values)
                if len(rows) >= 40:
                    break

            if not rows:
                continue

            header_cells = rows[0]
            headers = [(c.value if c.value is not None else "") for c in header_cells]
            headers_str = [str(h).strip() for h in headers]
            looks_like_table = sum(1 for h in headers_str if h) >= 2 and len(rows) >= 2

            if looks_like_table:
                cols = []
                for ci, h in enumerate(headers_str):
                    if not h and all(not (r[ci].value if ci < len(r) else None) for r in rows[1:5]):
                        continue
                    label = h or f"Columna {ci + 1}"
                    key = _slug(label, fallback=f"col_{ci+1}")
                    cols.append(
                        {
                            "key": key,
                            "label": label,
                            "type": infer_field_type(label),
                            "required": False,
                        }
                    )
                if cols:
                    name = f"tabla_{_slug(sheet_name)}"
                    if name in used_names:
                        name = f"{name}_{order}"
                    used_names.add(name)
                    fields.append(
                        _field(
                            name=name,
                            label=f"Tabla · {sheet_name}",
                            ftype="editable_table",
                            order=order,
                            section=section_title,
                            columns=cols,
                        )
                    )
                    order += 1

            for ri, row in enumerate(rows[:25]):
                for ci, cell in enumerate(row):
                    val = cell.value
                    if val is None:
                        continue
                    if isinstance(val, str) and val.startswith("="):
                        converted, warn = _convert_formula(val)
                        coord = getattr(cell, "coordinate", f"R{ri}C{ci}")
                        if converted:
                            label = f"Calculado {coord}"
                            name = _slug(f"calc_{coord}")
                            if name in used_names:
                                name = f"{name}_{order}"
                            used_names.add(name)
                            fld = _field(
                                name=name,
                                label=label,
                                ftype="calculated",
                                order=order,
                                section=section_title,
                                formula=converted,
                            )
                            fld["mode"] = "readonly"
                            fields.append(fld)
                            formulas.append({"cell": coord, "original": val, "converted": converted})
                            order += 1
                        else:
                            warnings.append(warn or f"Fórmula no convertida en {coord}.")
                            formulas.append(
                                {"cell": coord, "original": val, "converted": None, "warning": warn}
                            )
                        continue

                    if isinstance(val, str) and ci + 1 < len(row):
                        nxt = row[ci + 1].value
                        label = val.strip().rstrip(":")
                        if label and (nxt is None or nxt == ""):
                            name = _slug(label)
                            if name in used_names:
                                continue
                            used_names.add(name)
                            fields.append(
                                _field(
                                    name=name,
                                    label=label,
                                    ftype=infer_field_type(label),
                                    order=order,
                                    section=section_title,
                                )
                            )
                            order += 1
    finally:
        wb.close()

    if not fields:
        warnings.append("No se encontraron campos editables. Se creó un campo de observaciones.")
        fields.append(
            _field(
                name="observaciones",
                label="Observaciones",
                ftype="textarea",
                order=1,
                section="Datos generales",
            )
        )
        if not sections:
            sections = [{"id": "sec_general", "title": "Datos generales", "order": 0}]

    return {
        "detectedType": "excel",
        "confidence": 0.8 if len(fields) > 1 else 0.5,
        "suggestedName": suggested,
        "suggestedCode": "",
        "warnings": warnings,
        "sections": sections,
        "fields": fields,
        "tables": [f for f in fields if f["type"] == "editable_table"],
        "formulas": formulas,
        "layoutHtml": build_xlsx_layout(data, fields),
        "layoutMode": "document",
    }


def analyze_office_bytes(data: bytes, ext: str) -> dict[str, Any]:
    assert_safe_office_bytes(data, ext)
    if ext == ".docx":
        return analyze_docx(data)
    if ext in (".xlsx", ".xls"):
        if ext == ".xls":
            # openpyxl no lee .xls binario; intentar y fallar claro
            try:
                return analyze_xlsx(data)
            except ImportSecurityError:
                raise
            except Exception as exc:
                raise ImportSecurityError(
                    "El formato .xls antiguo no pudo interpretarse. Guarde el archivo como .xlsx."
                ) from exc
        return analyze_xlsx(data)
    raise ImportSecurityError("Tipo de archivo no soportado.")
