from __future__ import annotations

from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from flask import current_app
from sqlalchemy import Select, func, or_, select
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models.sgi import (
    ESTADO_BORRADOR,
    ESTADO_LABELS,
    ESTADOS_DOCUMENTO,
    SLUG_TO_TIPO,
    TIPO_LABELS,
    TIPO_SLUGS,
    TIPOS_DOCUMENTO,
    SgiDocumento,
    SgiDocumentoHistorial,
)
from app.models.user import User
from app.services import security_audit_service as audit_svc
from app.services.upload_paths import uploads_workspace_root

ACCION_ALTA = "alta"
ACCION_EDICION = "edicion"
ACCION_CAMBIO_ESTADO = "cambio_estado"
ACCION_ARCHIVO = "archivo_adjunto"
ACCION_BAJA = "baja"
ACCION_RECUPERACION = "recuperacion"

ESTADOS_ELIMINABLES: frozenset[str] = frozenset(
    {ESTADO_BORRADOR, "en_revision", "revisado"}
)

ALLOWED_ATTACHMENT_EXTENSIONS = frozenset({".pdf"})


def tipo_from_slug(slug: str | None) -> str | None:
    s = (slug or "").strip().lower()
    return SLUG_TO_TIPO.get(s)


def slug_from_tipo(tipo: str | None) -> str | None:
    t = (tipo or "").strip().upper()
    return TIPO_SLUGS.get(t)


def parse_iso_date(raw: str | None) -> date | None:
    s = (raw or "").strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def filter_args_from_request(values: Any, *, tipo_fijo: str | None = None) -> dict[str, Any]:
    return {
        "tipo": tipo_fijo or (values.get("tipo") or "").strip().upper(),
        "q": (values.get("q") or "").strip(),
        "estado": (values.get("estado") or "").strip(),
        "fecha_desde": values.get("fecha_desde"),
        "fecha_hasta": values.get("fecha_hasta"),
    }


def build_filtered_query(args: dict[str, Any]) -> Select[Any]:
    tipo = (args.get("tipo") or "").strip().upper()
    q_text = (args.get("q") or "").strip()
    estado = (args.get("estado") or "").strip()
    fd = parse_iso_date(args.get("fecha_desde"))
    fh = parse_iso_date(args.get("fecha_hasta"))

    q = select(SgiDocumento).order_by(SgiDocumento.codigo, SgiDocumento.id)

    if tipo and tipo in TIPOS_DOCUMENTO:
        q = q.where(SgiDocumento.tipo == tipo)

    if q_text:
        like = f"%{q_text}%"
        q = q.where(
            or_(
                SgiDocumento.codigo.ilike(like),
                SgiDocumento.titulo.ilike(like),
                SgiDocumento.tipo.ilike(like),
            )
        )

    if estado and estado in ESTADOS_DOCUMENTO:
        q = q.where(SgiDocumento.estado == estado)

    if fd is not None:
        q = q.where(
            or_(
                SgiDocumento.fecha_creacion_doc >= fd,
                func.date(SgiDocumento.created_at) >= fd,
            )
        )
    if fh is not None:
        q = q.where(
            or_(
                SgiDocumento.fecha_creacion_doc <= fh,
                func.date(SgiDocumento.created_at) <= fh,
            )
        )

    return _exclude_deleted(q)


def fetch_list(args: dict[str, Any]) -> list[SgiDocumento]:
    return list(db.session.scalars(build_filtered_query(args)).all())


def normalize_persona_campo(raw: str | None) -> str:
    """Nombre o área de elaboración/revisión/aprobación en mayúsculas."""
    return (raw or "").strip().upper()


def _apply_upper_text_attr(obj: Any, attr: str, *, max_len: int | None = None) -> bool:
    val = (getattr(obj, attr, None) or "").strip()
    if not val:
        return False
    upper = val.upper()
    if upper == val:
        return False
    setattr(obj, attr, upper[:max_len] if max_len else upper)
    return True


def ensure_documento_nombres_mayusculas(doc: SgiDocumento) -> bool:
    """Normaliza código, título y responsables del documento a mayúsculas."""
    changed = False
    for attr, max_len in (
        ("codigo", 64),
        ("titulo", 512),
        ("responsable_elaboracion", 256),
        ("responsable_revision", 256),
        ("responsable_aprobacion", 256),
    ):
        if _apply_upper_text_attr(doc, attr, max_len=max_len):
            changed = True
    return changed


def ensure_list_nombres_mayusculas(rows: list[SgiDocumento]) -> None:
    """Aplica mayúsculas en lote (p. ej. al abrir un listado con datos antiguos)."""
    changed = any(ensure_documento_nombres_mayusculas(doc) for doc in rows)
    if changed:
        db.session.commit()


def _exclude_deleted(q: Select[Any]) -> Select[Any]:
    return q.where(SgiDocumento.deleted_at.is_(None))


def _only_deleted(q: Select[Any]) -> Select[Any]:
    return q.where(SgiDocumento.deleted_at.is_not(None))


def documento_esta_eliminado(doc: SgiDocumento | None) -> bool:
    return doc is not None and doc.deleted_at is not None


def documento_codigo_visible(doc: SgiDocumento) -> str:
    return (doc.codigo_archivado or doc.codigo or "").strip()


def puede_eliminar_documento(doc: SgiDocumento | None) -> bool:
    if doc is None or documento_esta_eliminado(doc):
        return False
    return (doc.estado or "").strip().lower() in ESTADOS_ELIMINABLES


def get_documento(doc_id: int, *, incluir_eliminados: bool = False) -> SgiDocumento | None:
    row = db.session.get(SgiDocumento, int(doc_id))
    if row is None:
        return None
    if not incluir_eliminados and documento_esta_eliminado(row):
        return None
    return row


def historial_for(doc_id: int) -> list[SgiDocumentoHistorial]:
    return list(
        db.session.scalars(
            select(SgiDocumentoHistorial)
            .where(SgiDocumentoHistorial.documento_id == int(doc_id))
            .order_by(SgiDocumentoHistorial.fecha.desc(), SgiDocumentoHistorial.id.desc())
        ).all()
    )


def append_historial(documento_id: int, usuario_label: str, accion: str, detalle: str) -> None:
    db.session.add(
        SgiDocumentoHistorial(
            documento_id=int(documento_id),
            fecha=datetime.now(timezone.utc),
            usuario=(usuario_label or "").strip()[:256],
            accion=(accion or "").strip()[:64],
            detalle=(detalle or "").strip()[:8000],
        )
    )


def _record_security_audit(
    *,
    action: str,
    actor: User | None,
    entity_id: int | None,
    old_value: str | None = None,
    new_value: str | None = None,
    detail: str | None = None,
) -> None:
    audit_svc.record_event(
        action=action,
        module="sgi",
        actor=actor,
        entity_type="sgi_documento",
        entity_id=entity_id,
        old_value=old_value,
        new_value=new_value,
        detail=detail,
    )


def _codigo_duplicado(tipo: str, codigo: str, exclude_id: int | None = None) -> bool:
    q = select(func.count()).select_from(SgiDocumento).where(
        SgiDocumento.tipo == tipo,
        func.lower(SgiDocumento.codigo) == codigo.lower(),
        SgiDocumento.deleted_at.is_(None),
    )
    if exclude_id is not None:
        q = q.where(SgiDocumento.id != int(exclude_id))
    return int(db.session.scalar(q) or 0) > 0


def _codigo_archivado_en_papelera(tipo: str, codigo: str, exclude_id: int | None = None) -> bool:
    q = select(func.count()).select_from(SgiDocumento).where(
        SgiDocumento.tipo == tipo,
        func.lower(SgiDocumento.codigo_archivado) == codigo.lower(),
        SgiDocumento.deleted_at.is_not(None),
    )
    if exclude_id is not None:
        q = q.where(SgiDocumento.id != int(exclude_id))
    return int(db.session.scalar(q) or 0) > 0


def _parse_form(data: dict[str, Any], *, tipo_fijo: str | None = None) -> tuple[dict[str, Any] | None, str | None]:
    tipo = (tipo_fijo or (data.get("tipo") or "")).strip().upper()
    codigo = (data.get("codigo") or "").strip().upper()
    titulo = (data.get("titulo") or "").strip().upper()
    revision = (data.get("revision") or "").strip()
    estado = (data.get("estado") or ESTADO_BORRADOR).strip().lower()
    responsable_elab = normalize_persona_campo(data.get("responsable_elaboracion"))
    responsable_aprob = normalize_persona_campo(data.get("responsable_aprobacion"))
    observaciones = (data.get("observaciones") or "").strip()

    if tipo not in TIPOS_DOCUMENTO:
        return None, "Tipo de documento inválido."
    if len(codigo) < 1:
        return None, "El código del documento es obligatorio."
    if len(codigo) > 64:
        return None, "El código no puede superar 64 caracteres."
    if len(titulo) < 2:
        return None, "El título debe tener al menos 2 caracteres."
    if estado not in ESTADOS_DOCUMENTO:
        return None, "Estado inválido."

    return {
        "tipo": tipo,
        "codigo": codigo,
        "titulo": titulo[:512],
        "revision": revision[:32],
        "fecha_creacion_doc": parse_iso_date(data.get("fecha_creacion_doc")),
        "fecha_ultima_revision": parse_iso_date(data.get("fecha_ultima_revision")),
        "responsable_elaboracion": responsable_elab[:256],
        "responsable_aprobacion": responsable_aprob[:256],
        "estado": estado,
        "observaciones": observaciones[:8000],
    }, None


def create_documento(
    data: dict[str, Any],
    user_id: int,
    actor_label: str,
    *,
    tipo_fijo: str | None = None,
    actor: User | None = None,
) -> tuple[SgiDocumento | None, str | None]:
    parsed, err = _parse_form(data, tipo_fijo=tipo_fijo)
    if err or parsed is None:
        return None, err

    if _codigo_duplicado(parsed["tipo"], parsed["codigo"]):
        return None, f"Ya existe un documento {parsed['tipo']} con el código «{parsed['codigo']}»."

    row = SgiDocumento(
        **parsed,
        created_by_id=user_id,
        updated_by_id=user_id,
    )
    db.session.add(row)
    db.session.flush()
    append_historial(
        row.id,
        actor_label,
        ACCION_ALTA,
        f"{parsed['tipo']} {parsed['codigo']} — {parsed['titulo']}",
    )
    db.session.commit()
    db.session.refresh(row)

    _record_security_audit(
        action="sgi_documento_create",
        actor=actor,
        entity_id=row.id,
        new_value=f"{row.tipo}:{row.codigo}",
        detail=row.titulo,
    )
    return row, None


def update_documento(
    doc_id: int,
    data: dict[str, Any],
    user_id: int,
    actor_label: str,
    *,
    tipo_fijo: str | None = None,
    actor: User | None = None,
) -> tuple[bool, str]:
    row = get_documento(doc_id)
    if row is None:
        return False, "Documento no encontrado."

    parsed, err = _parse_form(data, tipo_fijo=tipo_fijo or row.tipo)
    if err or parsed is None:
        return False, err or "Datos inválidos."

    if parsed["tipo"] != row.tipo:
        return False, "No se puede cambiar el tipo de un documento existente."

    if _codigo_duplicado(parsed["tipo"], parsed["codigo"], exclude_id=row.id):
        return False, f"Ya existe otro documento {parsed['tipo']} con el código «{parsed['codigo']}»."

    old_estado = row.estado
    old_codigo = row.codigo
    changes: list[str] = []

    for field, value in parsed.items():
        if getattr(row, field) != value:
            changes.append(field)
            setattr(row, field, value)

    row.updated_at = datetime.now(timezone.utc)
    row.updated_by_id = user_id

    if changes:
        append_historial(
            row.id,
            actor_label,
            ACCION_EDICION,
            f"Campos modificados: {', '.join(changes)}",
        )
        if "estado" in changes:
            append_historial(
                row.id,
                actor_label,
                ACCION_CAMBIO_ESTADO,
                f"{ESTADO_LABELS.get(old_estado, old_estado)} → {ESTADO_LABELS.get(row.estado, row.estado)}",
            )
            _record_security_audit(
                action="sgi_documento_estado_change",
                actor=actor,
                entity_id=row.id,
                old_value=old_estado,
                new_value=row.estado,
                detail=f"{row.tipo}:{row.codigo}",
            )

    if row.es_procedimiento_visual:
        from app.services import sgi_procedimiento_service as proc_svc

        if proc_svc.ensure_visual_documento_titulo_sync(row):
            if "titulo" not in changes:
                changes.append("titulo")

    db.session.commit()

    _record_security_audit(
        action="sgi_documento_update",
        actor=actor,
        entity_id=row.id,
        old_value=old_codigo,
        new_value=row.codigo,
        detail=", ".join(changes) if changes else "sin cambios",
    )
    return True, "Documento actualizado." if changes else "Sin cambios."


def delete_documento(
    doc_id: int,
    actor_label: str,
    *,
    actor: User | None = None,
) -> tuple[bool, str]:
    """Mueve el documento a la papelera (soft delete)."""
    row = get_documento(doc_id, incluir_eliminados=True)
    if row is None:
        return False, "Documento no encontrado."
    if documento_esta_eliminado(row):
        return False, "El documento ya está en la papelera."
    if not puede_eliminar_documento(row):
        return False, "Solo se pueden eliminar documentos en borrador o en curso de revisión."

    codigo_visible = documento_codigo_visible(row)
    summary = f"{row.tipo} {codigo_visible} — {row.titulo}"
    row.codigo_archivado = codigo_visible[:64]
    suffix = f"__ELIM_{row.id}"
    base = codigo_visible[: max(1, 64 - len(suffix))]
    row.codigo = f"{base}{suffix}"[:64]
    row.deleted_at = datetime.now(timezone.utc)
    row.deleted_by_id = actor.id if actor is not None else None
    row.updated_at = datetime.now(timezone.utc)
    row.updated_by_id = actor.id if actor is not None else None

    append_historial(row.id, actor_label, ACCION_BAJA, f"Movido a papelera: {summary}")
    _record_security_audit(
        action="sgi_documento_delete",
        actor=actor,
        entity_id=row.id,
        old_value=summary,
        detail="soft_delete",
    )
    db.session.commit()
    return True, f"Documento movido a la papelera: {summary}"


def restore_documento(
    doc_id: int,
    actor_label: str,
    *,
    actor: User | None = None,
) -> tuple[bool, str]:
    row = get_documento(doc_id, incluir_eliminados=True)
    if row is None:
        return False, "Documento no encontrado."
    if not documento_esta_eliminado(row):
        return False, "El documento no está en la papelera."

    codigo_restaurar = (row.codigo_archivado or "").strip().upper()
    if not codigo_restaurar:
        return False, "No se pudo recuperar el código original del documento."
    if _codigo_duplicado(row.tipo, codigo_restaurar, exclude_id=row.id):
        return False, f"Ya existe un documento activo con código {codigo_restaurar}."

    summary = f"{row.tipo} {codigo_restaurar} — {row.titulo}"
    row.codigo = codigo_restaurar[:64]
    row.codigo_archivado = None
    row.deleted_at = None
    row.deleted_by_id = None
    row.updated_at = datetime.now(timezone.utc)
    row.updated_by_id = actor.id if actor is not None else None

    append_historial(row.id, actor_label, ACCION_RECUPERACION, f"Recuperado desde papelera: {summary}")
    _record_security_audit(
        action="sgi_documento_restore",
        actor=actor,
        entity_id=row.id,
        new_value=summary,
    )
    db.session.commit()
    return True, f"Documento recuperado: {summary}"


def fetch_deleted_list(args: dict[str, Any]) -> list[SgiDocumento]:
    tipo = (args.get("tipo") or "").strip().upper()
    q_text = (args.get("q") or "").strip()

    q = select(SgiDocumento).order_by(SgiDocumento.deleted_at.desc(), SgiDocumento.id.desc())
    q = _only_deleted(q)

    if tipo and tipo in TIPOS_DOCUMENTO:
        q = q.where(SgiDocumento.tipo == tipo)

    if q_text:
        like = f"%{q_text}%"
        q = q.where(
            or_(
                SgiDocumento.codigo_archivado.ilike(like),
                SgiDocumento.codigo.ilike(like),
                SgiDocumento.titulo.ilike(like),
            )
        )

    return list(db.session.scalars(q).all())


def _upload_max_bytes() -> int:
    try:
        return int(current_app.config.get("MAX_CONTENT_LENGTH") or 12 * 1024 * 1024)
    except (TypeError, ValueError):
        return 12 * 1024 * 1024


def allowed_attachment_suffix(filename: str) -> tuple[bool, str | None]:
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_ATTACHMENT_EXTENSIONS:
        return False, "Solo se permiten archivos PDF."
    return True, None


def save_attachment(
    doc_id: int,
    storage: FileStorage | None,
    user_id: int,
    actor_label: str,
    *,
    actor: User | None = None,
) -> tuple[bool, str]:
    row = get_documento(doc_id)
    if row is None:
        return False, "Documento no encontrado."
    if not storage or not (storage.filename or "").strip():
        return False, "No se seleccionó ningún archivo."

    ok_fn, fn_err = allowed_attachment_suffix(storage.filename or "")
    if not ok_fn:
        return False, fn_err or "Archivo no permitido."

    data = storage.read()
    size = len(data)
    if size > _upload_max_bytes():
        return False, "El archivo supera el tamaño máximo permitido."

    fn = secure_filename(storage.filename or "documento.pdf")
    if not fn:
        return False, "Nombre de archivo inválido."
    if not fn.lower().endswith(".pdf"):
        fn = f"{fn}.pdf"
    stem, _, ext = fn.rpartition(".")
    fn = f"{stem.upper()}.{ext.lower()}" if ext else stem.upper()

    base = uploads_workspace_root() / "sgi" / str(row.id)
    base.mkdir(parents=True, exist_ok=True)
    dest = base / fn
    dest.write_bytes(data)

    rel = Path("sgi") / str(row.id) / fn
    row.archivo_path = rel.as_posix()
    row.updated_at = datetime.now(timezone.utc)
    row.updated_by_id = user_id
    append_historial(row.id, actor_label, ACCION_ARCHIVO, f"Adjunto PDF: {fn} ({size} bytes)")
    db.session.commit()

    _record_security_audit(
        action="sgi_documento_attachment",
        actor=actor,
        entity_id=row.id,
        detail=fn,
    )
    return True, "Archivo PDF guardado."


def attachment_absolute_path(rel: str | None) -> Path | None:
    if not rel:
        return None
    from app.services.upload_paths import resolve_under_upload_roots

    return resolve_under_upload_roots(Path(rel))


def estado_visual_row(doc: SgiDocumento) -> str:
    if doc.estado == ESTADO_BORRADOR:
        return "sgi-row-borrador"
    if doc.estado in ("en_revision", "revisado"):
        return "sgi-row-revision"
    if doc.estado == "obsoleto":
        return "sgi-row-obsoleto"
    if doc.estado in ("aprobado", ESTADO_VIGENTE):
        return "sgi-row-vigente"
    return "sgi-row-borrador"


def build_export_xlsx(rows: list[SgiDocumento], *, tipo_label: str = "SGI") -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = tipo_label[:31]
    headers = [
        "Tipo",
        "Código",
        "Título",
        "Revisión",
        "Fecha creación doc.",
        "Fecha última revisión",
        "Resp. elaboración",
        "Resp. aprobación",
        "Estado",
        "Observaciones",
        "Creado por",
        "Fecha/hora creación",
        "Modificado por",
        "Fecha/hora modificación",
        "Tiene PDF",
    ]
    bold = Font(bold=True)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.alignment = Alignment(vertical="center", wrap_text=True)

    r = 2
    for d in rows:
        creador = ""
        if d.created_by:
            creador = (d.created_by.nombre_completo or d.created_by.username or "").strip()
        modif = ""
        if d.updated_by:
            modif = (d.updated_by.nombre_completo or d.updated_by.username or "").strip()
        ws.cell(row=r, column=1, value=d.tipo)
        ws.cell(row=r, column=2, value=d.codigo)
        ws.cell(row=r, column=3, value=d.titulo)
        ws.cell(row=r, column=4, value=d.revision or "")
        ws.cell(row=r, column=5, value=d.fecha_creacion_doc.isoformat() if d.fecha_creacion_doc else "")
        ws.cell(row=r, column=6, value=d.fecha_ultima_revision.isoformat() if d.fecha_ultima_revision else "")
        ws.cell(row=r, column=7, value=d.responsable_elaboracion or "")
        ws.cell(row=r, column=8, value=d.responsable_aprobacion or "")
        ws.cell(row=r, column=9, value=ESTADO_LABELS.get(d.estado, d.estado))
        ws.cell(row=r, column=10, value=d.observaciones or "")
        ws.cell(row=r, column=11, value=creador)
        ws.cell(row=r, column=12, value=d.created_at.isoformat() if d.created_at else "")
        ws.cell(row=r, column=13, value=modif)
        ws.cell(row=r, column=14, value=d.updated_at.isoformat() if d.updated_at else "")
        ws.cell(row=r, column=15, value="Sí" if d.archivo_path else "No")
        for c in range(1, 16):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    for col in range(1, 16):
        ws.column_dimensions[get_column_letter(col)].width = min(48, 14 + (6 if col == 3 else 0))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def counts_by_tipo() -> dict[str, int]:
    result: dict[str, int] = {t: 0 for t in TIPOS_DOCUMENTO}
    rows = db.session.execute(
        select(SgiDocumento.tipo, func.count())
        .where(SgiDocumento.deleted_at.is_(None))
        .group_by(SgiDocumento.tipo)
    ).all()
    for tipo, cnt in rows:
        if tipo in result:
            result[tipo] = int(cnt or 0)
    return result
