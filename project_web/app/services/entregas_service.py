from __future__ import annotations

import json
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from sqlalchemy import and_, func, not_, or_, select
from sqlalchemy.orm import selectinload

from app.extensions import db
from app.constants import ENTREGAS_MARCA_PRODUCTO_TERMINADO_TRAZA, ENTREGAS_STOCK_CATEGORIA
from app.models import Entrega, EntregaEvento, User
from app.services import operational_informed_stock as informed_stock
from app.services import stock_service
from app.utils.hipoclorito_producto import nombre_ledger_canonico_hipoclorito
from app.utils.datetime_operacion import now_operacion_naive_local

_DIAS_ES = ("Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo")

# Ventana de KPIs en gestión web: últimos N días corridos hacia atrás desde "ahora" (APP_TIMEZONE).
ENTREGAS_KPI_ROLLING_DAYS_DEFAULT = 30


def _fmt_cantidad_entrega_display(q: float) -> str:
    """Miles con punto; decimales con coma (si aplica)."""
    v = float(q or 0)
    if v <= 0:
        return "0"
    if abs(v - round(v)) < 1e-6:
        n = int(round(v))
        return f"{n:,}".replace(",", ".")
    neg = v < 0
    v = abs(v)
    whole = int(v)
    frac_cents = int(round((v - whole) * 100))
    if frac_cents >= 100:
        whole += 1
        frac_cents = 0
    w = f"{whole:,}".replace(",", ".")
    if neg:
        w = "-" + w
    if frac_cents:
        frac_s = f"{frac_cents:02d}".rstrip("0")
        return f"{w},{frac_s}"
    return w


def cantidad_programada_operativa(e: Entrega) -> float:
    return float(e.cantidad_programada if e.cantidad_programada is not None else (e.cantidad or 0))


def cantidad_cargada_operativa(e: Entrega) -> float:
    return float(e.cantidad_real_cargada if e.cantidad_real_cargada is not None else cantidad_programada_operativa(e))


def cantidad_entregada_operativa(e: Entrega) -> float:
    return float(e.cantidad_real_entregada if e.cantidad_real_entregada is not None else cantidad_programada_operativa(e))


def _cantidad_programada_sql():
    return func.coalesce(Entrega.cantidad_programada, Entrega.cantidad)


def cantidad_cargada_sql():
    return func.coalesce(Entrega.cantidad_real_cargada, Entrega.cantidad_programada, Entrega.cantidad)


def cantidad_entregada_sql():
    return func.coalesce(Entrega.cantidad_real_entregada, Entrega.cantidad_programada, Entrega.cantidad)


def validate_cantidad_real(raw: float | int | str | None, label: str) -> float:
    try:
        qty = float(str(raw if raw is not None else "").replace(",", ".").strip() or 0)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} debe ser numérica.") from exc
    if qty != qty:
        raise ValueError(f"{label} debe ser numérica.")
    if qty < 0:
        raise ValueError(f"{label} no puede ser negativa.")
    return qty


def cantidad_real_warning(programada: float, real: float) -> str | None:
    if programada <= 0:
        return None
    diff_ratio = abs(float(real) - float(programada)) / float(programada)
    if diff_ratio > 0.20:
        return "Atención: la cantidad real difiere más de 20% de la programada."
    return None


def entregas_kpis_rolling(dias: int = ENTREGAS_KPI_ROLLING_DAYS_DEFAULT) -> dict[str, Any]:
    """Totales de volumen y cantidad de operaciones en una ventana móvil de `dias` corridos.

    Criterio temporal: desde (ahora local de operación − `dias` días) hasta ahora, inclusive.
    Los timestamps comparados son strings ISO guardados por las acciones «Cargar» y «Entregado»
    (`cargada_at_iso`, `entregada_at_iso`), coherentes con `now_operacion_naive_local()`.

    - **Cargado**: suma de cantidad real cargada donde hay `cargada_at_iso` en la ventana.
      No incluye entregas solo programadas (`programada` sin carga).
    - **Entregado**: suma de cantidad real entregada donde `estado == "entregada"` y `entregada_at_iso` en la ventana.
      No incluye cargas sin cierre de entrega ni programación sin entregar.

    Se excluyen filas con `cantidad` <= 0. El módulo gestiona el volumen en **litros** (`cantidad` / columna «Litros»).
    """
    dias = max(1, int(dias or ENTREGAS_KPI_ROLLING_DAYS_DEFAULT))
    now = now_operacion_naive_local()
    start = now - timedelta(days=dias)
    start_s = start.isoformat(timespec="seconds")
    end_s = now.isoformat(timespec="seconds")

    qty_carga = cantidad_cargada_sql()
    q_carga = select(func.coalesce(func.sum(qty_carga), 0.0), func.count()).where(
        Entrega.cargada_at_iso.isnot(None),
        Entrega.cargada_at_iso != "",
        qty_carga > 0,
        Entrega.cargada_at_iso >= start_s,
        Entrega.cargada_at_iso <= end_s,
    )
    total_cargado, n_cargas = db.session.execute(q_carga).one()

    qty_ent = cantidad_entregada_sql()
    q_ent = select(func.coalesce(func.sum(qty_ent), 0.0), func.count()).where(
        Entrega.estado == "entregada",
        Entrega.entregada_at_iso.isnot(None),
        Entrega.entregada_at_iso != "",
        qty_ent > 0,
        Entrega.entregada_at_iso >= start_s,
        Entrega.entregada_at_iso <= end_s,
    )
    total_ent, n_ent = db.session.execute(q_ent).one()

    tc = float(total_cargado or 0)
    te = float(total_ent or 0)
    return {
        "periodo_dias": dias,
        "periodo_subtitulo": f"últimos {dias} días corridos",
        "unidad": "L",
        "unidad_larga": "litros",
        "total_cargado": tc,
        "total_entregado": te,
        "count_cargas": int(n_cargas or 0),
        "count_entregas": int(n_ent or 0),
        "total_cargado_display": _fmt_cantidad_entrega_display(tc),
        "total_entregado_display": _fmt_cantidad_entrega_display(te),
        "ventana_inicio_iso": start_s,
        "ventana_fin_iso": end_s,
    }


def dia_semana_es(fecha_hora: datetime) -> str:
    return _DIAS_ES[int(fecha_hora.weekday())]


def append_evento(
    entrega_id: int,
    tipo: str,
    at_iso: str,
    actor: User | None,
    actor_display: str,
    detalle: dict[str, Any] | None = None,
) -> None:
    db.session.add(
        EntregaEvento(
            entrega_id=int(entrega_id),
            tipo=(tipo or "").strip()[:32],
            at_iso=at_iso,
            actor_user_id=int(actor.id) if actor else None,
            actor_display=(actor_display or "").strip() or "sistema",
            detalle=json.dumps(detalle, ensure_ascii=False) if detalle else None,
        )
    )


def puede_editar_campos_completos(entrega: Entrega) -> bool:
    return str(entrega.estado or "") == "programada"


def puede_editar_logistica_tras_carga(entrega: Entrega) -> bool:
    return str(entrega.estado or "") == "cargada"


def puede_marcar_cargada(entrega: Entrega) -> bool:
    return str(entrega.estado or "") == "programada" and entrega.consumo_stock_id is None


def puede_marcar_entregada(entrega: Entrega) -> bool:
    if str(entrega.estado or "") == "entregada":
        return False
    if stock_service.producto_entrega_es_stock_hipoclorito(str(entrega.producto or "")):
        return str(entrega.estado or "") == "cargada"
    return str(entrega.estado or "") in ("programada", "cargada")


def ejecutar_cargada(entrega: Entrega, actor: User | None, ahora: datetime, cantidad_real: float | None = None) -> None:
    if not puede_marcar_cargada(entrega):
        raise ValueError("Esta entrega no admite la acción «Cargar».")
    op_name = _actor_operador(actor)
    iso = ahora.isoformat(timespec="seconds")
    consumo_id: int | None = None
    programada = cantidad_programada_operativa(entrega)
    qty_real = validate_cantidad_real(programada if cantidad_real is None else cantidad_real, "La cantidad a cargar")
    if stock_service.producto_entrega_es_stock_hipoclorito(str(entrega.producto or "")):
        cat = (entrega.stock_categoria or ENTREGAS_STOCK_CATEGORIA).strip()
        if cat != ENTREGAS_STOCK_CATEGORIA:
            raise ValueError("La categoría de stock de la entrega debe ser producto terminado.")
        marca = (ENTREGAS_MARCA_PRODUCTO_TERMINADO_TRAZA or "").strip()
        if not marca:
            raise ValueError("Marca de trazabilidad QDV no configurada (constante ENTREGAS_MARCA_PRODUCTO_TERMINADO_TRAZA).")
        entrega.stock_marca = marca
        if qty_real > 0:
            informed_stock.raise_if_carga_qty_exceeds_instant(qty_real)
            obs = f"Entrega #{entrega.id} · carga en camión"
            # Un solo nombre en el ledger aunque `Entrega.producto` use el alias comercial del PT.
            prod_ledger = nombre_ledger_canonico_hipoclorito()
            rec = stock_service.add_consumo_stock_record(
                cat,
                prod_ledger,
                marca,
                qty_real,
                op_name,
                observaciones=obs,
                equipo_id=int(entrega.stock_equipo_id) if entrega.stock_equipo_id else None,
                fecha_hora=ahora,
                skip_ledger_availability_check=True,
                ingreso_stock_id=None,
            )
            db.session.flush()
            consumo_id = int(rec.id)
    entrega.estado = "cargada"
    entrega.cargada_at_iso = iso
    entrega.cargada_by_user_id = int(actor.id) if actor else None
    entrega.cantidad_programada = programada
    entrega.cantidad_real_cargada = qty_real
    entrega.updated_at_iso = iso
    if consumo_id is not None:
        entrega.consumo_stock_id = consumo_id
    detalle_carga: dict[str, Any] = {
        "cantidad_programada": programada,
        "cantidad_real_cargada": qty_real,
        "diferencia_litros": qty_real - programada,
        "consumo_stock_id": consumo_id,
        "operador_stock": op_name,
    }
    if consumo_id is not None:
        detalle_carga["producto_entrega"] = str(entrega.producto or "").strip() or None
        detalle_carga["producto_ledger"] = nombre_ledger_canonico_hipoclorito()
    append_evento(
        int(entrega.id),
        "cargada",
        iso,
        actor,
        _actor_display(actor),
        detalle_carga,
    )


def ejecutar_entregada(entrega: Entrega, actor: User | None, ahora: datetime, cantidad_real: float | None = None) -> None:
    if not puede_marcar_entregada(entrega):
        raise ValueError("Esta entrega no admite la acción «Entregado».")
    iso = ahora.isoformat(timespec="seconds")
    lugar = (entrega.lugar_entrega or "").strip()
    chof = _actor_display(actor)
    programada = cantidad_programada_operativa(entrega)
    fallback = cantidad_cargada_operativa(entrega) if str(entrega.estado or "") == "cargada" else programada
    qty_real = validate_cantidad_real(fallback if cantidad_real is None else cantidad_real, "La cantidad a entregar")
    entrega.estado = "entregada"
    entrega.entregada_at_iso = iso
    entrega.entregada_by_user_id = int(actor.id) if actor else None
    entrega.entregada_chofer_nombre = chof
    entrega.entregada_lugar = lugar
    entrega.entregada_dia_semana = dia_semana_es(ahora)
    entrega.cantidad_programada = programada
    entrega.cantidad_real_entregada = qty_real
    entrega.updated_at_iso = iso
    append_evento(
        int(entrega.id),
        "entregada",
        iso,
        actor,
        chof,
        {
            "lugar_entrega": lugar,
            "cantidad_programada": programada,
            "cantidad_real_entregada": qty_real,
            "diferencia_litros": qty_real - programada,
            "fecha": ahora.strftime("%Y-%m-%d"),
            "hora": ahora.strftime("%H:%M"),
            "dia_semana": entrega.entregada_dia_semana,
            "responsable": chof,
        },
    )


def _actor_display(u: User | None) -> str:
    if u is None:
        return "sistema"
    full = (getattr(u, "nombre_completo", None) or "").strip()
    if full:
        return full
    return (u.username or "").strip() or "usuario"


def _actor_operador(u: User | None) -> str:
    return _actor_display(u)


def _entrega_options_query():
    return (
        selectinload(Entrega.producto_terminado),
        selectinload(Entrega.cliente_row),
        selectinload(Entrega.lugar_row),
        selectinload(Entrega.chofer_row),
    )


def _lunes_y_domingo_semana_conteniendo(d: date) -> tuple[date, date]:
    """Lunes = inicio de semana (ISO); domingo = lunes + 6 días."""
    lunes = d - timedelta(days=int(d.weekday()))
    domingo = lunes + timedelta(days=6)
    return lunes, domingo


def rango_semana_operacion_actual() -> tuple[date, date]:
    """Lunes y domingo de la semana que contiene la fecha de operación (hoy)."""
    hoy = now_operacion_naive_local().date()
    return _lunes_y_domingo_semana_conteniendo(hoy)


def get_entregas_visibles_semana_actual() -> list[Entrega]:
    """Vista principal: desde el lunes actual en fecha prevista + pendientes de semanas previas.

    Criterio de «hoy»: `now_operacion_naive_local()` (zona APP_TIMEZONE).
    Usa columna real `fecha_prevista` (texto comparable como AAAA-MM-DD).

    - (a) `fecha_prevista` desde el lunes de la semana actual en adelante, cualquier estado.
    - (b) `fecha_prevista` anterior al lunes actual y estado distinto de entregada.
    """
    lunes, _domingo = rango_semana_operacion_actual()
    lunes_s = lunes.isoformat()
    est = func.lower(func.coalesce(Entrega.estado, ""))
    cond_desde_lunes = Entrega.fecha_prevista >= lunes_s
    cond_backlog = and_(Entrega.fecha_prevista < lunes_s, est != "entregada")
    return list(
        db.session.scalars(
            select(Entrega)
            .options(*_entrega_options_query())
            .where(or_(cond_desde_lunes, cond_backlog))
            .order_by(Entrega.fecha_prevista.asc(), Entrega.id.asc())
        ).all()
    )


def listar_entregas() -> list[Entrega]:
    return get_entregas_visibles_semana_actual()


class FiltroHistorialEntregas:
    __slots__ = ("cliente_id", "lugar_entrega_id", "chofer_entrega_id", "estado", "fecha_desde", "fecha_hasta")

    def __init__(
        self,
        *,
        cliente_id: int | None = None,
        lugar_entrega_id: int | None = None,
        chofer_entrega_id: int | None = None,
        estado: str | None = None,
        fecha_desde: date | None = None,
        fecha_hasta: date | None = None,
    ) -> None:
        self.cliente_id = cliente_id
        self.lugar_entrega_id = lugar_entrega_id
        self.chofer_entrega_id = chofer_entrega_id
        self.estado = (estado or "").strip() or None
        self.fecha_desde = fecha_desde
        self.fecha_hasta = fecha_hasta


def get_historial_entregas_filtrado(filtro: FiltroHistorialEntregas | None = None) -> list[Entrega]:
    """Listado completo de entregas con filtros opcionales (para historial / export)."""
    f = filtro or FiltroHistorialEntregas()
    q = select(Entrega).options(*_entrega_options_query()).order_by(Entrega.fecha_prevista.asc(), Entrega.id.asc())
    if f.cliente_id is not None:
        q = q.where(Entrega.cliente_id == int(f.cliente_id))
    if f.lugar_entrega_id is not None:
        q = q.where(Entrega.lugar_entrega_id == int(f.lugar_entrega_id))
    if f.chofer_entrega_id is not None:
        q = q.where(Entrega.chofer_entrega_id == int(f.chofer_entrega_id))
    if f.estado is not None:
        q = q.where(Entrega.estado == f.estado)
    if f.fecha_desde is not None and f.fecha_hasta is not None:
        q = q.where(
            Entrega.fecha_prevista >= f.fecha_desde.isoformat(),
            Entrega.fecha_prevista <= f.fecha_hasta.isoformat(),
        )
    elif f.fecha_desde is not None:
        q = q.where(Entrega.fecha_prevista >= f.fecha_desde.isoformat())
    elif f.fecha_hasta is not None:
        q = q.where(Entrega.fecha_prevista <= f.fecha_hasta.isoformat())
    return list(db.session.scalars(q).all())


def entregas_estados_para_filtro() -> list[str]:
    rows = db.session.scalars(select(Entrega.estado).distinct().order_by(Entrega.estado.asc())).all()
    return [str(x) for x in rows if x is not None and str(x).strip() != ""]


def _entrega_fila_excel(e: Entrega) -> list[Any]:
    cli = e.cliente_row.nombre if e.cliente_row else e.cliente
    lug = e.lugar_row.nombre if e.lugar_row else e.lugar_entrega
    prod = e.producto_terminado.nombre if e.producto_terminado else e.producto
    ch = e.chofer_row.nombre if e.chofer_row else (e.chofer_previsto or "")
    return [
        e.id,
        e.fecha_prevista,
        cli,
        lug,
        prod,
        cantidad_programada_operativa(e),
        e.cantidad_real_cargada,
        e.cantidad_real_entregada,
        (e.unidad or "").strip() or "L",
        ch,
        e.estado,
        e.cargada_at_iso,
        e.entregada_at_iso,
        (e.observaciones or "").strip(),
    ]


def _safe_sheet_title(name: str) -> str:
    s = re.sub(r"[" + re.escape("[]:*?/\\") + r"]", "_", (name or "").strip())[:31]
    return s or "Hoja"


def _autosize_entregas_ws(ws, ncols: int, nrows: int) -> None:
    from openpyxl.utils import get_column_letter

    if nrows < 1:
        return
    ws.freeze_panes = "A2"
    last_col = get_column_letter(ncols)
    ws.auto_filter.ref = f"A1:{last_col}{nrows}"
    for c in range(1, ncols + 1):
        col_letter = get_column_letter(c)
        maxlen = 10
        for r in range(1, min(nrows, 500) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                maxlen = max(maxlen, min(len(str(v)), 60))
        ws.column_dimensions[col_letter].width = min(maxlen + 2, 55)


def exportar_historial_entregas_excel(filtro: FiltroHistorialEntregas | None = None) -> BytesIO:
    """Genera .xlsx con columnas legibles; respeta los mismos filtros que el historial."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    rows_ent = get_historial_entregas_filtrado(filtro)
    headers = [
        "ID",
        "Fecha prevista",
        "Cliente",
        "Lugar de entrega",
        "Producto",
        "Cantidad programada",
        "Cantidad real cargada",
        "Cantidad real entregada",
        "Unidad",
        "Chofer",
        "Estado",
        "Fecha/hora carga",
        "Fecha/hora entrega",
        "Observaciones",
    ]
    data_rows = [_entrega_fila_excel(e) for e in rows_ent]

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = _safe_sheet_title("Historial entregas")
    bold = Font(bold=True)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r_i, row in enumerate(data_rows, start=2):
        for c_i, val in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    _autosize_entregas_ws(ws, len(headers), 1 + len(data_rows))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def entrega_to_api_dict(e: Entrega) -> dict[str, Any]:
    """Forma estable para API / clientes offline (mismos datos que ve gestión web)."""
    cr = e.cliente_row
    lr = e.lugar_row
    pt = e.producto_terminado
    ch = e.chofer_row

    def _opt_int(v: int | None) -> int | None:
        return int(v) if v is not None else None

    return {
        "id": int(e.id),
        "estado": str(e.estado or ""),
        "cliente": str(e.cliente or ""),
        "lugar_entrega": str(e.lugar_entrega or ""),
        "producto": str(e.producto or ""),
        "cantidad": cantidad_programada_operativa(e),
        "cantidad_programada": cantidad_programada_operativa(e),
        "cantidad_real_cargada": float(e.cantidad_real_cargada) if e.cantidad_real_cargada is not None else None,
        "cantidad_real_entregada": float(e.cantidad_real_entregada) if e.cantidad_real_entregada is not None else None,
        "cantidad_operativa_cargada": cantidad_cargada_operativa(e),
        "cantidad_operativa_entregada": cantidad_entregada_operativa(e),
        "unidad": ((e.unidad or "").strip() or None),
        "fecha_prevista": str(e.fecha_prevista or ""),
        "observaciones": ((e.observaciones or "").strip() or None),
        "chofer_previsto": ((e.chofer_previsto or "").strip() or None),
        "cliente_id": _opt_int(e.cliente_id),
        "lugar_entrega_id": _opt_int(e.lugar_entrega_id),
        "producto_terminado_id": _opt_int(e.producto_terminado_id),
        "chofer_entrega_id": _opt_int(e.chofer_entrega_id),
        "catalogo": {
            "cliente_nombre": (cr.nombre if cr else None),
            "lugar_nombre": (lr.nombre if lr else None),
            "producto_terminado_nombre": (pt.nombre if pt else None),
            "chofer_nombre": (ch.nombre if ch else None),
        },
        "created_at_iso": e.created_at_iso,
        "updated_at_iso": e.updated_at_iso,
        "created_by_user_id": _opt_int(e.created_by_user_id),
        "cargada_at_iso": e.cargada_at_iso,
        "cargada_by_user_id": _opt_int(e.cargada_by_user_id),
        "consumo_stock_id": _opt_int(e.consumo_stock_id),
        "stock_categoria": e.stock_categoria,
        "stock_marca": e.stock_marca,
        "stock_equipo_id": _opt_int(e.stock_equipo_id),
        "entregada_at_iso": e.entregada_at_iso,
        "entregada_by_user_id": _opt_int(e.entregada_by_user_id),
        "entregada_chofer_nombre": e.entregada_chofer_nombre,
        "entregada_lugar": e.entregada_lugar,
        "entregada_dia_semana": e.entregada_dia_semana,
    }
