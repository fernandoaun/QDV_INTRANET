from __future__ import annotations

from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from flask import current_app
from sqlalchemy import Select, func, select
from sqlalchemy.orm import joinedload
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from app.extensions import db
from app.models import SectorVencimiento, Vencimiento, VencimientoHistorial
from app.services.deadline_alert_email_service import normalize_validate_email
from app.services.upload_paths import uploads_workspace_root
from app.utils.datetime_operacion import now_operacion_naive_local

ESTADO_VIGENTE = "vigente"
ESTADO_PROXIMO = "proximo_vencer"
ESTADO_VENCIDO = "vencido"
ESTADO_RENOVADO = "renovado"
ESTADO_INACTIVO = "inactivo"

ESTADO_LABELS: dict[str, str] = {
    ESTADO_VIGENTE: "Vigente",
    ESTADO_PROXIMO: "Próximo a vencer",
    ESTADO_VENCIDO: "Vencido",
    ESTADO_RENOVADO: "Renovado",
    ESTADO_INACTIVO: "Inactivo",
}

ACCION_ALTA = "alta"
ACCION_EDICION = "edicion"
ACCION_RENOVACION = "renovacion"
ACCION_CAMBIO_ESTADO = "cambio_estado"
ACCION_ARCHIVO = "archivo_adjunto"
ACCION_BAJA = "baja"
ACCION_MAIL_ENVIADO = "mail_enviado"
ACCION_MAIL_ERROR = "mail_error"


def _today_local() -> date:
    return now_operacion_naive_local().date()


def dias_restantes(fecha_vencimiento: date, today: date | None = None) -> int:
    t = today or _today_local()
    return (fecha_vencimiento - t).days


def estado_visual_row(v: Vencimiento, today: date | None = None) -> str:
    if not v.activo or v.estado == ESTADO_INACTIVO:
        return "venc-row-inactivo"
    if v.estado == ESTADO_RENOVADO:
        return "venc-row-renovado"
    t = today or _today_local()
    d = dias_restantes(v.fecha_vencimiento, t)
    if d < 0 or v.estado == ESTADO_VENCIDO:
        return "venc-row-vencido"
    if d <= 30 or v.estado == ESTADO_PROXIMO:
        return "venc-row-proximo"
    return "venc-row-vigente"


def _derived_estado_active(v: Vencimiento, today: date) -> str:
    d = dias_restantes(v.fecha_vencimiento, today)
    if d < 0:
        return ESTADO_VENCIDO
    if d <= 30:
        return ESTADO_PROXIMO
    return ESTADO_VIGENTE


def sync_derived_estado(v: Vencimiento, today: date | None = None) -> bool:
    t = today or _today_local()
    if not v.activo:
        return False
    if v.estado in (ESTADO_RENOVADO, ESTADO_INACTIVO):
        return False
    want = _derived_estado_active(v, t)
    if v.estado != want:
        v.estado = want
        return True
    return False


def append_historial(vencimiento_id: int, usuario_label: str, accion: str, detalle: str) -> None:
    db.session.add(
        VencimientoHistorial(
            vencimiento_id=int(vencimiento_id),
            fecha=datetime.now(timezone.utc),
            usuario=(usuario_label or "").strip()[:256],
            accion=(accion or "").strip()[:64],
            detalle=(detalle or "").strip()[:8000],
        )
    )


def list_sectores(*, solo_activos: bool = False) -> list[SectorVencimiento]:
    q = select(SectorVencimiento).order_by(SectorVencimiento.nombre)
    if solo_activos:
        q = q.where(SectorVencimiento.activo.is_(True))
    return list(db.session.scalars(q).all())


def get_sector(sid: int) -> SectorVencimiento | None:
    return db.session.get(SectorVencimiento, int(sid))


def create_sector(nombre: str, descripcion: str, user_id: int | None) -> tuple[bool, str]:
    n = (nombre or "").strip()
    if len(n) < 2:
        return False, "El nombre del sector debe tener al menos 2 caracteres."
    exists = db.session.scalar(
        select(func.count()).select_from(SectorVencimiento).where(func.lower(SectorVencimiento.nombre) == n.lower())
    )
    if int(exists or 0) > 0:
        return False, "Ya existe un sector con ese nombre."
    row = SectorVencimiento(nombre=n, descripcion=(descripcion or "").strip()[:512], activo=True, created_by_id=user_id)
    db.session.add(row)
    db.session.commit()
    return True, "Sector creado."


def update_sector(sid: int, nombre: str, descripcion: str, activo: bool) -> tuple[bool, str]:
    row = get_sector(sid)
    if row is None:
        return False, "Sector no encontrado."
    n = (nombre or "").strip()
    if len(n) < 2:
        return False, "El nombre del sector debe tener al menos 2 caracteres."
    dup = db.session.scalar(
        select(SectorVencimiento.id).where(
            func.lower(SectorVencimiento.nombre) == n.lower(),
            SectorVencimiento.id != int(sid),
        )
    )
    if dup is not None:
        return False, "Ya existe otro sector con ese nombre."
    row.nombre = n
    row.descripcion = (descripcion or "").strip()[:512]
    row.activo = bool(activo)
    db.session.commit()
    return True, "Sector actualizado."


def parse_iso_date(raw: str | None) -> date | None:
    s = (raw or "").strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def filter_args_from_request(values: Any) -> dict[str, Any]:
    return {
        "sector_id": values.get("sector_id"),
        "estado": values.get("estado"),
        "responsable": values.get("responsable"),
        "fecha_desde": values.get("fecha_desde"),
        "fecha_hasta": values.get("fecha_hasta"),
        "proximos": values.get("proximos"),
        "vencidos": values.get("vencidos"),
        "solo_activos": values.get("solo_activos", "1"),
    }


def build_filtered_query(args: dict[str, Any]) -> Select[Any]:
    sector_id = args.get("sector_id")
    estado = (args.get("estado") or "").strip()
    responsable = (args.get("responsable") or "").strip()
    fd = parse_iso_date(args.get("fecha_desde"))
    fh = parse_iso_date(args.get("fecha_hasta"))
    proximos = str(args.get("proximos") or "").strip().lower() in ("1", "true", "yes", "on")
    vencidos = str(args.get("vencidos") or "").strip().lower() in ("1", "true", "yes", "on")
    solo_activos = str(args.get("solo_activos") or "1").strip().lower() not in ("0", "false", "no")

    q = select(Vencimiento).options(joinedload(Vencimiento.sector)).order_by(Vencimiento.fecha_vencimiento, Vencimiento.id)

    if solo_activos:
        q = q.where(Vencimiento.activo.is_(True))

    if sector_id:
        try:
            q = q.where(Vencimiento.sector_id == int(sector_id))
        except (TypeError, ValueError):
            pass

    if estado and estado in ESTADO_LABELS:
        q = q.where(Vencimiento.estado == estado)

    if responsable:
        like = f"%{responsable}%"
        q = q.where(Vencimiento.responsable.ilike(like))

    if fd is not None:
        q = q.where(Vencimiento.fecha_vencimiento >= fd)
    if fh is not None:
        q = q.where(Vencimiento.fecha_vencimiento <= fh)

    t = _today_local()
    if proximos and not vencidos:
        q = q.where(
            Vencimiento.fecha_vencimiento >= t,
            Vencimiento.fecha_vencimiento <= date.fromordinal(t.toordinal() + 30),
        )
    elif vencidos and not proximos:
        q = q.where(Vencimiento.fecha_vencimiento < t)

    return q


def fetch_list_sync_estados(args: dict[str, Any]) -> list[Vencimiento]:
    q = build_filtered_query(args)
    rows = list(db.session.scalars(q).unique().all())
    t = _today_local()
    changed = False
    for v in rows:
        if sync_derived_estado(v, t):
            changed = True
    if changed:
        db.session.commit()
        for v in rows:
            db.session.refresh(v)
    return rows


def get_vencimiento(vid: int) -> Vencimiento | None:
    return db.session.scalar(
        select(Vencimiento).options(joinedload(Vencimiento.sector)).where(Vencimiento.id == int(vid)).limit(1)
    )


def historial_for(vid: int) -> list[VencimientoHistorial]:
    return list(
        db.session.scalars(
            select(VencimientoHistorial)
            .where(VencimientoHistorial.vencimiento_id == int(vid))
            .order_by(VencimientoHistorial.fecha.desc(), VencimientoHistorial.id.desc())
        ).all()
    )


def _validate_email_aviso(raw: str | None) -> tuple[str | None, str | None]:
    s = (raw or "").strip()
    if not s:
        return None, "El correo de aviso es obligatorio."
    norm = normalize_validate_email(s)
    if norm is None:
        return None, "El correo de aviso no es válido."
    return norm, None


def create_vencimiento(form: dict[str, Any], user_id: int | None, actor_label: str) -> tuple[Vencimiento | None, str | None]:
    sector_id = int(form.get("sector_id") or 0)
    nombre = (form.get("nombre") or "").strip()
    if not nombre:
        return None, "El nombre es obligatorio."
    sec = get_sector(sector_id)
    if sec is None or not sec.activo:
        return None, "Sector inválido o inactivo."
    fd = parse_iso_date(form.get("fecha_vencimiento"))
    if fd is None:
        return None, "La fecha de vencimiento es obligatoria."
    em_raw = form.get("email_aviso")
    em, err = _validate_email_aviso(em_raw)
    if err:
        return None, err

    now = datetime.now(timezone.utc)
    row = Vencimiento(
        sector_id=sector_id,
        nombre=nombre[:256],
        descripcion=(form.get("descripcion") or "").strip()[:4000],
        fecha_vencimiento=fd,
        responsable=(form.get("responsable") or "").strip()[:256],
        email_aviso=(em or "")[:256],
        observaciones=(form.get("observaciones") or "").strip()[:4000],
        estado=ESTADO_VIGENTE,
        activo=True,
        aviso_30_dias_enviado=False,
        fecha_aviso_30_dias=None,
        created_at=now,
        updated_at=now,
        created_by_id=user_id,
        updated_by_id=user_id,
    )
    sync_derived_estado(row)
    db.session.add(row)
    db.session.flush()
    append_historial(row.id, actor_label, ACCION_ALTA, f"Alta: {nombre} · sector={sec.nombre} · vence {fd.isoformat()}")
    db.session.commit()
    db.session.refresh(row)
    return row, None


def update_vencimiento(vid: int, form: dict[str, Any], user_id: int | None, actor_label: str) -> tuple[bool, str]:
    v = db.session.get(Vencimiento, int(vid))
    if v is None:
        return False, "Registro no encontrado."
    prev_fecha = v.fecha_vencimiento
    sector_id = int(form.get("sector_id") or v.sector_id)
    sec = get_sector(sector_id)
    if sec is None:
        return False, "Sector inválido."
    nombre = (form.get("nombre") or "").strip()
    if not nombre:
        return False, "El nombre es obligatorio."
    fd = parse_iso_date(form.get("fecha_vencimiento"))
    if fd is None:
        return False, "La fecha de vencimiento es obligatoria."
    em_raw = form.get("email_aviso")
    em, err = _validate_email_aviso(em_raw)
    if err:
        return False, err

    prev_email_lower = (v.email_aviso or "").strip().lower()
    bits: list[str] = []
    if v.sector_id != sector_id:
        bits.append(f"sector_id {v.sector_id}->{sector_id}")
    if v.nombre != nombre:
        bits.append("nombre")
    if v.descripcion != (form.get("descripcion") or "").strip():
        bits.append("descripción")
    if v.fecha_vencimiento != fd:
        bits.append(f"fecha_vencimiento {prev_fecha}->{fd}")
    if v.responsable != (form.get("responsable") or "").strip():
        bits.append("responsable")
    if (v.email_aviso or "").lower() != (em or "").lower():
        bits.append("email_aviso")
    if v.observaciones != (form.get("observaciones") or "").strip():
        bits.append("observaciones")

    v.sector_id = sector_id
    v.nombre = nombre[:256]
    v.descripcion = (form.get("descripcion") or "").strip()[:4000]
    v.fecha_vencimiento = fd
    v.responsable = (form.get("responsable") or "").strip()[:256]
    v.email_aviso = (em or "")[:256]
    v.observaciones = (form.get("observaciones") or "").strip()[:4000]
    v.updated_at = datetime.now(timezone.utc)
    v.updated_by_id = user_id

    if prev_fecha != fd or prev_email_lower != (em or "").strip().lower():
        v.aviso_30_dias_enviado = False
        v.fecha_aviso_30_dias = None
        bits.append("aviso_mail_reset_por_cambio")

    sync_derived_estado(v)
    detalle = "Cambios: " + (", ".join(bits) if bits else "(sin cambios registrados)")
    append_historial(v.id, actor_label, ACCION_EDICION, detalle)
    db.session.commit()
    return True, "Vencimiento actualizado."


def deactivate_vencimiento(vid: int, user_id: int | None, actor_label: str) -> tuple[bool, str]:
    v = db.session.get(Vencimiento, int(vid))
    if v is None:
        return False, "Registro no encontrado."
    if not v.activo:
        return False, "El registro ya estaba inactivo."
    v.activo = False
    v.estado = ESTADO_INACTIVO
    v.updated_at = datetime.now(timezone.utc)
    v.updated_by_id = user_id
    append_historial(v.id, actor_label, ACCION_BAJA, "Desactivación / baja lógica.")
    db.session.commit()
    return True, "Registro desactivado."


def renew_vencimiento(vid: int, nueva_fecha: date, user_id: int | None, actor_label: str) -> tuple[Vencimiento | None, str]:
    old = db.session.get(Vencimiento, int(vid))
    if old is None or not old.activo:
        return None, "Registro no encontrado o ya inactivo."
    now = datetime.now(timezone.utc)

    old.estado = ESTADO_RENOVADO
    old.activo = False
    old.updated_at = now
    old.updated_by_id = user_id
    append_historial(
        old.id,
        actor_label,
        ACCION_RENOVACION,
        f"Renovación: cerrado como renovado. Nueva fecha operativa en registro hijo: {nueva_fecha.isoformat()}",
    )

    nuevo = Vencimiento(
        sector_id=old.sector_id,
        nombre=old.nombre,
        descripcion=old.descripcion,
        fecha_vencimiento=nueva_fecha,
        responsable=old.responsable,
        email_aviso=old.email_aviso,
        observaciones=old.observaciones,
        archivo_path=None,
        estado=ESTADO_VIGENTE,
        activo=True,
        aviso_30_dias_enviado=False,
        fecha_aviso_30_dias=None,
        continuacion_de_id=old.id,
        created_at=now,
        updated_at=now,
        created_by_id=user_id,
        updated_by_id=user_id,
    )
    sync_derived_estado(nuevo)
    db.session.add(nuevo)
    db.session.flush()
    append_historial(
        nuevo.id,
        actor_label,
        ACCION_ALTA,
        f"Alta por renovación desde vencimiento id={old.id} · fecha {nueva_fecha.isoformat()}",
    )
    db.session.commit()
    db.session.refresh(nuevo)
    return nuevo, "Renovación registrada."


def _upload_max_bytes() -> int:
    return int(current_app.config.get("MAINTENANCE_ATTACHMENT_MAX_BYTES") or (12 * 1024 * 1024))


def save_attachment(vid: int, storage: FileStorage, user_id: int | None, actor_label: str) -> tuple[bool, str]:
    v = db.session.get(Vencimiento, int(vid))
    if v is None or not v.activo:
        return False, "Registro no encontrado o inactivo."
    if not storage or not (storage.filename or "").strip():
        return False, "No se seleccionó ningún archivo."
    data = storage.read()
    size = len(data)
    if size > _upload_max_bytes():
        return False, "El archivo supera el tamaño máximo permitido."
    fn = secure_filename(storage.filename or "adjunto")
    if not fn:
        return False, "Nombre de archivo inválido."

    base = uploads_workspace_root() / "vencimientos" / str(v.id)
    base.mkdir(parents=True, exist_ok=True)
    dest = base / fn
    dest.write_bytes(data)

    rel = Path("vencimientos") / str(v.id) / fn
    v.archivo_path = rel.as_posix()
    v.updated_at = datetime.now(timezone.utc)
    v.updated_by_id = user_id
    append_historial(v.id, actor_label, ACCION_ARCHIVO, f"Adjunto: {fn} ({size} bytes)")
    db.session.commit()
    return True, "Archivo guardado."


def attachment_absolute_path(rel: str | None) -> Path | None:
    if not rel:
        return None
    from app.services.upload_paths import resolve_under_upload_roots

    p = resolve_under_upload_roots(Path(rel))
    return p


def build_export_xlsx(rows: list[Vencimiento]) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Vencimientos"
    headers = [
        "Sector",
        "Nombre",
        "Descripción",
        "Fecha de vencimiento",
        "Días restantes",
        "Estado",
        "Responsable",
        "Email",
        "Observaciones",
    ]
    bold = Font(bold=True)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.alignment = Alignment(vertical="center", wrap_text=True)

    t = _today_local()
    r = 2
    for v in rows:
        sec = v.sector.nombre if v.sector else ""
        dr = dias_restantes(v.fecha_vencimiento, t)
        est = ESTADO_LABELS.get(v.estado, v.estado)
        ws.cell(row=r, column=1, value=sec)
        ws.cell(row=r, column=2, value=v.nombre)
        ws.cell(row=r, column=3, value=v.descripcion or "")
        ws.cell(row=r, column=4, value=v.fecha_vencimiento.isoformat())
        ws.cell(row=r, column=5, value=dr)
        ws.cell(row=r, column=6, value=est)
        ws.cell(row=r, column=7, value=v.responsable or "")
        ws.cell(row=r, column=8, value=v.email_aviso or "")
        ws.cell(row=r, column=9, value=v.observaciones or "")
        for c in range(1, 10):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = min(48, 14 + (4 if col == 3 else 0))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def candidatos_aviso_mail() -> list[Vencimiento]:
    """Vencimientos activos, no renovados, con ventana 0..30 días y aviso no enviado."""
    t = _today_local()
    lim_sup = date.fromordinal(t.toordinal() + 30)
    q = (
        select(Vencimiento)
        .options(joinedload(Vencimiento.sector))
        .where(
            Vencimiento.activo.is_(True),
            Vencimiento.estado != ESTADO_RENOVADO,
            Vencimiento.estado != ESTADO_INACTIVO,
            Vencimiento.aviso_30_dias_enviado.is_(False),
            Vencimiento.fecha_vencimiento >= t,
            Vencimiento.fecha_vencimiento <= lim_sup,
        )
        .order_by(Vencimiento.fecha_vencimiento, Vencimiento.id)
    )
    rows = list(db.session.scalars(q).unique().all())
    out: list[Vencimiento] = []
    for v in rows:
        if normalize_validate_email(v.email_aviso):
            out.append(v)
    return out


def reset_aviso_flags(v: Vencimiento) -> None:
    v.aviso_30_dias_enviado = False
    v.fecha_aviso_30_dias = None
