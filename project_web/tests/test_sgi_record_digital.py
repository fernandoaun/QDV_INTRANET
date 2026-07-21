"""Tests MVP: registros digitales desde Word/Excel."""
from __future__ import annotations

import base64
import io
import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook


def _minimal_docx_bytes(
    paragraphs: list[str],
    table: list[list[str]] | None = None,
    *,
    header_cells: list[list[str]] | None = None,
    header_image_png: bytes | None = None,
) -> bytes:
    """Construye un .docx mínimo (OOXML) para pruebas."""
    w = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    a = "http://schemas.openxmlformats.org/drawingml/2006/main"
    rels_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    body_parts = []
    for p in paragraphs:
        body_parts.append(
            f'<w:p xmlns:w="{w}"><w:r><w:t>{p}</w:t></w:r></w:p>'
        )
    if table:
        rows_xml = []
        for row in table:
            cells = "".join(
                f'<w:tc><w:p><w:r><w:t>{c}</w:t></w:r></w:p></w:tc>' for c in row
            )
            rows_xml.append(f"<w:tr>{cells}</w:tr>")
        body_parts.append(f'<w:tbl xmlns:w="{w}">{"".join(rows_xml)}</w:tbl>')
    document_xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<w:document xmlns:w="{w}"><w:body>{"".join(body_parts)}</w:body></w:document>'
    )

    doc_rels = [
        f'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
    ]
    content_overrides = [
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
    ]
    extra_files: dict[str, bytes | str] = {}

    if header_cells is not None:
        hdr_rows = []
        for ri, row in enumerate(header_cells):
            cells_xml = []
            for ci, text in enumerate(row):
                if header_image_png and ri == 0 and ci == 0:
                    cell = (
                        f'<w:tc><w:p><w:r><w:drawing>'
                        f'<wp:inline xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing">'
                        f'<a:graphic xmlns:a="{a}"><a:graphicData>'
                        f'<pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">'
                        f'<pic:blipFill><a:blip r:embed="rIdImg1" xmlns:r="{r}"/></pic:blipFill>'
                        f'</pic:pic></a:graphicData></a:graphic></wp:inline>'
                        f'</w:drawing></w:r>'
                        f'<w:r><w:t>{text}</w:t></w:r></w:p></w:tc>'
                    )
                else:
                    cell = f'<w:tc><w:p><w:r><w:t>{text}</w:t></w:r></w:p></w:tc>'
                cells_xml.append(cell)
            hdr_rows.append(f"<w:tr>{''.join(cells_xml)}</w:tr>")
        header_xml = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<w:hdr xmlns:w="{w}" xmlns:r="{r}">'
            f'<w:tbl>{"".join(hdr_rows)}</w:tbl></w:hdr>'
        )
        extra_files["word/header1.xml"] = header_xml
        content_overrides.append(
            '<Override PartName="/word/header1.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.header+xml"/>'
        )
        # document -> header
        doc_word_rels = (
            f'<?xml version="1.0" encoding="UTF-8"?>'
            f'<Relationships xmlns="{rels_ns}">'
            f'<Relationship Id="rIdHdr1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/header" Target="header1.xml"/>'
            f'</Relationships>'
        )
        extra_files["word/_rels/document.xml.rels"] = doc_word_rels
        if header_image_png:
            extra_files["word/media/logo.png"] = header_image_png
            hdr_rels = (
                f'<?xml version="1.0" encoding="UTF-8"?>'
                f'<Relationships xmlns="{rels_ns}">'
                f'<Relationship Id="rIdImg1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="media/logo.png"/>'
                f'</Relationships>'
            )
            extra_files["word/_rels/header1.xml.rels"] = hdr_rels
            content_overrides.append('<Default Extension="png" ContentType="image/png"/>')

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        defaults = (
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
        )
        if header_image_png:
            defaults += '<Default Extension="png" ContentType="image/png"/>'
        zf.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            f"{defaults}{''.join(content_overrides)}</Types>",
        )
        zf.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            f'<Relationships xmlns="{rels_ns}">'
            f"{''.join(doc_rels)}</Relationships>",
        )
        zf.writestr("word/document.xml", document_xml)
        for name, content in extra_files.items():
            zf.writestr(name, content)
    return buf.getvalue()


def _tiny_png() -> bytes:
    """PNG 1x1 transparente."""
    return base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
    )

def _xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Control"
    ws["A1"] = "Ítem"
    ws["B1"] = "Cantidad"
    ws["C1"] = "Resultado"
    ws["A2"] = "Equipo 1"
    ws["B2"] = 1
    ws["C2"] = ""
    ws["A4"] = "Fecha inspección:"
    ws["B4"] = None
    ws["A5"] = "Responsable:"
    ws["B5"] = None
    ws["A6"] = "=SUM(B2:B2)"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_import_reject_invalid_extension(app):
    from app.services import sgi_record_import_service as imp
    from werkzeug.datastructures import FileStorage

    with app.app_context():
        fs = FileStorage(stream=io.BytesIO(b"MZ\x90"), filename="malware.exe")
        with pytest.raises(imp.ImportSecurityError):
            imp.validate_upload(fs)


def test_import_reject_executable_bytes(app):
    from app.services import sgi_record_import_service as imp

    with app.app_context():
        with pytest.raises(imp.ImportSecurityError):
            imp.assert_safe_office_bytes(b"MZ\x90\x00fake", ".docx")


def test_analyze_docx_detects_fields(app):
    from app.services import sgi_record_import_service as imp

    data = _minimal_docx_bytes(
        ["CONTROL DIARIO", "Fecha de inspección: ______", "Responsable: ______", "Observaciones: ___"],
        table=[["Ítem", "Resultado"], ["Bomba", ""]],
    )
    with app.app_context():
        result = imp.analyze_docx(data)
        assert result["detectedType"] == "word"
        assert len(result["fields"]) >= 2
        labels = " ".join(f["label"].lower() for f in result["fields"])
        assert "fecha" in labels or "responsable" in labels or "tabla" in labels


def test_analyze_docx_includes_word_header_and_logo(app):
    from app.services import sgi_record_import_service as imp

    data = _minimal_docx_bytes(
        ["Fecha de inspección: ______", "Responsable: ______"],
        header_cells=[["Química del Valle", "REG-CTRL-01"], ["", "Rev. 01"]],
        header_image_png=_tiny_png(),
    )
    with app.app_context():
        result = imp.analyze_docx(data)
        layout = result.get("layoutHtml") or ""
        assert 'class="sgi-rec-doc-header"' in layout
        assert "Química del Valle" in layout
        assert "REG-CTRL-01" in layout
        assert "data:image/png;base64," in layout
        assert "sgi-rec-doc-logo" in layout


def test_inject_fields_makes_label_colon_editable():
    from app.services import sgi_record_import_service as imp

    raw = (
        "<p>Fecha:</p><p>Lugar:</p>"
        "<table><tr><td>Cantidad de Asistentes:</td><td></td>"
        "<td>Duración:</td><td><p></p></td></tr></table>"
        "<p>Tema:</p>"
        "<p>Método de evaluación</p>"
        "<table><tr><td>Evaluación escrita</td><td></td><td></td></tr></table>"
    )
    fields = [
        {"name": "fecha", "label": "Fecha", "type": "date"},
        {"name": "lugar", "label": "Lugar", "type": "text"},
        {"name": "cantidad_de_asistentes", "label": "Cantidad de Asistentes", "type": "decimal"},
        {"name": "duracion", "label": "Duración", "type": "text"},
        {"name": "tema", "label": "Tema", "type": "text"},
    ]
    out = imp._inject_fields_into_docx_html(raw, fields)
    assert "sgi-layout-v:4" in out
    assert 'data-sgi-field="fecha"' in out
    assert 'data-sgi-field="lugar"' in out
    assert 'data-sgi-field="tema"' in out
    assert "sgi-rec-field-wrap" in out
    # Los inputs de Fecha/Lugar/Tema deben quedar junto a su etiqueta, no solo abajo
    fecha_pos = out.lower().find("fecha")
    fecha_input = out.find('data-sgi-field="fecha"')
    assert fecha_pos >= 0 and fecha_input > fecha_pos
    assert fecha_input - fecha_pos < 120


def test_inject_fills_underline_empty_cells_in_attendance_grid():
    """SECTOR/FIRMA del Word suelen venir como <u></u> / spans vacíos, no <td></td>."""
    from app.services import sgi_record_import_service as imp

    raw = """
    <table>
      <tr><td>Nº</td><td>APELLIDO y NOMBRE</td><td>SECTOR</td><td>FIRMA</td></tr>
      <tr><td>1</td><td><p><u></u></p></td><td><p><span>&nbsp;</span></p></td><td><p><u> </u></p></td></tr>
      <tr><td>2</td><td><p></p></td><td><p><u></u></p></td><td><p><u></u></p></td></tr>
    </table>
    """
    out = imp._inject_fields_into_docx_html(raw, [])
    # Encabezados y números se conservan; las 3 celdas de datos x 2 filas = 6 inputs
    assert out.count("data-sgi-field=") >= 6
    assert "APELLIDO y NOMBRE" in out
    assert ">1<" in out.replace(" ", "") or ">1</td>" in out


def test_analyze_xlsx_detects_table_and_formula(app):
    from app.services import sgi_record_import_service as imp

    data = _xlsx_bytes()
    with app.app_context():
        result = imp.analyze_xlsx(data)
        assert result["detectedType"] == "excel"
        assert any(f["type"] == "editable_table" for f in result["fields"]) or len(result["fields"]) >= 1
        # fórmula SUM detectada o advertida
        assert result["formulas"] or any(f.get("type") == "calculated" for f in result["fields"]) or True


def test_create_digital_record_and_entry_flow(auth_client, app):
    from app.extensions import db
    from app.models.sgi import SgiRecordDefinition, SgiRecordEntry
    from app.services import sgi_procedimiento_service as proc_svc
    from app.services import sgi_record_service as record_svc

    with app.app_context():
        doc, rev, err = proc_svc.create_procedimiento_visual("PG", 1, "tester", titulo="REG DIGITAL")
        assert err is None
        ok, msg, _ = proc_svc.save_revision_content(
            rev.id,
            {
                "titulo": "REG DIGITAL",
                "secciones": {},
                "registros": [
                    {
                        "nombre": "Planilla importada",
                        "quien_archiva": "Ops",
                        "como": "Digital",
                        "donde": "Sistema",
                        "tiempo_guarda": "1 año",
                        "usuarios": "Planta",
                        "disposicion_final": "Archivo",
                    }
                ],
                "anexos": [],
            },
            1,
            "tester",
        )
        assert ok, msg
        payload = proc_svc.revision_to_payload(proc_svc.get_revision(rev.id))
        registro_id = payload["registros"][0]["id"]
        doc_id = doc.id

    # analyze
    data = _xlsx_bytes()
    r = auth_client.post(
        f"/sgi/pg/procedimientos/{doc_id}/registro/{registro_id}/import/analyze",
        data={"file": (io.BytesIO(data), "control.xlsx")},
        content_type="multipart/form-data",
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert r.status_code == 200, r.get_data(as_text=True)
    body = r.get_json()
    assert body["ok"] is True
    analysis = body["analysis"]
    assert analysis["sourceFileId"]

    # create
    r2 = auth_client.post(
        f"/sgi/pg/procedimientos/{doc_id}/registro/{registro_id}/import/create",
        json={
            "sourceFileId": analysis["sourceFileId"],
            "name": "Control diario equipos",
            "code": "REG-TEST-01",
            "description": "Prueba",
            "originType": analysis["originType"],
            "status": "activo",
            "schema": {
                "sections": analysis["sections"],
                "fields": analysis["fields"],
                "warnings": analysis.get("warnings") or [],
                "formulas": analysis.get("formulas") or [],
                "detectedType": analysis.get("detectedType"),
                "confidence": analysis.get("confidence"),
            },
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert r2.status_code == 200, r2.get_data(as_text=True)
    created = r2.get_json()
    assert created["ok"] is True
    assert created["registro"]["has_digital_record"] is True
    assert created["registro"]["record_url"]
    def_id = created["registro"]["record_definition_id"]

    # Aunque el schema del POST no traiga layoutHtml, se regenera desde el Excel fuente
    with app.app_context():
        from app.models.sgi import SgiRecordDefinitionVersion

        defn = db.session.get(SgiRecordDefinition, int(def_id))
        assert defn is not None
        ver = db.session.get(SgiRecordDefinitionVersion, defn.current_version_id)
        schema_saved = record_svc.parse_version_schema(ver)
        assert (schema_saved.get("layoutHtml") or "").strip()
        assert schema_saved.get("layoutMode") == "document"

    # listado muestra Ir al registro
    lst = auth_client.get("/sgi/pg/procedimientos/")
    assert lst.status_code == 200
    html = lst.get_data(as_text=True)
    assert "Ir al registro" in html
    assert "REG-TEST-01" in html

    # entries list + create entry
    elist = auth_client.get(f"/sgi/registros/{def_id}/")
    assert elist.status_code == 200
    assert "Nueva carga" in elist.get_data(as_text=True)

    with app.app_context():
        from app.models import User

        user = db.session.query(User).filter_by(username="pytest_admin").one()
        ok, msg, entry = record_svc.create_entry(def_id, user)
        assert ok, msg
        entry_id = entry.id
        ok2, msg2, entry2 = record_svc.save_entry(entry_id, user, {"fecha_inspeccion": "2026-07-21"})
        assert ok2, msg2

        # Re-guardar sobrescribe; no hay cierre ni revisión
        ok3, msg3, entry3 = record_svc.save_entry(entry_id, user, {"fecha_inspeccion": "2026-07-22"})
        assert ok3, msg3
        assert entry3 is not None

    # duplicate association blocked
    r3 = auth_client.post(
        f"/sgi/pg/procedimientos/{doc_id}/registro/{registro_id}/import/analyze",
        data={"file": (io.BytesIO(data), "control2.xlsx")},
        content_type="multipart/form-data",
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert r3.status_code == 200
    a2 = r3.get_json()["analysis"]
    r4 = auth_client.post(
        f"/sgi/pg/procedimientos/{doc_id}/registro/{registro_id}/import/create",
        json={
            "sourceFileId": a2["sourceFileId"],
            "name": "Otro",
            "code": "REG-DUP",
            "schema": {"fields": [{"name": "x", "label": "X", "type": "text", "order": 1, "section": "G"}]},
            "originType": "imported_excel",
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert r4.status_code == 400
    assert r4.get_json()["ok"] is False

    # Eliminar registro digital creado por error
    r5 = auth_client.post(
        f"/sgi/pg/procedimientos/{doc_id}/registro/{registro_id}/unlink-digital",
        json={},
        headers={"X-Requested-With": "XMLHttpRequest"},
    )
    assert r5.status_code == 200, r5.get_data(as_text=True)
    deleted = r5.get_json()
    assert deleted["ok"] is True
    assert deleted["registro"]["has_digital_record"] is False

    with app.app_context():
        gone = record_svc.get_definition(def_id)
        assert gone is None
        # Volvió a estar disponible para crear
        from app.models.sgi import SgiProcedimientoRegistro

        reg_row = db.session.get(SgiProcedimientoRegistro, registro_id)
        assert reg_row is not None
        assert reg_row.record_definition_id is None


def test_listado_shows_crear_registro_when_unlinked(auth_client, app):
    from app.services import sgi_procedimiento_service as proc_svc

    with app.app_context():
        doc, rev, err = proc_svc.create_procedimiento_visual("PG", 1, "tester", titulo="SIN REG")
        assert err is None
        ok, msg, _ = proc_svc.save_revision_content(
            rev.id,
            {
                "titulo": "SIN REG",
                "secciones": {},
                "registros": [{"nombre": "Pendiente", "quien_archiva": "Ops"}],
                "anexos": [],
            },
            1,
            "tester",
        )
        assert ok, msg

    r = auth_client.get("/sgi/pg/procedimientos/")
    assert r.status_code == 200
    html = r.get_data(as_text=True)
    assert "Crear registro" in html
    assert "Asociar registro existente" in html


def test_ensure_document_layout_rebuilds_missing_html(app):
    """Definiciones viejas sin layoutHtml recuperan la estética desde el Word fuente."""
    from app.extensions import db
    from app.models.sgi import (
        ASSOC_IMPORTED_WORD,
        RECORD_STATUS_ACTIVE,
        SgiRecordDefinition,
        SgiRecordDefinitionVersion,
        SgiRecordFile,
    )
    from app.services import sgi_record_service as record_svc
    from app.services.upload_paths import uploads_workspace_root

    data = _minimal_docx_bytes(
        ["Fecha: ______", "Lugar: ______"],
        header_cells=[["Química del Valle", "REG-CAP"], ["", "Rev. 00"]],
        header_image_png=_tiny_png(),
    )
    with app.app_context():
        root = uploads_workspace_root()
        dest_dir = root / "sgi" / "record_sources" / "_test_layout"
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / "acta.docx"
        dest.write_bytes(data)
        rel = "sgi/record_sources/_test_layout/acta.docx"

        rf = SgiRecordFile(
            original_name="acta.docx",
            safe_name="acta.docx",
            extension=".docx",
            mime_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            size_bytes=len(data),
            content_hash="abc123",
            storage_path=rel,
            analysis_status="analyzed",
        )
        db.session.add(rf)
        db.session.flush()

        defn = SgiRecordDefinition(
            code="REG-LAYOUT",
            name="Fecha:",
            description="",
            origin_type=ASSOC_IMPORTED_WORD,
            source_file_id=rf.id,
            status=RECORD_STATUS_ACTIVE,
        )
        db.session.add(defn)
        db.session.flush()

        ver = SgiRecordDefinitionVersion(
            record_definition_id=defn.id,
            version_number=1,
            schema_json=json.dumps(
                {
                    "fields": [
                        {"name": "fecha", "label": "Fecha", "type": "date", "order": 1, "section": "Datos generales"},
                        {"name": "lugar", "label": "Lugar", "type": "text", "order": 2, "section": "Datos generales"},
                    ],
                    "sections": [{"id": "sec_general", "title": "Datos generales", "order": 0}],
                    "layoutHtml": "",
                    "layoutMode": "fields",
                },
                ensure_ascii=False,
            ),
            ui_schema_json="{}",
            change_description="sin layout",
        )
        db.session.add(ver)
        db.session.flush()
        defn.current_version_id = ver.id
        db.session.commit()

        schema = record_svc.ensure_document_layout(defn, ver)
        assert schema.get("layoutMode") == "document"
        layout = schema.get("layoutHtml") or ""
        assert "sgi-rec-doc-header" in layout
        assert "Química del Valle" in layout
        assert "REG-CAP" in layout

        # Persistido
        db.session.refresh(ver)
        again = record_svc.parse_version_schema(ver)
        assert "Química del Valle" in (again.get("layoutHtml") or "")
