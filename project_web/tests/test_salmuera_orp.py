from __future__ import annotations

from datetime import date
from io import BytesIO
import re


def _csrf(html: str) -> str:
    match = re.search(r'name="csrf_token"\s+value="([^"]+)"', html)
    assert match is not None
    return match.group(1)


def _valid_salmuera_payload(csrf_token: str, *, orp: str) -> dict[str, str]:
    return {
        "csrf_token": csrf_token,
        "action": "guardar",
        "electrolizador": "1",
        "cantidad_celdas": "2",
        "voltajes": "3.0, 3.1",
        "amperaje": "120",
        "caudal_agua_l_h": "10",
        "caudal_salmuera_l_h": "20",
        "hipo_conc": "1.2",
        "hipo_exceso_soda": "2.3",
        "sal_temp": "24.5",
        "sal_conc": "260",
        "sal_ph": "4.8",
        "soda_conc": "30",
        "declor_ph": "1.2",
        "orp": orp,
        "observaciones": "Control ORP pytest",
    }


def test_salmuera_orp_optional_negative_history_and_export(app, auth_client):
    from openpyxl import load_workbook

    from app.services.historicos_export_service import build_historicos_workbook

    page = auth_client.get("/produccion/salmuera?fecha=2026-04-30")
    assert page.status_code == 200
    html = page.get_data(as_text=True)
    assert "ORP (mV)" in html

    resp = auth_client.post(
        "/produccion/salmuera?fecha=2026-04-30",
        data=_valid_salmuera_payload(_csrf(html), orp="-123,5"),
        headers={"X-Requested-With": "XMLHttpRequest", "Accept": "application/json"},
    )
    assert resp.status_code == 200
    payload = resp.get_json()
    assert payload["ok"] is True
    assert payload["registro"]["orp"] == -123.5

    hist = auth_client.get("/produccion/salmuera/historial?desde=2026-04-30&hasta=2026-04-30")
    assert hist.status_code == 200
    hist_html = hist.get_data(as_text=True)
    assert "ORP (mV)" in hist_html
    assert "-123.5" in hist_html

    with app.app_context():
        bio, err = build_historicos_workbook(
            ["salmuera"],
            date.fromisoformat("2026-04-30"),
            date.fromisoformat("2026-04-30"),
        )
    assert err is None
    assert bio is not None
    wb = load_workbook(BytesIO(bio.getvalue()))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    orp_col = headers.index("ORP (mV)") + 1
    assert ws.cell(row=2, column=orp_col).value == -123.5


def test_salmuera_orp_empty_does_not_block_and_invalid_is_rejected(auth_client):
    page = auth_client.get("/produccion/salmuera?fecha=2026-04-30")
    assert page.status_code == 200
    csrf_token = _csrf(page.get_data(as_text=True))

    blank = auth_client.post(
        "/produccion/salmuera?fecha=2026-04-30",
        data=_valid_salmuera_payload(csrf_token, orp=""),
        headers={"X-Requested-With": "XMLHttpRequest", "Accept": "application/json"},
    )
    assert blank.status_code == 200
    blank_payload = blank.get_json()
    assert blank_payload["registro"]["orp"] is None

    page = auth_client.get("/produccion/salmuera?fecha=2026-04-30")
    invalid = auth_client.post(
        "/produccion/salmuera?fecha=2026-04-30",
        data=_valid_salmuera_payload(_csrf(page.get_data(as_text=True)), orp="abc"),
        headers={"X-Requested-With": "XMLHttpRequest", "Accept": "application/json"},
    )
    assert invalid.status_code == 400
    assert "ORP (mV) debe ser numérico" in invalid.get_json()["error"]
