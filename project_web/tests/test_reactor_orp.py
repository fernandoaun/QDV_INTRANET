from __future__ import annotations

from datetime import date
from io import BytesIO
import re


def _csrf(html: str) -> str:
    match = re.search(r'name="csrf_token"\s+value="([^"]+)"', html)
    assert match is not None
    return match.group(1)


def _valid_reactor_payload(csrf_token: str, *, orp: str) -> dict[str, str]:
    return {
        "csrf_token": csrf_token,
        "ph": "8.2",
        "temperatura": "24.5",
        "densidad": "1.2",
        "concentracion_tabla": "260",
        "exceso_naoh": "0.1",
        "exceso_na2co3": "0.2",
        "orp": orp,
        "observaciones": "Control ORP pytest",
    }


def test_reactor_orp_negative_history_and_export(app, auth_client):
    from openpyxl import load_workbook

    from app.extensions import db
    from app.models import ReactorRegistro, SalmueraRegistro
    from app.services.historicos_export_service import build_historicos_workbook

    page = auth_client.get("/produccion/reactor?fecha=2026-04-30")
    assert page.status_code == 200
    html = page.get_data(as_text=True)
    assert "ORP (mV)" in html

    salmuera_page = auth_client.get("/produccion/salmuera?fecha=2026-04-30")
    assert salmuera_page.status_code == 200
    assert "ORP (mV)" not in salmuera_page.get_data(as_text=True)

    resp = auth_client.post(
        "/produccion/reactor?fecha=2026-04-30",
        data=_valid_reactor_payload(_csrf(html), orp="-123,5"),
    )
    assert resp.status_code in (302, 303)

    with app.app_context():
        row = db.session.query(ReactorRegistro).one()
        assert row.orp == -123.5
        assert not hasattr(SalmueraRegistro, "orp")

    hist = auth_client.get("/produccion/reactor/historial?desde=2026-04-30&hasta=2026-04-30")
    assert hist.status_code == 200
    hist_html = hist.get_data(as_text=True)
    assert "ORP (mV)" in hist_html
    assert "-123.5" in hist_html

    with app.app_context():
        bio, err = build_historicos_workbook(
            ["reactor"],
            date.fromisoformat("2026-04-30"),
            date.fromisoformat("2026-04-30"),
        )
    assert err is None
    assert bio is not None
    wb = load_workbook(BytesIO(bio.getvalue()))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    orp_col = headers.index("ORP (mV)") + 1
    assert ws.cell(row=2, column=orp_col).value == -123.5


def test_reactor_orp_empty_does_not_block_and_invalid_is_rejected(app, auth_client):
    from app.extensions import db
    from app.models import ReactorRegistro

    page = auth_client.get("/produccion/reactor?fecha=2026-04-30")
    assert page.status_code == 200

    blank = auth_client.post(
        "/produccion/reactor?fecha=2026-04-30",
        data=_valid_reactor_payload(_csrf(page.get_data(as_text=True)), orp=""),
    )
    assert blank.status_code in (302, 303)
    with app.app_context():
        row = db.session.query(ReactorRegistro).one()
        assert row.orp is None

    page = auth_client.get("/produccion/reactor?fecha=2026-04-30")
    invalid = auth_client.post(
        "/produccion/reactor?fecha=2026-04-30",
        data=_valid_reactor_payload(_csrf(page.get_data(as_text=True)), orp="abc"),
    )
    assert invalid.status_code == 200
    assert "ORP (mV) debe ser numérico" in invalid.get_data(as_text=True)
    with app.app_context():
        assert db.session.query(ReactorRegistro).count() == 1
