from __future__ import annotations

from datetime import datetime
from io import BytesIO
import re


def _csrf(html: str) -> str:
    match = re.search(r'name="csrf_token"\s+value="([^"]+)"', html)
    assert match is not None
    return match.group(1)


def _seed_ingreso(app) -> int:
    from app.extensions import db
    from app.models import IngresoStock
    from app.services import stock_service

    with app.app_context():
        stock_service.save_ingreso(
            "materia_prima",
            "Soda ajuste pytest",
            "Marca pytest",
            "2027-12-31",
            "L- AJ-1",
            10.0,
            "admin",
            unidad="kg",
            fecha="2026-04-30",
            hora="08:00",
        )
        row = db.session.query(IngresoStock).filter_by(producto="Soda ajuste pytest").one()
        return int(row.id)


def test_admin_stock_ajuste_changes_stock_without_consumo(app, auth_client):
    from app.extensions import db
    from app.models import ConsumoStock, ProductoCatalogo, StockAjuste
    from app.services import stock_service

    ingreso_id = _seed_ingreso(app)

    page = auth_client.get(
        "/produccion/stock/ajustes?categoria=materia_prima&producto=Soda%20ajuste%20pytest&marca=Marca%20pytest"
    )
    assert page.status_code == 200
    html = page.get_data(as_text=True)
    assert "Ajustes administrativos de stock" in html

    neg = auth_client.post(
        "/produccion/stock/ajustes",
        data={
            "csrf_token": _csrf(html),
            "categoria": "materia_prima",
            "producto": "Soda ajuste pytest",
            "marca": "Marca pytest",
            "ingreso_stock_id": str(ingreso_id),
            "tipo": "negativo",
            "cantidad": "2",
            "motivo": "Diferencia de inventario",
            "observaciones": "Recuento físico",
        },
    )
    assert neg.status_code in (302, 303)

    with app.app_context():
        assert db.session.query(ConsumoStock).count() == 0
        ajuste = db.session.query(StockAjuste).one()
        assert ajuste.cantidad == -2
        assert stock_service.stock_actual("materia_prima", "Soda ajuste pytest", "Marca pytest") == 8.0
        assert stock_service.consumos_recientes("materia_prima", "Soda ajuste pytest") == []
        catalogo = db.session.query(ProductoCatalogo).filter_by(
            categoria="materia_prima",
            nombre_producto="Soda ajuste pytest",
        ).one()
        catalogo.stock_minimo_alerta = 9.0
        db.session.commit()
        alertas = stock_service.alertas_bajo_stock(limit=30)
        alerta = next(a for a in alertas if a["producto"] == "Soda ajuste pytest")
        assert alerta["stock_actual"] == 8.0
        assert alerta["nivel_alerta"] == "critico"

    page = auth_client.get(
        "/produccion/stock/ajustes?categoria=materia_prima&producto=Soda%20ajuste%20pytest&marca=Marca%20pytest"
    )
    pos = auth_client.post(
        "/produccion/stock/ajustes",
        data={
            "csrf_token": _csrf(page.get_data(as_text=True)),
            "categoria": "materia_prima",
            "producto": "Soda ajuste pytest",
            "marca": "Marca pytest",
            "ingreso_stock_id": str(ingreso_id),
            "tipo": "positivo",
            "cantidad": "3",
            "motivo": "Sobrante detectado",
        },
    )
    assert pos.status_code in (302, 303)
    with app.app_context():
        assert db.session.query(ConsumoStock).count() == 0
        assert db.session.query(StockAjuste).count() == 2
        assert stock_service.stock_actual("materia_prima", "Soda ajuste pytest", "Marca pytest") == 11.0


def test_stock_ajuste_negative_cannot_exceed_available(app, auth_client):
    from app.extensions import db
    from app.models import StockAjuste

    ingreso_id = _seed_ingreso(app)
    page = auth_client.get(
        "/produccion/stock/ajustes?categoria=materia_prima&producto=Soda%20ajuste%20pytest&marca=Marca%20pytest"
    )
    resp = auth_client.post(
        "/produccion/stock/ajustes",
        data={
            "csrf_token": _csrf(page.get_data(as_text=True)),
            "categoria": "materia_prima",
            "producto": "Soda ajuste pytest",
            "marca": "Marca pytest",
            "ingreso_stock_id": str(ingreso_id),
            "tipo": "negativo",
            "cantidad": "99",
            "motivo": "Error",
        },
    )
    assert resp.status_code == 200
    assert "saldo negativo" in resp.get_data(as_text=True)
    with app.app_context():
        assert db.session.query(StockAjuste).count() == 0


def test_stock_ajustes_export(app, auth_client):
    from openpyxl import load_workbook

    from app.services.historicos_export_service import build_historicos_workbook

    ingreso_id = _seed_ingreso(app)
    page = auth_client.get(
        "/produccion/stock/ajustes?categoria=materia_prima&producto=Soda%20ajuste%20pytest&marca=Marca%20pytest"
    )
    resp = auth_client.post(
        "/produccion/stock/ajustes",
        data={
            "csrf_token": _csrf(page.get_data(as_text=True)),
            "categoria": "materia_prima",
            "producto": "Soda ajuste pytest",
            "marca": "Marca pytest",
            "ingreso_stock_id": str(ingreso_id),
            "tipo": "positivo",
            "cantidad": "1.5",
            "motivo": "Ajuste gerencia",
        },
    )
    assert resp.status_code in (302, 303)

    from app.utils.datetime_operacion import now_operacion_naive_local

    with app.app_context():
        export_day = now_operacion_naive_local().date()
        bio, err = build_historicos_workbook(
            ["stock_ajustes"],
            export_day,
            export_day,
        )
    assert err is None
    assert bio is not None
    wb = load_workbook(BytesIO(bio.getvalue()))
    ws = wb.active
    assert ws.title == "Stock_ajustes"
    headers = [cell.value for cell in ws[1]]
    assert "Cantidad firmada" in headers
    assert ws.cell(row=2, column=headers.index("Cantidad firmada") + 1).value == 1.5

def test_stock_ajustes_non_admin_blocked(mant_client):
    blocked = mant_client.get("/produccion/stock/ajustes", follow_redirects=False)
    assert blocked.status_code in (302, 303)


def test_stock_existencias_can_be_queried_by_day_and_time(app, auth_client):
    from app.extensions import db
    from app.models import StockAjuste
    from app.services import stock_service

    with app.app_context():
        stock_service.save_ingreso(
            "materia_prima",
            "Soda historica pytest",
            "Marca pytest",
            "2027-12-31",
            "L-HIST-1",
            10.0,
            "admin",
            unidad="kg",
            fecha="2026-04-30",
            hora="08:00",
        )
        stock_service.add_consumo_stock_record(
            "materia_prima",
            "Soda historica pytest",
            "Marca pytest",
            3.0,
            "admin",
            fecha_hora=datetime(2026, 4, 30, 10, 0),
            skip_ledger_availability_check=True,
        )
        db.session.add(
            StockAjuste(
                categoria="materia_prima",
                producto="Soda historica pytest",
                marca="Marca pytest",
                cantidad=2.0,
                fecha="2026-04-30",
                hora="11:00",
                operador="admin",
                motivo="Ajuste pytest",
                created_at_iso="2026-04-30T11:00:00",
            )
        )
        db.session.commit()

        items_09 = stock_service.stock_consolidado_as_of("materia_prima", "2026-04-30", "09:00")
        item_09 = next(x for x in items_09 if x["producto"] == "Soda historica pytest")
        assert item_09["stock"] == 10.0

        items_1030 = stock_service.stock_consolidado_as_of("materia_prima", "2026-04-30", "10:30")
        item_1030 = next(x for x in items_1030 if x["producto"] == "Soda historica pytest")
        assert item_1030["stock"] == 7.0

        items_12 = stock_service.stock_consolidado_as_of("materia_prima", "2026-04-30", "12:00")
        item_12 = next(x for x in items_12 if x["producto"] == "Soda historica pytest")
        assert item_12["stock"] == 9.0

    resp = auth_client.get(
        "/produccion/stock/ver?categoria=materia_prima&fecha_consulta=2026-04-30&hora_consulta=10:30"
    )
    assert resp.status_code == 200
    html = resp.get_data(as_text=True)
    assert "Consulta histórica hasta 2026-04-30 10:30" in html
    assert "Soda historica pytest" in html
